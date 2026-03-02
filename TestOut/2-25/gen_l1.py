"""
L1工程特性清单生成器 — LAH.3WA.201.B 燃油供给系统
使用方法：python3 gen_l1.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
CONFIG = {
    "输出路径": "阶段2_L1工程特性清单.xlsx",
}

# ============================================================
# L1_DATA
# 列顺序: [ID, 特性分类, 工程特性名称, 目标值/要求, 原文描述, 类别, 文件来源, 章节, 页码, 备注]
# 类别: "C(关键)"=BsM-Sa/Sa/Z/SOTIF/Crash / "A(一般)"=其他
# ============================================================
L1_DATA = [
    # ===== 系统压力条件 =====
    [1, "环境条件", "油箱内部压力-汽油常规油箱", "真空50hPa / 正压100hPa", "Gasoline conventional fuel tank: Vacuum 50 hPa; positive pressure 100 hPa", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "BsM-SOTIF"],
    [2, "环境条件", "油箱内部压力-汽油增压油箱", "真空150hPa / 正压400kPa", "Gasoline pressurized fuel tank: Vacuum 150 hPa; positive pressure 400 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "BsM-SOTIF"],
    [3, "环境条件", "油箱内部压力-柴油", "真空150hPa / 正压100hPa", "Diesel: Vacuum 150 hPa; positive pressure 100 hPa", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "BsM-SOTIF"],
    [4, "环境条件", "供油温度范围-汽油", "按燃油标准蒸汽压≤15psi", "For gasoline fuels, applies at least to vapor pressures up to 15 psi", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "BsM-Sa/Z, Lfd.168"],
    [5, "环境条件", "供油温度范围-柴油", "CFPP-5K ~ 90°C", "Diesel fuels from 5 K colder than CFPP up to 90°C", "C(关键)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-Sa/Z, Lfd.188"],
    [6, "环境条件", "柴油运动粘度范围", "1.2~6.5 mm²/s @40°C", "Kinematic viscosity (at 40°C) of 1.2 to 6.5 mm²/s", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "BsM-Sa/Z, Lfd.169"],
    [7, "环境条件", "燃油箱呼吸高度补偿", "-10/+20 mm", "Height compensation of -10/+20 mm due to fuel tank breathing", "C(关键)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-SOTIF, Lfd.206"],

    # ===== FSI/MPI 供油性能 =====
    [8, "供油性能-FSI", "FSI/MPI工作电压范围", "6~16V", "Operating voltage range: 6 to 16 V", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.1", "25", "Lfd.328-332"],
    [9, "供油性能-FSI", "FSI/MPI最大工作电流", "12A", "Max operating current: 12 A", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.1", "25", ""],
    [10, "供油性能-FSI", "FSI/MPI冲击电流限制", "≤3倍工作电流, ≤100ms", "Inrush current: max 100 ms, max 3x operating current", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.1", "25", ""],
    [11, "供油性能-FSI", "FSI/MPI堵转电流", "≤3倍工作电流", "Stall current: max 3x operating current", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.1", "25", ""],
    [12, "供油性能-FSI", "FSI建压时间-EA211", "≤500ms至410kPa @11V,Qv20l/h", "EA211: t≤500ms to p 410kPa @U 11V, Qv 20l/h", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.4", "26", "BsM-Sa/Z, Lfd.340-343"],
    [13, "供油性能-FSI", "FSI建压时间-EA888", "<500ms至610kPa @6V,Qv40l/h", "EA888: t<500ms to p 610kPa @U 6V, Qv 40l/h", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.4", "26", "BsM-Sa/Z"],
    [14, "供油性能-FSI", "FSI冷启动温度", "-40°C @6V", "Cold start at -40°C, 6V for fuels up to E30/E85/E100", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.5", "27", "BsM-Sa/Z, Lfd.344-359"],
    [15, "供油性能-FSI", "FSI热启动温度", "35~70°C", "Hot start at 35-70°C", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.6", "28", "Lfd.360-375"],
    [16, "供油性能-FSI", "FSI最大系统压力", "700kPa", "Maximum pressure: 700 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.8", "29", "Lfd.378-388"],
    [17, "供油性能-FSI", "FSI限压阀关闭压力", ">600kPa", "Closing pressure: >600 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.8", "29", ""],
    [18, "供油性能-FSI", "FSI限压阀开启压力", ">620kPa @≤1l/h", "Opening pressure: >620 kPa at ≤1 l/h", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.8", "29", ""],
    [19, "供油性能-FSI", "FSI流量临时衰减上限", "≤10%（温度范围内）", "Temporary reduction max 10% over defined fuel temperature range", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.2.3", "26", "Lfd.337-339"],

    # ===== TDI 供油性能 =====
    [20, "供油性能-TDI", "TDI工作电压范围", "6~16V", "TDI operating voltage range: 6 to 16 V", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.1", "31", "Lfd.397-400"],
    [21, "供油性能-TDI", "TDI最大工作电流", "12A", "TDI max operating current: 12 A", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.1", "31", ""],
    [22, "供油性能-TDI", "TDI工作压力范围-供油", "50~700kPa", "TDI feed operating pressure: 50-700 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.2", "31", "Lfd.401-402"],
    [23, "供油性能-TDI", "TDI工作压力范围-回油", "0~100kPa", "TDI return operating pressure: 0-100 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.2", "31", ""],
    [24, "供油性能-TDI", "TDI建压时间", "<200ms至550kPa", "TDI pressure build-up: <200ms to 550 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.4", "32", "BsM-Sa/Z, Lfd.411-415"],
    [25, "供油性能-TDI", "TDI冷启动性能", "<200ms, >12l/h @-24°C,9V", "Cold start: <200ms, >12l/h at -24°C, 9V", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.5", "32", "BsM-Sa/Z, Lfd.416-421"],
    [26, "供油性能-TDI", "TDI热启动性能", "<200ms, >12l/h @35~90°C,9V", "Hot start: <200ms, >12l/h at 35-90°C, 9V", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.7", "32", "Lfd.424-429"],
    [27, "供油性能-TDI", "TDI限压阀开启压力", "620~640kPa @>30l/h", "TDI opening pressure: 620-640 kPa at >30 l/h", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.8", "33", "Lfd.430-443"],
    [28, "供油性能-TDI", "TDI最大系统压力", "700kPa", "TDI maximum pressure: 700 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.8", "33", ""],
    [29, "供油性能-TDI", "TDI限压阀关闭压力", "≥62kPa", "TDI closing pressure: >=62 kPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.8", "33", ""],
    [30, "供油性能-TDI", "TDI泄漏量限制", "<1l/h（0至关闭压力）", "TDI permissible leakage: <1 l/h for 0 to closing pressure", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3.8", "33", ""],
    [31, "供油性能-TDI", "TDI泵型要求", "EC电子换向泵", "Supply system designed with EC electric fuel pump", "A(一般)", "LAH.3WA.201.B", "5.3.1.3.3", "31", "Lfd.395"],
    [32, "供油性能-TDI", "TDI防结冰设计", "必须可靠防止功能影响性结冰", "Design must reliably prevent icing affecting functionality", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3", "31", "Lfd.396"],
    [33, "供油性能-TDI", "TDI高压泵最低循环量", "持续循环（与燃油消耗无关）", "Minimum fuel quantity continuously circulated regardless of consumption", "C(关键)", "LAH.3WA.201.B", "5.3.1.3.3", "31", "Lfd.393"],

    # ===== 控制策略 =====
    [34, "控制策略", "PWM接口方式", "双向PWM（ECM↔KPE）", "Bidirectional pulse-width modulation (PWM) interface between ECM and KPE", "C(关键)", "LAH.3WA.201.B", "4.1", "5", "Lfd.38"],
    [35, "控制策略", "低压摆动适应下限(策略1)", "50kPa", "Wobbling adaption lower limit: 50 kPa", "C(关键)", "LAH.3WA.201.B", "4.1.3", "6", "Lfd.54"],
    [36, "控制策略", "低压下限(策略2)", "300kPa", "Application strategy 2: lower limit 300 kPa in pre-control characteristic map", "C(关键)", "LAH.3WA.201.B", "4.1.3", "7", "Lfd.59"],
    [37, "控制策略", "应急策略-最大驱动持续运行", "Kl.15 off/on复位", "Emergency strategy: continuous max actuation until terminal 15 off/on reset", "C(关键)", "LAH.3WA.201.B", "4.1.3", "7", "BsM-Sa, Lfd.60"],

    # ===== 效率 =====
    [38, "效率", "DC汽油流量泵效率", ">20% @430kPa", "Efficiency >20% at 430 kPa operating pressure with DC flow pump for gasoline", "A(一般)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-E.T, Lfd.192"],
    [39, "效率", "EC汽油流量泵效率", ">30% @430kPa", "Efficiency >30% at 430 kPa operating pressure with EC flow pump for gasoline", "A(一般)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-E.T, Lfd.193"],
    [40, "效率", "EC柴油容积泵效率", ">35% @450kPa", "Efficiency >35% at 450 kPa operating pressure with EC displacement pump for diesel", "A(一般)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-E.T, Lfd.194, geändert"],
    [41, "效率", "柴油最大效率设计点", "约60%最大模块流量", "Diesel max efficiency designed for ~60% of max module delivery rate", "A(一般)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-E.T, Lfd.196"],
    [42, "效率", "汽油最大效率设计点", "约40%最大模块流量", "Gasoline max efficiency designed for ~40% of max module delivery rate", "A(一般)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-E.T, Lfd.197"],
    [43, "效率", "DC流量散差（电功率-流量）", "≤20%容积泵/≤30%流量泵", "DC: variation ≤20% displacement / ≤30% flow pumps", "A(一般)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-E.T, Lfd.199"],
    [44, "效率", "EC流量散差（PWM-流量）", "≤15%容积泵/≤25%流量泵", "EC: distribution ≤15% displacement / ≤25% flow pumps", "A(一般)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-E.T, Lfd.200"],

    # ===== 密封性 =====
    [45, "密封性", "停放4周泄漏量", "<1ml", "Loss in volume <1 ml when unused for up to 4 weeks", "C(关键)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-Sa/Z, Lfd.185"],
    [46, "密封性", "停放1周泄漏量", "<0.25ml", "Loss in volume <0.25 ml when unused for up to 1 week", "C(关键)", "LAH.3WA.201.B", "5.3.1", "13", "BsM-Sa/Z, Lfd.186"],
    [47, "密封性", "虹吸效应防止", "必须始终防止虹吸", "Siphoning effects must always be prevented", "C(关键)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-S, Lfd.201"],

    # ===== 保压功能（仅汽油） =====
    [48, "保压功能", "关泵瞬时压降", "≤10% (按LAH.DUM.201.C)", "Momentary pressure drop max 10% per LAH.DUM.201.C", "C(关键)", "LAH.3WA.201.B", "5.3.1.4", "33", "Lfd.444-450"],
    [49, "保压功能", "关泵30min内继续压降", "≤20%", "Further 20% drop within 30 min after pump shutoff", "C(关键)", "LAH.3WA.201.B", "5.3.1.4", "33", ""],

    # ===== 涡流杯填充 =====
    [50, "涡流杯填充", "涡流杯最小可用容积", "≥0.4L", "Usable swirl pot volume of at least 0.4 l", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.1", "16", "BsM-SOTIF, Lfd.225"],
    [51, "涡流杯填充", "初始建压时间-涡流杯18mm液位", "<15s（汽油Lfd.216.1）", "Initial starting behavior: build required pressure <15s with surrounding fuel level at swirl pot of 18 mm", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.1", "15", "BsM-Sa/Z, Lfd.216"],
    [52, "涡流杯填充", "缺油重启-涡流杯18mm", "<10s建压", "Restarting after breakdown: swirl pot filled to 18mm, pressure within <10s", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.1", "16", "BsM-Sa/Z, Lfd.226"],
    [53, "涡流杯填充", "涡流杯倾斜交付率", ">40s @35%燃油消耗, 30°倾角", "Swirl pot delivery rate for ICE available for >40s at 35% of max consumption, 30° inclination", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.1", "16", "BsM-Sa/Z, Lfd.224"],
    [54, "涡流杯填充", "射流泵中断后恢复", "<2s恢复定义流量", "Suction jet pump: ensure defined delivery rate within <2s following each supply interruption", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.2", "17", "BsM-Z, Lfd.251"],
    [55, "涡流杯填充", "射流泵≤10s中断后恢复", "达到定义流量", "Defined suction jet pump delivery rate after ≤10s supply interruption", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.2", "17", "BsM-Z, Lfd.252"],

    # ===== 过滤性能 =====
    [56, "过滤性能", "精滤器容尘量", ">12g (ISO 19438)", "Dirt-holding capacity >12g as per ISO 19438", "C(关键)", "LAH.3WA.201.B", "5.4.7", "57", "BsM-Sa/Z, Lfd.855"],
    [57, "过滤性能", "精滤器最大压降", "45kPa @100%负载", "Max pressure drop 45 kPa for 100% loaded state as per ISO 19438", "C(关键)", "LAH.3WA.201.B", "5.4.7", "57", "BsM-Sa/Z, Lfd.856"],
    [58, "过滤性能", "精滤器分离效率 @3μm", "≥65%", "Separation efficiency ≥3μm: 65%", "C(关键)", "LAH.3WA.201.B", "5.4.7", "57", "BsM-Sa/Z, Lfd.857"],
    [59, "过滤性能", "精滤器分离效率 @5μm", "≥82%", "Separation efficiency ≥5μm: 82%", "C(关键)", "LAH.3WA.201.B", "5.4.7", "57", "BsM-Sa/Z, Lfd.857"],
    [60, "过滤性能", "精滤器分离效率 @10μm", "≥95%", "Separation efficiency ≥10μm: 95%", "C(关键)", "LAH.3WA.201.B", "5.4.7", "57", "BsM-Sa/Z, Lfd.857"],
    [61, "过滤性能", "精滤器分离效率 @20μm", "≥99%", "Separation efficiency ≥20μm: 99%", "C(关键)", "LAH.3WA.201.B", "5.4.7", "58", "BsM-Sa/Z, Lfd.857"],
    [62, "过滤性能", "精滤器新品清洁度", "50-100μm≤500个, 100-200μm≤75个, 200-500μm≤2个, >500μm=0个", "Filter cleanliness: 50-100μm/500, 100-200μm/75, 200-500μm/2, >500μm/0", "C(关键)", "LAH.3WA.201.B", "5.4.7", "58", "BsM-Sa/Z, Lfd.858"],

    # ===== 吸油滤网 =====
    [63, "吸油滤网", "EKP吸油滤网最小面积", "≥6500mm²", "Min total surface area of strainer upstream of main pump: 6500 mm²", "C(关键)", "LAH.3WA.201.B", "5.4.8", "58", "BsM-Sa/Z, Lfd.863"],
    [64, "吸油滤网", "射流泵滤网网目", "600μm (基于4mm混合管径)", "Suction jet pump strainer mesh size: 600μm based on 4mm mixing tube diameter", "C(关键)", "LAH.3WA.201.B", "5.4.5", "57", "BsM-Z, Lfd.835"],
    [65, "吸油滤网", "滤网自由截面积", "≥10倍混合管径", "Free cross section of extraction strainer: 10x mixing tube diameter", "C(关键)", "LAH.3WA.201.B", "5.4.5", "57", "BsM-Z, Lfd.836"],

    # ===== 诊断功能 =====
    [66, "诊断功能", "EC泵故障类型数量", "9种故障(Table 3)", "EC fault definitions: 9 faults per Table 3", "C(关键)", "LAH.3WA.201.B", "5.3.3.1", "40", "BsM-Sa, Lfd.547-550"],
    [67, "诊断功能", "DC泵故障类型数量", "7种故障(Table 4)", "DC fault definitions: 7 faults per Table 4", "C(关键)", "LAH.3WA.201.B", "5.3.3.1", "40", "BsM-Sa, [冲突] DC缺Fault 8(过温)和Fault 9(低压转速),确认是否N/A"],
    [68, "诊断功能", "诊断有效电压范围", "9~16V", "Diagnostics in input voltage range 9V to 16V", "C(关键)", "LAH.3WA.201.B", "5.3.3.1", "40", "BsM-Sa, Lfd.548"],
    [69, "诊断功能", "故障信号传输方式", "PWM线接地短路（定义时间t_low）", "Fault code via PWM interface by short circuiting PWM_KPE to ground for defined time t_low", "C(关键)", "LAH.3WA.201.B", "5.3.3.1", "40", "Lfd.547"],

    # ===== 事件存储 =====
    [70, "事件存储", "最小故障事件存储数", "≥200条（FIFO循环）", "Min 200 fault events, FIFO circular buffer", "C(关键)", "LAH.3WA.201.B", "5.3.4.2", "53", "Lfd.782-789"],
    [71, "事件存储", "非易失性存储要求", "断电后保持", "Internal non-volatile memory for subsequent analyses", "C(关键)", "LAH.3WA.201.B", "5.3.4", "53", "Lfd.773-804"],

    # ===== 电气接口-KPE =====
    [72, "电气接口", "KPE线缆最大长度(EMC)", "≤300mm", "Cable length between EC KPE and electrical load ≤300mm for EMC", "C(关键)", "LAH.3WA.201.B", "4.2.3", "9", "BsM-E.EMZ, Lfd.117"],
    [73, "电气接口", "KPE初始化时间", "≤20ms至就绪状态", "Max 20ms to ready-for-operation state", "C(关键)", "LAH.3WA.201.B", "5.3.1.7.1", "36", "Lfd.491-496"],
    [74, "电气接口", "PWM输入信号频率", "100Hz ±5Hz", "PWM_KPE input signal: 100 Hz ±5 Hz", "C(关键)", "LAH.3WA.201.B", "5.4.11.2.3.1", "73", "Lfd.1131-1144"],
    [75, "电气接口", "PWM占空比范围", "10%~90%", "PWM duty cycle: 10%-90%", "C(关键)", "LAH.3WA.201.B", "5.4.11.2.3.1", "73", ""],
    [76, "电气接口", "EC泵驱动PWM频率", "9~25kHz", "Output signal for rotating field: PWM 9-25 kHz", "A(一般)", "LAH.3WA.201.B", "5.4.11.2.3.2", "74", "Lfd.1145-1149"],

    # ===== KPE硬件/软件 =====
    [77, "KPE设计", "处理器容量余量", "≥30%", "Min 30% computer capacity reserve", "C(关键)", "LAH.3WA.201.B", "5.4.11.1.2", "66", "Lfd.972-977"],
    [78, "KPE设计", "看门狗故障响应时间", "10ms", "Watchdog critical fault response time: 10 ms", "C(关键)", "LAH.3WA.201.B", "5.4.11.1.5", "67", "Lfd.1008-1020"],

    # ===== 功能安全(FuSi) =====
    [79, "功能安全", "FuSi-ASIL适用性", "ASIL A~D按ISO 26262", "ASIL per ISO 26262 for safety-relevant parts", "C(关键)", "LAH.3WA.201.B", "2.10", "1-3", "neu hinzugefügt, Lfd.19.c-19.x"],
    [80, "功能安全", "碰撞工况EKP停机时间", "120ms内断开EKP供电", "KPE must deactivate EKP supply within 120 ms in crash event", "C(关键)", "LAH.3WA.201.B", "2.10", "3", "BsM-Sa.Crash, Lfd.19.ad-19.af"],
    [81, "功能安全", "泵停止请求响应-PWM异常", "120ms内断开EKP", "Deactivate EKP within 120ms if implausible PWM signal detected", "C(关键)", "LAH.3WA.201.B", "2.10", "3", "BsM-Sa.Crash, Lfd.19.ae"],
    [82, "功能安全", "泵停止请求响应-PWM断线", "120ms内断开EKP", "Deactivate EKP within 120ms if broken PWM line detected", "C(关键)", "LAH.3WA.201.B", "2.10", "3", "BsM-Sa.Crash, Lfd.19.af"],
    [83, "功能安全", "DIA文档义务", "必须填写VW DIA模板", "Contractor must evaluate and comment on DIA template", "C(关键)", "LAH.3WA.201.B", "2.10", "3", "Lfd.19.ag"],
    [84, "功能安全", "容错时间定义", "待与OEM协商", "Fault tolerance time to be agreed with purchaser", "C(关键)", "LAH.3WA.201.B", "2.10", "2", "[冲突] Lfd.19.u-19.x标注需协商,与120ms硬性要求需确认一致性"],

    # ===== 碰撞安全 =====
    [85, "碰撞安全", "碰撞无泄漏", "法兰无泄漏,失效点在油箱内定义位置", "Blind flange must not leak; component failure at defined locations inside fuel tank", "C(关键)", "LAH.3WA.201.B", "5.3.1", "12", "BsM-Sa/Z, Lfd.181"],
    [86, "碰撞安全", "碰撞后管路保持", "油箱内管路不可断开泄燃油", "Supply/return lines within fuel tank secured: fuel cannot escape in crash", "C(关键)", "LAH.3WA.201.B", "5.3.1", "12", "BsM-Sa/Z, Lfd.182"],

    # ===== 泄漏保护 =====
    [87, "泄漏保护", "泄漏保护阀设计压力-汽油常规", "16kPa + 80cm液柱", "Leakage protection: conventional gasoline 16kPa + hydrostatic height 80cm", "C(关键)", "LAH.3WA.201.B", "5.4.4", "56", "BsM-Sa/Z, Lfd.831"],
    [88, "泄漏保护", "泄漏保护阀设计压力-汽油增压", "41kPa + 80cm液柱", "Leakage protection: pressurized gasoline 41kPa + hydrostatic height 80cm", "C(关键)", "LAH.3WA.201.B", "5.4.4", "56", "BsM-Sa/Z, Lfd.831"],
    [89, "泄漏保护", "泄漏保护阀设计压力-柴油", "17kPa + 80cm液柱", "Leakage protection: diesel 17kPa + hydrostatic height 80cm", "C(关键)", "LAH.3WA.201.B", "5.4.4", "56", "BsM-Sa/Z, Lfd.831"],

    # ===== 油位测量 =====
    [90, "油位测量", "油位传感器总电阻", "340±1Ω 或公差带≤2Ω", "Total resistance: 340 ±1 Ohm or tolerance band ≤2 Ohm", "C(关键)", "LAH.3WA.201.B", "5.3.3.6", "52", "Lfd.762-772"],
    [91, "油位测量", "杠杆臂偏转角", ">100° <130°", "Lever arm deflection angle >100° <130°", "A(一般)", "LAH.3WA.201.B", "5.4.10.1", "60", "Lfd.877-897"],
    [92, "油位测量", "一种变体适配所有燃油", "通用设计", "One variant applicable to all fuel types", "C(关键)", "LAH.3WA.201.B", "5.3.2", "38", "[冲突] Lfd.522 vs 汽柴油物理差异, 待LAH.DUM.201.F确认"],

    # ===== 驻车加热器 =====
    [93, "驻车加热器", "加热器燃油管路流量", "630ml/h @20hPa", "Parking heater fuel line: 630 ml/h at 20 hPa", "C(关键)", "LAH.3WA.201.B", "5.3.1.5", "34", "Lfd.451-466"],
    [94, "驻车加热器", "加热器倾斜提取角", "≤11°锥角", "Fuel extraction at inclines up to 11° cone angle", "A(一般)", "LAH.3WA.201.B", "5.3.1.5", "34", ""],
    [95, "驻车加热器", "涡流杯底部预留高度", "≥25mm+20mm弹簧=45mm", "Swirl pot emptying only down to 45mm above bottom (25mm + 20mm telescopic spring)", "C(关键)", "LAH.3WA.201.B", "5.3.1.5", "34", "BsM-Sa/Z"],

    # ===== 压力跌落 =====
    [96, "压力稳定性", "压力跌落限制", ">50kPa瞬态不可接受", "Pressure dips >50 kPa in supply not permissible", "C(关键)", "LAH.3WA.201.B", "5.3.1.1.2", "18", "BsM-Sa/Z, Lfd.271"],

    # ===== 排放 =====
    [97, "排放", "HC排放限值-PZEV/LEV III", "2 mg HC/天", "PZEV/LEV III: 2 mg HC/day", "C(关键)", "LAH.3WA.201.B", "5.16.9", "75", "Lfd.1209"],
    [98, "排放", "HC排放限值-其他", "20 mg HC/天", "All others: 20 mg HC/day", "C(关键)", "LAH.3WA.201.B", "5.16.9", "75", "Lfd.1210"],

    # ===== 管路 =====
    [99, "管路连接", "管路材料标准", "符合TL 82253", "All pressure parts including joints comply with TL 82253", "C(关键)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-Sa/Z, Lfd.202"],
    [100, "管路连接", "油位传感器间距", "≥10mm(含公差)", "At least 10mm clearance to fuel level sensor envelope", "C(关键)", "LAH.3WA.201.B", "5.4.6", "57", "BsM-Sa, Lfd.844"],

    # ===== 设计稳定性 =====
    [101, "设计稳定性", "启停系统适应性", "适配启停系统运行", "Supply system ASSY suitable for start-stop system operation", "C(关键)", "LAH.3WA.201.B", "5.3.1", "11", "Lfd.170"],
    [102, "设计稳定性", "模块防倾/防扭", "不允许在动态工况下倾斜/扭转", "Tilting or twisting of fuel supply module not permissible", "C(关键)", "LAH.3WA.201.B", "5.3.1", "14", "BsM-Sa/Z, Lfd.208"],
    [103, "设计稳定性", "污染条件下供油", "油箱有脏污/污染时仍可供油", "Must supply and extract fuel even with dirt or contamination in fuel tank", "C(关键)", "LAH.3WA.201.B", "5.3.1", "12", "BsM-Sa/Z, Lfd.177"],
    [104, "设计稳定性", "E100兼容设计", "滤网需兼容E100或可无损更换", "Design for E100 operation, or allow strainer change without other design changes", "C(关键)", "LAH.3WA.201.B", "5.4.8", "58", "BsM-Sa/Z, Lfd.862"],

    # ===== 电流监控 =====
    [105, "电流监控", "EC泵Fault 3:相电流过流", "触发并报ECM", "Fault 3: Phase overcurrent - trigger and report to ECM", "C(关键)", "LAH.3WA.201.B", "5.3.3.4.3", "44", "BsM-Sa, Lfd.614-634"],
    [106, "电流监控", "EC泵Fault 4:堵转", "触发并报ECM", "Fault 4: Blocked rotor - trigger and report to ECM", "C(关键)", "LAH.3WA.201.B", "5.3.3.4.4", "45", "BsM-Sa, Lfd.635-657"],
    [107, "电流监控", "EC泵Fault 8:末级过温故障", "触发并停止泵", "Fault 8: Overtemperature final stage fault - trigger and stop pump", "C(关键)", "LAH.3WA.201.B", "5.3.3.4.8", "49", "BsM-Sa, Lfd.721-743"],

    # ===== 系统级试验 =====
    [108, "试验要求", "试验规范引用", "LAH.3WA.201.C + F.SYS.3WA.201", "Testing per LAH.3WA.201.C and F.SYS.3WA.201", "A(一般)", "LAH.3WA.201.B", "6", "77", "Lfd.1240"],
]

# ============================================================
# GAP_DATA
# ============================================================
GAP_DATA = [
    ["BT-LAH通用基础规范：环境温度/振动/燃油/EMC/碰撞具体参数", "~30%需求参数值缺失(约70处引用)", "BT-LAH 'Fuel Systems'", "高"],
    ["LAH.DUM.201.C汽油泵零件规格：流量曲线/效率/耐久谱", "汽油泵全部L2特性", "LAH.DUM.201.C", "高"],
    ["LAH.DUM.201.E柴油泵零件规格：流量曲线/粘度/冷启动", "柴油泵全部L2特性", "LAH.DUM.201.E", "高"],
    ["LAH.DUM.201.F油位传感器规格：电阻值/精度/耐久性", "油位传感器L2特性", "LAH.DUM.201.F", "高"],
    ["VW 80000电气电子通用要求：EMC等级/ESD/电压保护", "KPE电气设计基础", "VW 80000", "高"],
    ["EP 21200.11系统耐久性试验计划", "全部耐久验证方案(49处引用)", "EP 21200.11", "中"],
    ["NNC试验规程(H/AA/AN/AK等)", "系统级验证方案", "NNC.201.021.xx系列", "中"],
    ["TL 82253压力件供货规范", "管路材料和强度标准", "TL 82253", "中"],
    ["NNC.201.021.xy占位符", "试验编号未最终确定", "待VW确认", "低"],
]

# ============================================================
# CONFLICT_DATA
# ============================================================
CONFLICT_DATA = [
    ["MPI Entfall vs 430kPa加载谱保留", "5.3.1.3.1全Entfall", "Lfd.292-293保留430kPa", "430kPa属FSI/MPI电子调压,加载谱有效,Freigabereport注明"],
    ["TDI压力参数三档", "550kPa运行验证", "660±20kPa耐久标称/680kPa最大背压", "获取LAH.DUM.201.E核对,特性清单标注三个使用场景"],
    ["EC 9故障 vs DC 7故障", "EC: 9种(含过温+低压转速)", "DC: 7种(缺Fault 8/9)", "按泵类型分别列DVP&R,确认DC缺失项为N/A"],
    ["120ms硬性停机 vs FuSi容错时间待协商", "Lfd.19.ad-af: 120ms", "Lfd.19.u-x: 需协商", "优先与VW确认ASIL等级,确保120ms满足时序"],
    ["油位传感器通用 vs 汽柴油差异", "Lfd.522: 通用设计", "汽油50hPa/柴油150hPa真空差异", "获取LAH.DUM.201.F确认是否区分燃油类型"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
CLASS_C_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
CONFLICT_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_HIGH = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
GAP_MID = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_LOW = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

L1_HEADERS = ["ID", "特性分类", "工程特性名称", "目标值/要求", "原文描述", "类别", "文件来源", "章节", "页码", "备注"]
L1_WIDTHS = [6, 15, 40, 30, 50, 10, 20, 10, 6, 30]


def apply_header(ws, headers, widths):
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"


def build_l1(wb):
    ws = wb.active
    ws.title = "工程特性清单"
    apply_header(ws, L1_HEADERS, L1_WIDTHS)
    for row in L1_DATA:
        ws.append(row)
    for r in range(2, len(L1_DATA) + 2):
        cls = ws.cell(r, 6).value  # 类别列
        note = str(ws.cell(r, 10).value or "")  # 备注列
        for c in range(1, len(L1_HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP_TOP
            if "[冲突]" in note:
                cell.fill = CONFLICT_FILL
            elif str(cls).startswith("C"):
                cell.fill = CLASS_C_FILL


def build_gaps(wb):
    ws = wb.create_sheet("缺口与冲突")

    # 缺口部分
    ws.cell(1, 1, "— 缺口 —").font = Font(bold=True, size=12)
    gap_headers = ["缺失信息", "影响范围", "所需文档", "优先级"]
    gap_widths = [40, 30, 30, 10]
    for c, (h, w) in enumerate(zip(gap_headers, gap_widths), 1):
        cell = ws.cell(2, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = w

    priority_fills = {"高": GAP_HIGH, "中": GAP_MID, "低": GAP_LOW}
    for i, row in enumerate(GAP_DATA, 3):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
        p = row[3] if len(row) > 3 else ""
        if p in priority_fills:
            for c in range(1, len(gap_headers) + 1):
                ws.cell(i, c).fill = priority_fills[p]

    # 冲突部分
    conflict_start = len(GAP_DATA) + 5
    ws.cell(conflict_start, 1, "— 冲突 —").font = Font(bold=True, size=12)
    conflict_headers = ["参数", "文档A值", "文档B值", "建议处理"]
    conflict_widths = [30, 30, 30, 30]
    for c, (h, w) in enumerate(zip(conflict_headers, conflict_widths), 1):
        cell = ws.cell(conflict_start + 1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, w
        )
    for i, row in enumerate(CONFLICT_DATA, conflict_start + 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
            cell.fill = CONFLICT_FILL


def main():
    wb = openpyxl.Workbook()
    build_l1(wb)
    if GAP_DATA or CONFLICT_DATA:
        build_gaps(wb)
    wb.save(CONFIG["输出路径"])
    c_count = sum(1 for r in L1_DATA if len(r) > 5 and str(r[5]).startswith("C"))
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"L1特性: {len(L1_DATA)} 条 (C类:{c_count}, A类:{len(L1_DATA)-c_count})")
    if GAP_DATA:
        print(f"缺口: {len(GAP_DATA)} 项")
    if CONFLICT_DATA:
        print(f"冲突: {len(CONFLICT_DATA)} 项")


if __name__ == "__main__":
    main()
