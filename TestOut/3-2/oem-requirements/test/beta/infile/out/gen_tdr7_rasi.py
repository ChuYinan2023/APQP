"""
TDR7 RASI矩阵生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr7_rasi.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR7_RASI矩阵.xlsx",
}

# ============================================================
# ROLES — 角色列表（从SDT团队表提取）
# ============================================================
ROLES = [
    "PM", "DE", "QE", "TE", "ME", "PE", "SCM", "客户SQE",
]

# ============================================================
# DATA — RASI分配
# 列顺序: [序号, 交付物, PM, DE, QE, TE, ME, PE, SCM, 客户SQE]
# 值: "R"=负责 / "A"=决策 / "S"=支持 / "I"=知情 / ""=无关
# ============================================================
DATA = [
    #                                PM    DE    QE    TE    ME    PE    SCM   SQE
    [1,  "TDR1 成本分解",           "A",  "S",  "",   "",   "S",  "R",  "",   "I"],
    [2,  "TDR3 CDS开发表",         "R",  "S",  "S",  "",   "S",  "",   "",   "I"],
    [3,  "TDR4 3D模型",            "I",  "R",  "",   "",   "",   "",   "",   "A"],
    [4,  "TDR5 SDT团队表",         "R",  "",   "",   "",   "",   "",   "",   "I"],
    [5,  "TDR6 例外事项清单",      "A",  "R",  "S",  "S",  "S",  "",   "",   "I"],
    [6,  "TDR7 RASI矩阵",         "R",  "",   "",   "",   "",   "",   "",   "I"],
    [7,  "TDR8 补充材料清单",      "A",  "S",  "S",  "R",  "",   "S",  "",   "I"],
    [8,  "DFMEA",                   "I",  "R",  "A",  "S",  "S",  "",   "",   "I"],
    [9,  "PFMEA",                   "I",  "S",  "A",  "S",  "R",  "",   "",   "I"],
    [10, "DV/PV Plan",              "I",  "S",  "A",  "R",  "",   "",   "",   "I"],
    [11, "DV测试执行",             "I",  "S",  "S",  "R",  "",   "",   "",   "I"],
    [12, "PV测试执行",             "I",  "",   "S",  "R",  "",   "",   "",   "A"],
    [13, "模具设计",               "I",  "R",  "",   "",   "A",  "S",  "",   ""],
    [14, "模具制造跟踪",           "I",  "",   "",   "",   "R",  "A",  "",   ""],
    [15, "A5 特殊特性清单",         "I",  "R",  "A",  "S",  "",   "",   "",   "I"],
    [16, "A6 质量目标表",           "I",  "S",  "R",  "S",  "",   "",   "",   "A"],
    [17, "PPAP提交",               "I",  "S",  "R",  "S",  "S",  "",   "S",  "A"],
    [18, "IMDS提交",               "I",  "",   "R",  "",   "",   "S",  "",   "I"],
    [19, "包装方案",               "I",  "",   "",   "",   "S",  "",   "R",  "A"],
    [20, "Run@Rate",               "I",  "",   "A",  "",   "R",  "",   "S",  "I"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

RASI_FILLS = {
    "R": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "A": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "S": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "I": PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RASI矩阵"

    # 表头
    headers = ["序号", "交付物"] + ROLES
    widths = [6, 30] + [10] * len(ROLES)
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "C2"

    # 数据
    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = CENTER
            cell.font = BODY_FONT
            v = str(cell.value or "").upper()
            if v in RASI_FILLS:
                cell.fill = RASI_FILLS[v]
        ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 图例
    lr = len(DATA) + 4
    ws.cell(lr, 1, "图例").font = Font(name='Arial', bold=True, size=11)
    for i, (k, desc) in enumerate([("R", "Responsible 负责执行"), ("A", "Accountable 最终决策"),
                                    ("S", "Support 支持协助"), ("I", "Informed 知情通报")]):
        ws.cell(lr + 1 + i, 1, k).font = BODY_FONT
        ws.cell(lr + 1 + i, 1).fill = RASI_FILLS[k]
        ws.cell(lr + 1 + i, 1).alignment = CENTER
        ws.cell(lr + 1 + i, 2, desc).font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"交付物: {len(DATA)} 项, 角色: {len(ROLES)} 人")


if __name__ == "__main__":
    main()
