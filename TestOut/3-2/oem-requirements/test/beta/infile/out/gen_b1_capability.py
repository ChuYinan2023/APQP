"""
B1能力声明模板生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_b1_capability.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "B1_能力声明模板.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 评估项, 客户要求, 我方现状, 判定, 备注]
# 判定: "符合"/"例外"/"待确认" — 有company-profile时AI自动填，无则留空让用户填
# 客户要求: 从CTS/PF/SSTS提取（绿色底）
# 我方现状: 从company-profile提取或用户填写（黄色底）
# ============================================================
DATA = [
    # === 质量体系 ===
    ["质量体系", "IATF 16949认证", "IATF 16949:2016", "已认证（2027-06到期）", "符合", ""],
    ["质量体系", "ISO 14001认证", "ISO 14001:2015", "已认证（2026-12到期）", "符合", ""],
    ["质量体系", "VDA 6.3审核", "VDA 6.3", "A级（2026-06到期）", "符合", ""],
    ["质量体系", "IMDS注册", "IMDS材料声明", "已注册", "符合", ""],
    ["质量体系", "FMEA工具", "DFMEA/PFMEA per 00270/00271", "APIS IQ 8.0", "符合", ""],
    ["质量体系", "SPC过程控制", "Cpk≥1.67(CC)/1.33(SC)", "QI-Macros + 在线SPC", "符合", ""],
    # === 制造工艺 ===
    ["制造能力", "注塑成型", "PA66 GF30, PA12, POM注塑件", "海天80T-260T×8台, 50万件/年", "符合", "Damper/QC/管件"],
    ["制造能力", "CNC弯管", "燃油管路弯管成型", "CNC弯管机×3, 30万件/年", "符合", "钢管/铝管"],
    ["制造能力", "管端成型", "扩口/缩口/翻边", "管端成型机×2, 30万件/年", "符合", ""],
    ["制造能力", "超声波焊接", "塑料件焊接", "超声波焊接机×2, 40万件/年", "符合", ""],
    ["制造能力", "装配+在线气密", "总成装配+100%气密检测", "半自动装配线×2, 50万件/年", "符合", "气密检测集成"],
    ["制造能力", "激光打标", "零件追溯码/Logo", "光纤激光机×1", "符合", "per PS-4480"],
    ["制造能力", "表面处理-镀镍", "per PS-8688钢管涂层", "无内部产线", "例外", "委外：XX表面处理厂"],
    ["制造能力", "表面处理-涂装", "涂装处理", "无内部涂装线", "例外", "委外; 如仅含塑料管则不适用"],
    # === 测试能力 ===
    ["测试能力", "爆破压力", "0-100 bar爆破试验", "爆破试验机, 0-100 bar", "符合", ""],
    ["测试能力", "泄漏检测(VLD)", "VLD 15μm×3mm", "氦检漏仪, 10μm VLD", "符合", "CC特性"],
    ["测试能力", "在线气密", "100%差压检漏", "差压检漏仪, 0.1 kPa精度", "符合", "在线100%"],
    ["测试能力", "脉冲耐久", "0-40 bar, 300k~600k次", "脉冲试验台 4工位, 0-40 bar", "符合", ""],
    ["测试能力", "拉脱力", "管路≥450N / QC≥450N", "万能试验机 0-50 kN (Zwick)", "符合", ""],
    ["测试能力", "温度循环", "-50~180°C环境试验", "环境箱×2, -50~180°C", "符合", ""],
    ["测试能力", "盐雾试验", "腐蚀测试", "盐雾箱, 480h", "待确认", "SAE J2334需循环腐蚀，待确认兼容性"],
    ["测试能力", "清洁度", "ISO 16232, ≤1.5mg/dm²", "清洁度分析仪 (ISO 16232)", "符合", "颗粒计数"],
    ["测试能力", "燃油兼容性", "柴油C5/C6/C7浸泡", "燃油浸泡箱 (E10/E85/B20)", "待确认", "需确认柴油C系列燃料配方"],
    ["测试能力", "流量测试", "压降/流阻", "流量计 0-200 L/h", "符合", ""],
    ["测试能力", "尺寸测量", "GD&T per ASME Y14.5", "CMM三坐标 0.003mm (Hexagon)", "符合", ""],
    ["测试能力", "振动耐久", "R95C90频谱试验", "无振动试验台", "例外", "委外：国家实验室"],
    ["测试能力", "MiniSHED", "蒸发排放 per LP.7A005", "无SHED设备", "例外", "委外第三方实验室"],
    ["测试能力", "静电消散", "per SAE J1645 / DC≤1MΩ", "待确认", "待确认", "需确认是否有专用设备"],
    ["测试能力", "碎石冲击", "SAE J400, -30°C", "无碎石设备", "例外", "委外第三方实验室"],
    ["测试能力", "Charpy冲击", "-30°C低温冲击", "待确认Charpy附件", "待确认", "万能试验机可能需附件"],
    # === 软件工具 ===
    ["软件工具", "3D CAD", "UG/NX .prt 或 Parasolid", "CATIA V5 R2021", "例外", "需格式转换或采购NX"],
    ["软件工具", "2D CAD", "CAD出图 per CS.00080", "AutoCAD 2023", "符合", "需CS.00080确认差异"],
    ["软件工具", "CAE仿真", "虚拟分析(TDR#2)", "ANSYS 2023R1", "符合", "TDR#2标注Not Applicable"],
    ["软件工具", "ERP", "供应链管理", "SAP S/4HANA", "符合", ""],
    ["软件工具", "PLM", "产品生命周期管理", "无PLM (文件服务器)", "待确认", "客户如要求需采购Teamcenter"],
    # === 设计能力 ===
    ["设计能力", "Damper设计", "Damper子总成（参考P003161）", "有注塑+装配能力", "待确认", "基于参考图纸自行调整设计"],
    ["设计能力", "产品经验", "燃油管路总成开发", "主要产品：燃油管路、QC接头、塑料管件", "符合", "Stellantis/VW/Toyota经验"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
REQ_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PEND_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "评估项", "客户要求", "我方现状", "判定", "备注"]
WIDTHS = [15, 25, 30, 30, 10, 25]
JUDGMENT_MAP = {"符合": PASS_FILL, "例外": FAIL_FILL, "待确认": PEND_FILL}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "能力声明"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # 类别列
        ws.cell(r, 1).font = CAT_FONT
        ws.cell(r, 1).fill = CAT_FILL
        # 客户要求（绿色）
        ws.cell(r, 3).fill = REQ_FILL
        # 我方现状（黄色，待填）
        if not ws.cell(r, 4).value:
            ws.cell(r, 4).fill = INPUT_FILL
        # 判定列颜色
        j = str(ws.cell(r, 5).value or "")
        if j in JUDGMENT_MAP:
            ws.cell(r, 5).fill = JUDGMENT_MAP[j]
        elif not j:
            ws.cell(r, 5).fill = INPUT_FILL

    # 合并同类别
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"评估项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
