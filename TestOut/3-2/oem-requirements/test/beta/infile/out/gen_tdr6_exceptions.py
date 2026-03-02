"""
TDR6例外事项清单生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr6_exceptions.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR6_例外事项清单.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 序号, 客户要求, 要求来源, 判定, 我方现状, 替代方案, 备注]
# 判定: "符合"/"例外"/"待确认" — 有company-profile时AI自动填
# 同类别行自动合并
# ============================================================
DATA = [
    # === 软件工具 ===
    ["软件工具", 1, "3D CAD: UG (Unigraphics) .prt 或 Parasolid 格式",
     "TDR#4", "例外",
     "使用 CATIA V5 R2021，无UG/NX许可证",
     "方案1: 通过STEP/Parasolid中间格式转换交付;\n方案2: 如客户严格要求原生UG格式，需采购NX许可证",
     "建议与客户确认是否接受Parasolid格式"],
    ["软件工具", 2, "CAD出图规则 per CS.00080",
     "CTS Table 5", "例外",
     "CATIA V5出图，非UG出图规则",
     "参照CS.00080要求，使用CATIA按等效规则出图",
     "需获取CS.00080确认差异"],
    ["软件工具", 3, "PLM/Teamcenter集成",
     "TDR通用要求", "待确认",
     "无PLM系统，使用文件服务器管理",
     "如客户要求可采购Teamcenter; 或通过文件包交付",
     "确认客户是否强制要求PLM"],
    # === 制造工艺 ===
    ["制造工艺", 4, "表面处理-镀镍（钢管涂层 per PS-8688）",
     "PF.90197 §6.2.1", "例外",
     "无内部镀镍产线",
     "委外：XX表面处理厂（已签长期协议）",
     "需PS-8688确认具体涂层要求"],
    ["制造工艺", 5, "表面处理-涂装",
     "PF.90197 §6.2.1", "例外",
     "无内部涂装线",
     "委外处理",
     "如仅含塑料管则不适用"],
    # === 测试能力 ===
    ["测试能力", 6, "振动耐久试验 (R95C90, Table 8频谱)",
     "PF.90197 §9.4", "例外",
     "无振动耐久试验台",
     "委外：国家实验室（已有合作）",
     "DV/PV各45件"],
    ["测试能力", 7, "MiniSHED蒸发排放测试 (per LP.7A005)",
     "PF.90197 §8.2", "例外",
     "无MiniSHED设备",
     "委外：具备SHED能力的第三方实验室",
     "CC特性<E>; 需LP.7A005程序"],
    ["测试能力", 8, "静电消散测试 (per SAE J1645)",
     "PF.90197 §6.2.5", "待确认",
     "VLD氦检有，静电消散专用测试待确认",
     "如无：委外或采购静电测试设备",
     "CC特性<S>; 需SAE J1645标准"],
    ["测试能力", 9, "SAE J2334循环腐蚀 (180 cycles)",
     "PF.90298 §5.1", "待确认",
     "有盐雾箱(480h)，SAE J2334需循环腐蚀箱",
     "如无循环腐蚀箱：委外第三方实验室",
     "需确认现有盐雾箱是否满足J2334要求"],
    ["测试能力", 10, "碎石冲击 (SAE J400, -30°C)",
     "PF.90298 §7.6", "待确认",
     "无碎石冲击设备",
     "委外第三方实验室",
     "需SAE J400标准"],
    ["测试能力", 11, "Charpy低温冲击 (-30°C)",
     "PF.90298 §6.3.5", "待确认",
     "万能试验机有，Charpy冲击附件待确认",
     "如无：委外或采购Charpy冲击附件",
     "QC接头动态冲击测试"],
    # === 设计能力 ===
    ["设计能力", 12, "Damper子总成设计（参考图纸P003161）",
     "图纸P003161", "待确认",
     "有注塑+装配能力，需基于参考设计调整",
     "基于Nobel Plastiques参考图纸，自行优化设计",
     "我方需调整设计，非直接复制"],
    ["设计能力", 13, "CAE虚拟分析",
     "TDR#2", "符合",
     "ANSYS 2023R1基础模块",
     "—",
     "TDR#2标注Not Applicable"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PEND_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "序号", "客户要求", "要求来源", "判定", "我方现状", "替代方案", "备注"]
WIDTHS = [15, 6, 30, 15, 10, 25, 25, 20]
JUDGMENT_MAP = {"符合": PASS_FILL, "例外": FAIL_FILL, "待确认": PEND_FILL}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "例外事项清单"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # 类别列
        ws.cell(r, 1).font = CAT_FONT
        ws.cell(r, 1).fill = CAT_FILL
        # 判定列颜色
        j = str(ws.cell(r, 5).value or "")
        if j in JUDGMENT_MAP:
            ws.cell(r, 5).fill = JUDGMENT_MAP[j]
        elif not j:
            ws.cell(r, 5).fill = INPUT_FILL
        # 我方现状（无值则黄色）
        if not ws.cell(r, 6).value:
            ws.cell(r, 6).fill = INPUT_FILL

    # 合并同类别
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    # 统计
    sr = len(DATA) + 4
    ws.cell(sr, 1, "统计").font = Font(name='Arial', bold=True, size=11)
    for i, (label, fill) in enumerate([("符合", PASS_FILL), ("例外", FAIL_FILL), ("待确认", PEND_FILL)]):
        ws.cell(sr + 1 + i, 1, label).font = BODY_FONT
        ws.cell(sr + 1 + i, 1).fill = fill
        ws.cell(sr + 1 + i, 2, f'=COUNTIF(E2:E{len(DATA)+1},"{label}")').font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"要求项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
