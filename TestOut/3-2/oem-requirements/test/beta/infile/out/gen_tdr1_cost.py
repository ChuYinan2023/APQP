"""
TDR1成本分解生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr1_cost.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR1_成本分解.xlsx",
    "零件号": "FC00SAA78530",
    "零件名": "Fuel supply line — diesel filter to engine (KP1 A&B)",
    "管理费率": 0.15,
    "废料率": 0.03,
    "利润率": 0.08,
}

# ============================================================
# BOM_DATA — 原材料（从Step1 BOM提取）
# 列顺序: [序号, 零件/材料, 材料牌号, 单位, 用量, 单价(¥), 备注]
# 注：完整BOM需3D CAD确认，以下为基于Damper BOM+通用管路结构的初步估算
# ============================================================
BOM_DATA = [
    [1, "Damper外壳 (Capot)", "PA66+PA6", "kg", 0.012, 28.00, "注塑件"],
    [2, "Damper堵头 (Plug)", "PA66 GF30", "kg", 0.005, 28.00, "注塑件"],
    [3, "Damper阻尼器体", "PA66 GF30", "kg", 0.008, 28.00, "注塑件"],
    [4, "Damper弹簧", "弹簧钢丝 SWP-B", "kg", 0.003, 45.00, "4.5bar"],
    [5, "Damper膜片 (Membrane)", "54U6002+AgN91S", "件", 1, 0, "单价待确认"],
    [6, "PA12管路 (φ8×10)", "PA12 Rilsan AMNO", "m", 0.8, 65.00, "挤出管,长度待3D确认"],
    [7, "QC快插接头×2", "POM Delrin 500P", "kg", 0.015, 32.00, "2件"],
    [8, "O-Ring (燃油侧FKM)", "FKM 75A", "件", 2, 0, "单价待确认"],
    [9, "O-Ring (外部FVMQ)", "FVMQ", "件", 2, 0, "单价待确认"],
    [10, "卡扣/固定夹", "PA66 GF30", "kg", 0.003, 28.00, ""],
    [11, "波纹保护套管", "PA12 φ12", "m", 0.3, 3.80, ""],
]

# ============================================================
# PROCESS_DATA — 加工工艺
# 列顺序: [序号, 工序, 设备, 时间(min), 费率(¥/min), 备注]
# ============================================================
PROCESS_DATA = [
    [1, "PA12管挤出", "挤出生产线", 0.60, 1.20, ""],
    [2, "注塑-Damper件(3件)", "海天160T", 0.90, 1.50, "Capot+Plug+Body"],
    [3, "注塑-QC接头(2件)", "海天80T", 0.45, 1.50, ""],
    [4, "注塑-卡扣", "海天80T", 0.20, 1.50, ""],
    [5, "管端成型", "管端成型机", 0.30, 1.80, "2端"],
    [6, "超声波焊接-Damper", "超声波焊接机", 0.40, 1.00, ""],
    [7, "Damper子装配", "半自动线", 0.50, 0.80, "弹簧+膜片+壳体"],
    [8, "总成装配", "半自动线", 0.80, 0.80, "管路+QC+Damper+卡扣"],
    [9, "清洗", "超声波清洗", 0.30, 0.60, "ISO 16232"],
    [10, "气密检测(100%)", "差压检漏仪", 0.30, 0.50, "在线100%"],
    [11, "VLD泄漏检测(100%)", "氦检漏仪", 0.40, 0.50, "CC特性"],
    [12, "激光打标", "光纤激光机", 0.15, 0.30, "追溯码"],
    [13, "外观检查", "目视", 0.20, 0.40, ""],
    [14, "包装", "手工", 0.30, 0.20, ""],
]

# ============================================================
# OTHER_DATA — 其他费用
# 列顺序: [序号, 项目, 金额(¥), 分摊方式, 备注]
# ============================================================
OTHER_DATA = [
    [1, "注塑模具-Damper(3腔)", 240000, "50万件分摊", "Capot+Plug+Body各1模"],
    [2, "注塑模具-QC接头", 80000, "50万件分摊", ""],
    [3, "注塑模具-卡扣", 80000, "50万件分摊", ""],
    [4, "弯管夹具", 15000, "—", "管型专用"],
    [5, "装配工装", 25000, "—", ""],
    [6, "气密检具", 35000, "—", "含密封夹具"],
    [7, "检具(Go/NoGo)", 40000, "—", "关键尺寸2套"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name='Arial', bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
MONEY_FMT = '¥ #,##0.00'


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成本分解"

    # 项目信息
    ws['A1'] = '零件号'
    ws['B1'] = CONFIG['零件号']
    ws['C1'] = '零件名'
    ws['D1'] = CONFIG['零件名']
    for c in ['A', 'C']:
        ws[f'{c}1'].font = Font(name='Arial', bold=True, size=10)
    row = 3

    # === 原材料 ===
    ws.cell(row, 1, "一、原材料成本").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    bom_headers = ["序号", "零件/材料", "材料牌号", "单位", "用量", "单价(¥)", "小计(¥)"]
    bom_widths = [6, 20, 18, 6, 10, 12, 12]
    for c, (h, w) in enumerate(zip(bom_headers, bom_widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    bom_start = row + 1
    for i, r in enumerate(BOM_DATA):
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.cell(row, 7, f'=E{row}*F{row}')
        ws.cell(row, 7).number_format = MONEY_FMT
        ws.cell(row, 7).fill = FORMULA_FILL
        ws.cell(row, 7).border = THIN
    bom_end = row
    row += 1
    ws.cell(row, 5, "材料小计").font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 7, f'=SUM(G{bom_start}:G{bom_end})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 7).number_format = MONEY_FMT
    ws.cell(row, 7).fill = FORMULA_FILL
    ws.cell(row, 7).border = THIN
    mat_total_row = row

    # === 加工工艺 ===
    row += 2
    ws.cell(row, 1, "二、加工成本").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    proc_headers = ["序号", "工序", "设备", "时间(min)", "费率(¥/min)", "小计(¥)", "备注"]
    for c, h in enumerate(proc_headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    proc_start = row + 1
    for i, r in enumerate(PROCESS_DATA):
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.cell(row, 6, f'=D{row}*E{row}')
        ws.cell(row, 6).number_format = MONEY_FMT
        ws.cell(row, 6).fill = FORMULA_FILL
        ws.cell(row, 6).border = THIN
    proc_end = row
    row += 1
    ws.cell(row, 4, "加工小计").font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 6, f'=SUM(F{proc_start}:F{proc_end})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 6).number_format = MONEY_FMT
    ws.cell(row, 6).fill = FORMULA_FILL
    ws.cell(row, 6).border = THIN
    proc_total_row = row

    # === 其他费用 ===
    row += 2
    ws.cell(row, 1, "三、其他费用").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    other_headers = ["序号", "项目", "金额(¥)", "分摊方式", "备注"]
    for c, h in enumerate(other_headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    for r in OTHER_DATA:
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT

    # === 汇总 ===
    row += 2
    ws.cell(row, 1, "四、成本汇总").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    summary = [
        ("材料成本", f'=G{mat_total_row}'),
        ("废料成本", f'=G{mat_total_row}*{CONFIG["废料率"]}'),
        ("加工成本", f'=F{proc_total_row}'),
        ("管理费", f'=F{proc_total_row}*{CONFIG["管理费率"]}'),
        ("利润", f'=(G{mat_total_row}+F{proc_total_row})*{CONFIG["利润率"]}'),
    ]
    for label, formula in summary:
        row += 1
        ws.cell(row, 2, label).font = BODY_FONT
        ws.cell(row, 2).border = THIN
        ws.cell(row, 3, formula).font = BODY_FONT
        ws.cell(row, 3).number_format = MONEY_FMT
        ws.cell(row, 3).fill = FORMULA_FILL
        ws.cell(row, 3).border = THIN
    row += 1
    ws.cell(row, 2, "单件总成本").font = Font(name='Arial', bold=True, size=11)
    ws.cell(row, 2).border = THIN
    sum_rows = [row - i for i in range(len(summary), 0, -1)]
    ws.cell(row, 3, f'=SUM(C{sum_rows[0]}:C{sum_rows[-1]})').font = Font(name='Arial', bold=True, size=11)
    ws.cell(row, 3).number_format = MONEY_FMT
    ws.cell(row, 3).fill = FORMULA_FILL
    ws.cell(row, 3).border = THIN

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"BOM: {len(BOM_DATA)} 项, 工序: {len(PROCESS_DATA)} 项, 其他: {len(OTHER_DATA)} 项")


if __name__ == "__main__":
    main()
