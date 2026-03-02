"""
A5特殊特性清单生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_a5_special_chars.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "A5_特殊特性清单.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [序号, 特性名称, 类别, 规范要求, 来源文件, 验证方法, 控制方法建议, 备注]
# 类别: "CC"=关键特性(安全/法规) / "SC"=重要特性(功能/耐久)
# ============================================================
DATA = [
    # === CC 关键特性（安全/法规）===
    [1, "爆破压力", "CC",
     "RT: ≥8×工作压力; 115°C: ≥3×工作压力;\n后处理后≥75%原值且≥3×工作压力",
     "PF.90197 §7.4", "爆破试验 (DV/PV各15件, P99C90)",
     "100%在线气密检测 + 抽检爆破", ""],
    [2, "泄漏 (VLD)", "CC",
     "液体管: VLD 15μm×3mm, 150 PSI;\n蒸汽管: VLD 20μm×3mm, 10 PSI",
     "PF.90197 §7.2 <S,E>", "真空泄漏检测 (DV/PV各15件, P99C90)",
     "100%在线VLD检测", "标注<S>(安全)+<E>(排放)"],
    [3, "静电消散 (管路)", "CC",
     "per SAE J1645",
     "PF.90197 §6.2.5 <S>", "静电消散测试 (DV/PV各15件, P99C90)",
     "材料入库检 + 首件确认", "标注<S>(安全); 需SAE J1645标准"],
    [4, "静电消散 (QC接头)", "CC",
     "DC电阻 ≤1 MΩ @ 500V",
     "PF.90298 §6.4.1 <S>", "绝缘电阻测试 (DV/PV各15件, P99C90)",
     "100%在线电阻检测", "标注<S>(安全)"],
    [5, "MiniSHED蒸发排放", "CC",
     "per LP.7A005",
     "PF.90197 §8.2 <E>", "MiniSHED测试 (DV 15件P99C90, PV 15件R95C90)",
     "材料认证+过程控制", "标注<E>(排放); 需LP.7A005程序"],
    [6, "限用物质合规", "CC",
     "per CS-9003 / CS.00265 / GADSL",
     "PF.90298 Table 1, CTS Table 4", "IMDS登记 + 材料成分声明",
     "供应商IMDS提交 + 禁用物质筛查", ""],
    [7, "碰撞安全", "CC",
     "per FMVSS 301",
     "PF.90197 §8.1, PF.90298 §8.1", "车辆级碰撞测试",
     "STELLANTIS负责车辆级测试", "供应商提供零件"],
    # === SC 重要特性（功能/耐久）===
    [8, "拉脱力 (管路)", "SC",
     "RT燃油管≥450N, 蒸汽管≥222N;\n115°C: ≥115N / ≥65N",
     "PF.90197 §6.3.5", "拉脱力试验 (DV/PV各15件, P99C90)",
     "首件+抽检拉脱力", ""],
    [9, "拉脱力 (QC接头)", "SC",
     "液体: 未浸泡≥450N, 已浸泡≥297N;\n蒸汽: 未浸泡≥222N, 已浸泡≥75N",
     "PF.90298 §7.3", "拉脱力试验 (DV/PV各15件, P99C90)",
     "首件+抽检拉脱力", "需SAE J2044标准"],
    [10, "清洁度 (管路)", "SC",
     "汽油/蒸汽管≤1.5 mg/dm²;\n柴滤下游: per Table 6分级限值",
     "PF.90197 §7.1", "ISO 16232清洁度检测 (DV/PV各15件, P99C90)",
     "过程清洁度控制 + 抽检", ""],
    [11, "清洁度 (QC接头)", "SC",
     "≤0.75 mg/dm²",
     "PF.90298 §6.2.4", "ISO 16232清洁度检测",
     "过程清洁度控制 + 抽检", ""],
    [12, "脉冲压力耐久", "SC",
     "P1(回油0-1bar), P2(供油2-5bar),\nP5(UniJet柴油0-5bar); 300k~600k次",
     "PF.90197 §9.5", "脉冲耐久试验 (DV/PV各45件, R95C90)",
     "过程参数控制(焊接/装配)", ""],
    [13, "振动耐久", "SC",
     "R95/C90, 通用频谱 per Table 8",
     "PF.90197 §9.4", "振动耐久试验 (DV/PV各45件, R95C90)",
     "过程参数控制", ""],
    [14, "寿命循环", "SC",
     "per SAE J2045 §4.6, R95C90",
     "PF.90197 §9.3", "寿命循环试验 (DV/PV各45件, R95C90)",
     "过程参数控制", "需SAE J2045标准"],
    [15, "燃油兼容性", "SC",
     "1048h循环 (4×250h@120°C+12h@130°C),\n柴油C5/C6/C7",
     "PF.90197 §6.3.6", "燃油浸泡试验 (DV/PV各15件, P99C90)",
     "材料认证 + 批次追溯", "需SD-11597燃料清单"],
    [16, "腐蚀耐受", "SC",
     "15年功能寿命, per CS.00251;\nQC: SAE J2334 180 cycles",
     "PF.90197 §5.3, PF.90298 §5.1", "循环腐蚀试验 (DV/PV各15件, P99C90)",
     "表面处理工艺控制", "需CS.00251 + SAE J2334标准"],
    [17, "接头插入力", "SC",
     "<11mm: ≤67N; ≥11mm: ≤111N",
     "PF.90197 §6.2.4, PF.90298 §7.2", "插入力测试 (DV/PV各15件, P99C90)",
     "QC接头尺寸SPC控制", ""],
    [18, "压降/流阻", "SC",
     "液体≤400 kPa, 蒸汽≤3.4 kPa;\n成型管≥80%截面积",
     "PF.90197 §7.3", "流阻测试 (DV/PV各15件, P99C90)",
     "弯管工艺+模具尺寸控制", ""],
    [19, "QC高温爆破", "SC",
     "液体QC ≥3447 kPa(500 psi);\n蒸汽QC ≥689 kPa(100 psi)",
     "PF.90298 §6.3.6", "高温爆破试验",
     "材料+装配过程控制", ""],
    [20, "QC侧向载荷", "SC",
     "8mm→225N, 10mm→310N, 12mm+→400N",
     "PF.90298 §7.4", "侧向载荷试验",
     "接头结构设计验证", ""],
    [21, "QC动态冲击", "SC",
     "Charpy -30°C: 金属50J/混合30J/塑料10J",
     "PF.90298 §6.3.5", "低温冲击试验",
     "材料认证+批次控制", ""],
    [22, "卡扣插拔力", "SC",
     "安装≤20N(双指)/40N(全手);\n拆卸≥100N",
     "PF.90197 §6.3.4", "卡扣力测试 (DV/PV各15件, P99C90)",
     "卡扣模具尺寸控制", ""],
    [23, "材料拉伸性能", "SC",
     "拉伸强度≥20 N/mm²(新件),\n延伸率≥160%",
     "PF.90197 §6.2.1 Table 3", "EN ISO 527-2拉伸试验",
     "材料入货检验", "需EN ISO 527-2标准"],
    [24, "QC锁扣疲劳", "SC",
     "30次拆装, R95C90",
     "PF.90298 §9.3", "锁扣拆装疲劳试验",
     "锁扣设计+材料控制", "需SAE J2044标准"],
    [25, "QC摇摆耐久", "SC",
     "蒸汽360,000次/液体600,000次,\n±12\" stroke",
     "PF.90298 §9.5", "摇摆耐久试验",
     "QC结构+装配过程控制", ""],
    [26, "氯化钙耐受(EMEA)", "SC",
     "CaCl2 50%, 60±2°C, 200h",
     "PF.90197 §5.4", "氯化钙浸泡试验 (DV/PV各15件, P99C90)",
     "材料认证", ""],
    [27, "化学耐受", "SC",
     "耐汽车化学品, 试后满足6.1/7.2/7.4",
     "PF.90197 §5.1", "化学品浸泡试验 (DV/PV各15件, P99C90)",
     "材料认证", ""],
    [28, "NVH评分", "SC",
     "主观评分 ≥ 8 (Table 4 Rating Scale)",
     "PF.90197 §6.3.2", "NVH主观评估 (STELLANTIS负责)",
     "设计优化", "STELLANTIS负责测试"],
    [29, "Damper耐久", "SC",
     "1,000,000 cycles, 2~13.5 bar @ 98°C",
     "图纸P003161 Note 9", "脉冲耐久试验",
     "Damper装配+弹簧过程控制", "参考公司图纸, 我方需调整设计"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CC_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
SC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

HEADERS = ["序号", "特性名称", "类别", "规范要求", "来源文件", "验证方法", "控制方法建议", "备注"]
WIDTHS = [6, 25, 8, 30, 20, 20, 25, 20]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "特殊特性清单"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        cls = str(ws.cell(r, 3).value or "")
        fill = CC_FILL if cls == "CC" else SC_FILL if cls == "SC" else None
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    wb.save(CONFIG["输出路径"])
    cc = sum(1 for r in DATA if len(r) > 2 and r[2] == "CC")
    sc = sum(1 for r in DATA if len(r) > 2 and r[2] == "SC")
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"特殊特性: {len(DATA)} 条 (CC:{cc}, SC:{sc})")


if __name__ == "__main__":
    main()
