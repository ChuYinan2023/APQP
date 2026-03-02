"""
A6质量目标表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_a6_quality_targets.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "A6_质量目标表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 指标名称, 目标值, 来源文件, 验证方法, 备注]
# 类别: 市场质量目标 / 可靠性目标 / 过程能力 / 清洁度目标 / 安全法规 / 环保合规 / 持续符合性
# 同类别的行会自动合并单元格
# ============================================================
DATA = [
    # === 市场质量目标 ===
    ["市场质量目标", "CCP (Customer Car Profile)", "承诺共享目标达成", "SSTS Quality Targets", "CCP跟踪系统", ""],
    ["市场质量目标", "QT (Quality Tracking)", "承诺共享目标达成", "SSTS Quality Targets", "QT跟踪系统", ""],
    ["市场质量目标", "ICP (Initial Customer Perception)", "0 for single component", "SSTS Quality Targets", "初期客户感知评估", ""],
    ["市场质量目标", "可靠性 c/1000 @3MIS", "0.00 c/1000 (Underbody)", "SSTS Quality Targets", "现场故障跟踪", ""],
    ["市场质量目标", "可靠性 c/1000 @12MIS", "0.035 c/1000 (Underbody)", "SSTS Quality Targets", "现场故障跟踪", ""],
    # === 可靠性目标 ===
    ["可靠性目标", "脉冲压力耐久", "P1/P2/P5 300k~600k次, R95C90", "PF.90197 §9.5", "脉冲耐久试验 (DV/PV各45件)", ""],
    ["可靠性目标", "振动耐久", "通用频谱 per Table 8, R95C90", "PF.90197 §9.4", "振动耐久试验 (DV/PV各45件)", ""],
    ["可靠性目标", "寿命循环", "per SAE J2045 §4.6, R95C90", "PF.90197 §9.3", "寿命循环试验 (DV/PV各45件)", "需SAE J2045标准"],
    ["可靠性目标", "QC摇摆耐久", "液体600k次/蒸汽360k次, ±12\"", "PF.90298 §9.5", "摇摆耐久试验", ""],
    ["可靠性目标", "QC锁扣疲劳", "30次拆装, R95C90", "PF.90298 §9.3", "锁扣疲劳试验", "需SAE J2044标准"],
    ["可靠性目标", "Damper耐久", "1,000,000 cycles, 2~13.5 bar @98°C", "图纸P003161 Note 9", "脉冲耐久试验", "参考公司图纸, 我方调整设计"],
    ["可靠性目标", "Bench Demonstration", "per PF.90068 Annex A", "SSTS Quality Targets §R29", "台架验证试验", "需PF.90068 Annex A(最低样本量)"],
    # === 功能试验 ===
    ["功能试验", "爆破压力", "RT≥8×WP; 115°C≥3×WP;\n后处理≥75%原值且≥3×WP", "PF.90197 §7.4", "爆破试验 (DV/PV各15件, P99C90)", ""],
    ["功能试验", "泄漏 (VLD)", "液:15μm×3mm/150PSI;\n汽:20μm×3mm/10PSI", "PF.90197 §7.2", "真空泄漏检测 (DV/PV各15件, P99C90)", "CC特性 <S,E>"],
    ["功能试验", "拉脱力 (管路)", "RT燃油管≥450N/蒸汽管≥222N;\n115°C: ≥115N/≥65N", "PF.90197 §6.3.5", "拉脱力试验 (DV/PV各15件, P99C90)", ""],
    ["功能试验", "拉脱力 (QC)", "液体:未浸泡≥450N/已浸泡≥297N;\n蒸汽:未浸泡≥222N/已浸泡≥75N", "PF.90298 §7.3", "拉脱力试验 (P99C90)", "需SAE J2044标准"],
    ["功能试验", "压降/流阻", "液体≤400kPa; 蒸汽≤3.4kPa;\n成型管≥80%截面积", "PF.90197 §7.3", "流阻测试 (DV/PV各15件, P99C90)", ""],
    ["功能试验", "接头插入力", "<11mm≤67N; ≥11mm≤111N", "PF.90197 §6.2.4", "插入力测试 (DV/PV各15件, P99C90)", ""],
    ["功能试验", "QC高温爆破", "液体QC≥3447kPa; 蒸汽QC≥689kPa", "PF.90298 §6.3.6", "高温爆破试验", ""],
    ["功能试验", "静电消散 (管路)", "per SAE J1645", "PF.90197 §6.2.5", "静电消散测试 (P99C90)", "CC特性 <S>; 需SAE J1645"],
    ["功能试验", "静电消散 (QC)", "DC电阻≤1MΩ @500V", "PF.90298 §6.4.1", "绝缘电阻测试 (P99C90)", "CC特性 <S>"],
    # === 过程能力 ===
    ["过程能力", "CC特性 Cpk", "Cpk ≥ 1.67", "CTS / SSTS", "SPC统计", "爆破/泄漏/静电等CC项"],
    ["过程能力", "SC特性 Cpk", "Cpk ≥ 1.33", "CTS / SSTS", "SPC统计", "拉脱力/清洁度/尺寸等SC项"],
    # === 清洁度目标 ===
    ["清洁度目标", "管路清洁度", "汽油/蒸汽管≤1.5 mg/dm²;\n柴滤下游 per Table 6分级限值", "PF.90197 §7.1", "ISO 16232清洁度检测 (P99C90)", ""],
    ["清洁度目标", "QC接头清洁度", "≤0.75 mg/dm²", "PF.90298 §6.2.4", "ISO 16232清洁度检测", ""],
    # === 安全法规 ===
    ["安全法规", "碰撞安全", "per FMVSS 301", "PF.90197 §8.1", "车辆级碰撞测试 (STELLANTIS)", "供应商提供零件"],
    ["安全法规", "蒸发排放 MiniSHED", "per LP.7A005", "PF.90197 §8.2", "MiniSHED测试 (DV P99C90, PV R95C90)", "CC特性 <E>; 需LP.7A005"],
    # === 环保合规 ===
    ["环保合规", "IMDS登记", "100%材料成分声明", "CTS Table 4, CS-9003", "IMDS系统提交", "需CS-9003限用物质标准"],
    ["环保合规", "GADSL合规", "无禁用物质", "CS.00265", "材料成分筛查", ""],
    # === 耐环境目标 ===
    ["耐环境目标", "腐蚀耐受", "15年功能寿命, per CS.00251", "PF.90197 §5.3", "循环腐蚀试验 (P99C90)", "需CS.00251标准"],
    ["耐环境目标", "QC腐蚀", "SAE J2334 180 cycles", "PF.90298 §5.1", "SAE J2334循环腐蚀", "需SAE J2334标准"],
    ["耐环境目标", "燃油兼容性", "1048h循环, 柴油C5/C6/C7", "PF.90197 §6.3.6", "燃油浸泡试验 (P99C90)", "需SD-11597燃料清单"],
    ["耐环境目标", "氯化钙耐受(EMEA)", "CaCl2 50%, 60±2°C, 200h", "PF.90197 §5.4", "氯化钙浸泡试验 (P99C90)", ""],
    ["耐环境目标", "化学品耐受", "耐汽车化学品, 试后满足泄漏/爆破", "PF.90197 §5.1", "化学品浸泡试验 (P99C90)", ""],
    ["耐环境目标", "温度范围", "-40°C ~ +130°C continuous", "图纸P003161, PF.90197 §5.2", "温度循环试验", "尼龙管连续90°C, 短期115°C"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "指标名称", "目标值", "来源文件", "验证方法", "备注"]
WIDTHS = [18, 25, 30, 20, 20, 20]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质量目标表"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        cat_cell = ws.cell(r, 1)
        cat_cell.font = CAT_FONT
        cat_cell.fill = CAT_FILL

    # 合并同类别单元格
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    wb.save(CONFIG["输出路径"])
    cats = len(set(r[0] for r in DATA)) if DATA else 0
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"质量指标: {len(DATA)} 条, {cats} 个类别")


if __name__ == "__main__":
    main()
