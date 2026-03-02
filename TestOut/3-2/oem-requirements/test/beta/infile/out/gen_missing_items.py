"""
缺失项跟踪表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_missing_items.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "缺失项跟踪表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [序号, 缺失文档, 引用来源, 紧急度, 来源方, 影响范围, 处理状态]
# 紧急度: 最高/高/中/低
# 来源方: 向客户要/自行购买/内部收集
# 处理状态: 未发起/已发起/已收到/不适用
# ============================================================
DATA = [
    # === 最高（阻塞报价）===
    [1, "3D CAD 数据 (UG/Parasolid)", "TDR §4, SSTS §2.4.3, Deliverables #10/#14", "最高", "向客户要",
     "TDR#4无法完成；Layout/Packaging无法验证；注：我方用CATIA V5，UG格式列为例外项", "未发起"],
    [2, "PF.90068 — 认证程序 Annex A\n(最低样本量定义)", "SSTS Quality Targets §R29", "最高", "向客户要",
     "DV/PV plan中最低样本量无法确定", "未发起"],
    [3, "SAE J2044 — Quick Connect Coupling Spec", "PF.90197 §1.2/6.2.4/6.3.5, PF.90298 §6.3.1/7.1~7.4/9.3~9.5", "最高", "自行购买",
     "QC接头全部功能/耐久测试程序依赖此标准", "未发起"],
    [4, "SAE J2045 — Fuel System Tubing Performance", "PF.90197 §5.1/6.3.5/7.1/9.3", "最高", "自行购买",
     "管路拉脱力/清洁度/寿命循环核心测试程序", "未发起"],
    [5, "SAE J2260 — Nonmetallic Fuel System Tubing", "PF.90197 §1.2/6.2.1", "最高", "自行购买",
     "塑料管材料/结构/测试基础标准", "未发起"],
    # === 高（TKO前需要）===
    [6, "CD.80064 — Fuel Bundle Core Documents", "PF.90197 Table 1, §10.1", "高", "向客户要",
     "核心设计参考文件包，不可供应商下载", "未发起"],
    [7, "SD-11597 — Fuel Compatibility List", "PF.90197 §6.3.6, PF.90298 §5.5", "高", "向客户要",
     "燃油兼容性测试的燃料清单（C5/C6/C7/B20/B30等级定义）", "未发起"],
    [8, "SD-M0008/03 — Plastic Fuel Line Approvals", "PF.90197 §1.2, Table 1", "高", "向客户要",
     "塑料管材/接头预批准清单", "未发起"],
    [9, "CS.00251 — Corrosion Requirements", "PF.90197 §5.3", "高", "向客户要",
     "腐蚀测试程序（15年功能寿命验证）", "未发起"],
    [10, "QR.00001 — GPAT", "PF.90197 Table 1, PF.90298 Table 1, CTS Table 2", "高", "向客户要",
     "全球产品保证测试要求", "未发起"],
    [11, "07740 — Qualification Procedure", "CTS Table 2, SSTS §4 Validation", "高", "向客户要",
     "DV/PV plan格式要求（Annex 2）", "未发起"],
    [12, "CS-9003 — Restricted Substances", "PF.90298 Table 1, CTS Table 4", "高", "向客户要",
     "限用物质合规（IMDS前置条件）", "未发起"],
    [13, "QR-10012 — Dimensional Quality Requirements", "PF.90197 §6.2.2, PF.90298 §6.2.2", "高", "向客户要",
     "尺寸验证程序和公差标准", "未发起"],
    # === 中（DV前到位即可）===
    [14, "SAE J2334 — Laboratory Cyclic Corrosion Test", "PF.90197 §5.3, PF.90298 §5.1", "中", "自行购买",
     "循环腐蚀测试方法（180 cycles）", "未发起"],
    [15, "SAE J400 — Chip Resistance Test", "PF.90197 §5.3, PF.90298 §7.6", "中", "自行购买",
     "碎石冲击测试方法", "未发起"],
    [16, "SAE J1645 — Fuel System Electrostatic Charge", "PF.90197 §6.2.5, PF.90298 §6.4.1", "中", "自行购买",
     "静电消散测试方法", "未发起"],
    [17, "DIN 51604 — FAM Testing Fluid", "PF.90197 §6.2.1, PF.90298 §5.5", "中", "自行购买",
     "FAM B测试液标准", "未发起"],
    [18, "EN ISO 527-2 — Plastics Tensile Properties", "PF.90197 §6.2.1 Table 3", "中", "自行购买",
     "塑料拉伸性能测试方法", "未发起"],
    [19, "DIN 53504 — Rubber Tensile Properties", "PF.90197 Table 1", "中", "自行购买",
     "橡胶拉伸应力/应变测试", "未发起"],
    [20, "LP.7A004 — Fuel System Electrostatic Charge", "PF.90197 §6.2.5", "中", "向客户要",
     "静电测试程序（不可供应商下载）", "未发起"],
    [21, "LP.7A005 — Mini-SHED Emission Test", "PF.90197 §8.2", "中", "向客户要",
     "蒸发排放MiniSHED测试程序", "未发起"],
    [22, "PS-8688 — Coatings for Steel Tubing", "PF.90197 §6.2.1, Table 1", "中", "向客户要",
     "钢管涂层标准", "未发起"],
    [23, "CS.00081 — Corrosion Requirement (QC)", "PF.90298 §5.1", "中", "向客户要",
     "QC接头腐蚀测试引用标准", "未发起"],
    [24, "MS.50017 — PA Nylon (Harmonized)", "CTS Table 3", "中", "向客户要",
     "PA尼龙材料规范", "未发起"],
    [25, "MS.50015 — TPV EPDM/PP (Harmonized)", "CTS Table 3", "中", "向客户要",
     "热塑性硫化胶材料规范", "未发起"],
    [26, "9.14618 — Rubber O-Rings Seals", "CTS Table 3", "中", "向客户要",
     "O型密封圈规范", "未发起"],
    # === 低（可后续确认）===
    [27, "FMVSS 301 — Fuel System Integrity", "PF.90197 §8.1, PF.90298 §8.1", "低", "自行购买",
     "碰撞安全（STELLANTIS负责车辆级测试）", "未发起"],
    [28, "SAE J2027 — Protective Covers", "PF.90197 §1.2", "低", "自行购买",
     "保护套标准（材料选择参考）", "未发起"],
    [29, "SAE J1677 — Steel Tubing Test", "PF.90197 Table 1", "低", "自行购买",
     "钢管测试（如不含钢管则不适用）", "未发起"],
    [30, "EN 14214 — FAME for Diesel Engines", "PF.90197 §6.3.6.2", "低", "自行购买",
     "生物柴油B30参考标准", "未发起"],
    [31, "CS.00080 — CAD Drawings Rules", "CTS Table 5", "低", "向客户要",
     "CAD出图规则（列为例外项，我方用CATIA）", "未发起"],
    [32, "9.01111 — Ergonomics", "CTS §5.2", "低", "向客户要",
     "人机工程规范（辅助参考）", "未发起"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
URGENCY_FILLS = {
    "最高": PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid"),
    "高": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "中": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

HEADERS = ["序号", "缺失文档", "引用来源", "紧急度", "来源方", "影响范围", "处理状态"]
WIDTHS = [6, 35, 25, 8, 12, 30, 10]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "缺失项跟踪"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        urgency = str(ws.cell(r, 4).value or "")
        fill = URGENCY_FILLS.get(urgency)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    # 数据验证
    dv_urgency = DataValidation(type="list", formula1='"最高,高,中,低"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"未发起,已发起,已收到,不适用"', allow_blank=True)
    dv_source = DataValidation(type="list", formula1='"向客户要,自行购买,内部收集"', allow_blank=True)
    ws.add_data_validation(dv_urgency)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_source)
    last = len(DATA) + 50
    dv_urgency.add(f"D2:D{last}")
    dv_status.add(f"G2:G{last}")
    dv_source.add(f"E2:E{last}")

    # 统计区
    sr = len(DATA) + 4
    ws.cell(sr, 1, "统计").font = Font(name='Arial', bold=True, size=11)
    for i, (label, formula) in enumerate([
        ("最高", f'=COUNTIF(D2:D{len(DATA)+1},"最高")'),
        ("高", f'=COUNTIF(D2:D{len(DATA)+1},"高")'),
        ("中", f'=COUNTIF(D2:D{len(DATA)+1},"中")'),
        ("低", f'=COUNTIF(D2:D{len(DATA)+1},"低")'),
    ]):
        ws.cell(sr + 1 + i, 1, label).font = BODY_FONT
        ws.cell(sr + 1 + i, 2, formula).font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"缺失项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
