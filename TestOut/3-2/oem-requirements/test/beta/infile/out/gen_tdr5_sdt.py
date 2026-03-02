"""
TDR5 SDT团队表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr5_sdt.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR5_SDT团队表.xlsx",
}

# ============================================================
# SUPPLIER_DATA — 供应商团队（从company-profile/SDT团队表.xlsx自动填）
# 列顺序: [角色, 姓名, 职务, 电话, 邮箱, 参与阶段]
# ============================================================
SUPPLIER_DATA = [
    ["PM 项目经理", "王强", "高级项目经理", "138-0001-0001", "wang.qiang@nobelauto.com", "全程"],
    ["DE 设计工程师", "李明", "主任工程师", "138-0001-0002", "li.ming@nobelauto.com", "阶段1-3"],
    ["QE 质量工程师", "张伟", "质量主管", "138-0001-0003", "zhang.wei@nobelauto.com", "全程"],
    ["TE 测试工程师", "刘洋", "测试主管", "138-0001-0004", "liu.yang@nobelauto.com", "阶段2-5"],
    ["ME 制造工程师", "陈浩", "工艺主管", "138-0001-0005", "chen.hao@nobelauto.com", "阶段2-5"],
    ["PE 采购工程师", "赵磊", "采购主管", "138-0001-0006", "zhao.lei@nobelauto.com", "阶段1-3"],
    ["SCM 物流", "孙静", "物流经理", "138-0001-0007", "sun.jing@nobelauto.com", "阶段3-5"],
    ["质量总监", "周建国", "质量副总", "138-0001-0008", "zhou.jianguo@nobelauto.com", "关键节点"],
]

# ============================================================
# CUSTOMER_DATA — 客户对口人（部分已知，其余黄色待填）
# 列顺序: [角色, 姓名, 公司, 邮箱, 备注]
# ============================================================
CUSTOMER_DATA = [
    ["PC 采购联系人", "Yuhan Zhang", "Stellantis", "", "来源: SSTS"],
    ["PR 采购审批人", "Grace Tan", "Stellantis", "", "来源: SSTS"],
    ["SQE", "", "Stellantis", "", ""],
    ["技术对口", "", "Stellantis", "", ""],
    ["PF.90197 Author", "Fernando Sada", "Stellantis NA", "fernando.sada1@stellantis.com", "Fuel Systems"],
    ["PF.90197 Co-Author", "Bruno Le Moine", "Stellantis EMEA", "bruno.lemoine@stellantis.com", "Fuel Systems"],
]

# ============================================================
# COMM_DATA — 沟通协议
# 列顺序: [沟通类型, 频率, 参与人, 工具, 备注]
# ============================================================
COMM_DATA = [
    ["项目周会", "每周", "PM+DE+QE+ME", "Teams/腾讯会议", "周一上午"],
    ["里程碑评审", "按节点", "全员", "现场", "Gate Review"],
    ["质量问题快报", "即时", "QE→PM→客户SQE", "邮件+电话", "2小时内"],
    ["进度月报", "每月", "PM→客户PC", "邮件", "月末"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name='Arial', bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)


def write_table(ws, row, title, headers, widths, data, yellow_empty=False):
    ws.cell(row, 1, title).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(headers))
    row += 1
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, w)
    for r_data in data:
        row += 1
        for c, v in enumerate(r_data, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
            if yellow_empty and not v:
                cell.fill = INPUT_FILL
    return row + 2


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SDT团队表"
    row = 1

    row = write_table(ws, row, "一、供应商团队",
                      ["角色", "姓名", "职务", "电话", "邮箱", "参与阶段"],
                      [18, 10, 15, 16, 30, 10],
                      SUPPLIER_DATA)

    row = write_table(ws, row, "二、客户对口人",
                      ["角色", "姓名", "公司", "邮箱", "备注"],
                      [18, 10, 15, 30, 15],
                      CUSTOMER_DATA, yellow_empty=True)

    row = write_table(ws, row, "三、沟通协议",
                      ["沟通类型", "频率", "参与人", "工具", "备注"],
                      [15, 8, 25, 15, 15],
                      COMM_DATA)

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"供应商: {len(SUPPLIER_DATA)}, 客户: {len(CUSTOMER_DATA)}, 沟通: {len(COMM_DATA)}")


if __name__ == "__main__":
    main()
