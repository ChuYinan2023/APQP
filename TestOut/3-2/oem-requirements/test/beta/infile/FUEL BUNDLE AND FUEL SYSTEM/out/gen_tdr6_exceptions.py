"""
TDR6 例外事项清单生成器（基于客户模板格式）
客户模板: TDR_技术要求清单模板.xlsx (Stellantis)
增加供应商判定列和替代方案列
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR6_例外事项清单.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [Item, Reference Specification, Reference Paragraph,
#          Name, Requirement Description, Limit/Indication,
#          判定, 供应商现状, 替代方案/备注]
# 判定: "✅ 符合" / "❌ 例外" / "🔵 待确认"
# ============================================================
DATA = [
    # ===== 设计工具与数据管理 =====
    [1, "SSTS Deliverables", "Item 1",
     "3D CAD — UG/Teamcenter", "供应商须具备UG/NX + Teamcenter协同能力",
     "UG/NX + Teamcenter",
     "❌ 例外", "CATIA V5 R2021，无PLM/Teamcenter", "方案1: 采购UG NX Seat + Teamcenter许可证（约¥15万/年）；方案2: 使用Parasolid中间格式交换，协商客户接受"],
    [2, "SSTS Deliverables", "Item 8/12",
     "PLM数据管理", "CAD数据须通过Teamcenter管理和交付",
     "Teamcenter PLM",
     "❌ 例外", "文件服务器管理，无PLM系统", "随UG采购一并部署Teamcenter；或协商使用客户Teamcenter供应商门户"],

    # ===== 制造能力 =====
    [3, "CTS §3.1", "§3.1",
     "金属管内镍镀处理", "金属管内部须镍镀防腐处理",
     "镍镀层（厚度待定）",
     "🔵 待确认", "无内部电镀能力，需委外", "已有合格电镀供应商（XX电镀厂），IATF认证，可提供委外方案"],
    [4, "PF.90197", "§6.2.2",
     "多层管壁结构", "塑料管须符合多层壁结构设计要求",
     "多层管结构",
     "✅ 符合", "具备多层管挤出能力（PA12/EVOH/PA6等）", ""],
    [5, "PF.90197", "§6.3.4",
     "卡扣制造", "卡扣装配力≤40N，拆卸力≥100N",
     "装配≤40N / 拆卸≥100N",
     "✅ 符合", "注塑成型（PA66 GF30），8台注塑机", ""],
    [6, "PPTX Slide 3", "BOM",
     "Damper总成供应", "Damper工作压力4.5bar",
     "4.5 bar",
     "✅ 符合", "Nobel Auto为Damper现有供应商", "自供件，无需外购"],

    # ===== 爆破/压力测试 =====
    [7, "PF.90197", "§7.4",
     "爆破压力测试（常温）", "管路总成爆破压力≥36bar @23°C",
     "≥ 36 bar @23°C",
     "✅ 符合", "爆破测试仪 0-100 bar", ""],
    [8, "PF.90197", "§7.4",
     "爆破压力测试（高温）", "管路总成爆破压力≥13.5bar @115°C",
     "≥ 13.5 bar @115°C",
     "✅ 符合", "爆破测试仪 + 环境箱配合", ""],
    [9, "PF.90298", "§6.3.7",
     "QC接头爆破", "QC接头爆破压力≥3447kPa",
     "≥ 3447 kPa (500 psig)",
     "✅ 符合", "爆破测试仪 0-100 bar", ""],

    # ===== 泄漏测试 =====
    [10, "PF.90197", "§7.2",
     "VLD泄漏检测", "液体管VLD 15μm×3mm，测试压力≥150PSI",
     "VLD 15μm×3mm",
     "✅ 符合", "氦气检漏仪 10μm VLD", "满足精度要求"],
    [11, "PF.90298", "§7.1.1-7.1.3",
     "QC接头泄漏率", "低压/高压/真空泄漏≤0.5cc/min",
     "≤ 0.5 cc/min",
     "✅ 符合", "差压检漏仪 0.1kPa精度 + 氦检", ""],
    [12, "PF.90298", "§7.1.4",
     "QC在线检漏（100%）", "100%逐件在线泄漏检测",
     "100% inline, all pass",
     "✅ 符合", "装配线集成检漏工位，MES锁定", ""],

    # ===== 拉脱力测试 =====
    [13, "PF.90197", "§6.3.5",
     "接头拉脱力", "常温≥450N，高温115°C≥115N",
     "≥450N(RT) / ≥115N(115°C)",
     "✅ 符合", "万能材料试验机 0-50kN (Zwick)", ""],
    [14, "PF.90298", "§7.3",
     "QC拉脱力", "未暴露≥450N，已暴露燃油≥297N",
     "≥450N / ≥297N",
     "✅ 符合", "万能材料试验机 0-50kN (Zwick)", ""],

    # ===== 耐久测试 =====
    [15, "PF.90197", "§9.5",
     "脉冲压力耐久（P5等级）", "0~2×工作压力，300K/600K次",
     "P5: 300,000/600,000 cycles",
     "✅ 符合", "脉冲测试台 0-40bar 0-4Hz，4工位", ""],
    [16, "PF.90197", "§9.4",
     "振动耐久测试", "振动循环测试 45件 R95C90",
     "振动谱（详见CD.80064）",
     "❌ 例外", "无振动测试台", "委外至国家级实验室（CATARC/中汽中心），已有合作关系"],
    [17, "PF.90197", "§9.3",
     "寿命循环测试", "15年/150,000miles寿命验证",
     "Life cycle test 45件 R95C90",
     "✅ 符合", "脉冲台+环境箱组合验证", ""],

    # ===== 环境/腐蚀测试 =====
    [18, "PF.90197", "§5.3",
     "耐腐蚀测试", "SAE J2334循环腐蚀 + 盐雾",
     "循环腐蚀 + 盐雾",
     "✅ 符合", "盐雾箱 480h能力", ""],
    [19, "PF.90197", "§5.4",
     "氯化物应力腐蚀", "CaCl₂ 50% @80°C（EMEA） + ZnCl₂",
     "CaCl₂/ZnCl₂",
     "🔵 待确认", "盐雾箱可做基础测试，CaCl₂专项需确认方法", "需购置CaCl₂浸泡夹具，或委外"],
    [20, "PF.90197", "§5.1",
     "耐化学介质", "柴油+B7/B30+自氧化燃油浸泡",
     "Diesel + Bio-diesel + Sour gas",
     "✅ 符合", "燃油浸泡箱 E10/E85/B20能力", "需确认B30燃油供应"],
    [21, "PF.90197", "§5.2",
     "温度循环测试", "-40°C~180°C温度循环",
     "-50°C ~ 180°C",
     "✅ 符合", "环境箱 -50°C~180°C × 2台", ""],

    # ===== 清洁度 =====
    [22, "PF.90197", "§7.1",
     "清洁度检测", "≤1.5mg/dm²，颗粒分级 ISO 16232",
     "颗粒分级 Table 6",
     "✅ 符合", "清洁度分析仪 ISO 16232", ""],
    [23, "PF.90298", "§6.2.4",
     "QC接头清洁度", "≤0.75mg/dm²",
     "≤ 0.75 mg/dm²",
     "✅ 符合", "清洁度分析仪 ISO 16232", ""],

    # ===== 流量/压降 =====
    [24, "PF.90197", "§7.3",
     "流量/压降测试", "400kPa，成形截面≥80%",
     "400 kPa, ≥80% cross-section",
     "✅ 符合", "流量计 0-200L/h", ""],

    # ===== 特殊测试 =====
    [25, "PF.90197", "§8.2",
     "MiniSHED蒸发排放", "MiniSHED测试 DV P99C90 / PV R95C90",
     "LP.7A005",
     "❌ 例外", "无MiniSHED测试设备", "委外至第三方实验室（SGS/TÜV），或采购MiniSHED设备（约¥80万）"],
    [26, "PF.90298", "§6.4.1",
     "ESD静电放散", "电阻≤1,000,000Ω",
     "≤ 1MΩ",
     "🔵 待确认", "无专用ESD测试仪，需采购", "采购绝缘电阻测试仪（约¥2万），技术上可行"],
    [27, "PF.90197", "§6.2.5",
     "静电荷消散", "静电消散能力（LP.7A004）",
     "LP.7A004",
     "🔵 待确认", "需获取LP.7A004确认测试方法", "测试方法依赖缺失文档LP.7A004"],
    [28, "PF.90298", "§6.3.5",
     "动态冲击测试", "金属≥50J，塑料≥10J",
     "≥50J(metal) / ≥10J(plastic)",
     "🔵 待确认", "万能试验机可改造做冲击，或委外", "需采购冲摆锤冲击仪（约¥5万），或委外"],
    [29, "PF.90197", "§5.5.2",
     "碎石冲击（SAE J400）", "碎石冲击试验",
     "SAE J400",
     "❌ 例外", "无碎石冲击设备", "委外至第三方实验室"],
    [30, "PF.90298", "§6.2.1.2",
     "DSC熔点测试", "QC材料DSC测试（PV阶段）",
     "DSC",
     "❌ 例外", "无DSC设备", "委外至材料实验室（约¥500/次）"],
    [31, "PF.90197", "§5.8",
     "防火/火焰蔓延测试", "FMVSS 301（如适用）",
     "FMVSS 301",
     "🔵 待确认", "无防火测试能力", "需确认是否适用（EMEA市场可能不要求FMVSS 301）；如需要则委外"],

    # ===== 尺寸/外观 =====
    [32, "PF.90197", "§6.3.8",
     "尺寸测量", "关键尺寸三坐标测量",
     "CMM 0.003mm",
     "✅ 符合", "三坐标测量机 Hexagon 0.003mm", ""],
    [33, "PF.90197", "§6.1",
     "外观检查", "100%目视检查",
     "100% visual",
     "✅ 符合", "检验工位 + 标准样件", ""],

    # ===== 材料力学 =====
    [34, "PF.90197", "Table 3",
     "材料拉伸测试", "新件/燃油浸泡后/热老化后 拉伸性能",
     "≥20/15/15 N/mm²",
     "✅ 符合", "万能材料试验机 Zwick + 环境箱", ""],

    # ===== 质量体系 =====
    [35, "SSTS", "Quality",
     "IATF 16949认证", "供应商须IATF 16949认证",
     "IATF 16949:2016",
     "✅ 符合", "IATF 16949认证，有效期至2027-06", ""],
    [36, "PF.90197", "§3.1",
     "材料/过程认可（QR.00001 GPAT）", "材料和过程变更须GPAT审批",
     "QR.00001",
     "✅ 符合", "有GPAT流程经验（Stellantis现有供应商）", ""],
    [37, "SSTS", "Deliverables",
     "年度自认证（SVP）", "每年重复执行验证计划",
     "Annual SVP",
     "✅ 符合", "已有年度验证计划流程", ""],
    [38, "CS.00265", "全文",
     "有害物质管控", "REACH/RoHS/IMDS合规",
     "CS.00265",
     "✅ 符合", "IMDS已注册，材料合规管理流程完善", ""],

    # ===== 产能 =====
    [39, "SSTS", "Volume",
     "产能满足", "年产能须满足项目需求（待确认具体数量）",
     "待客户提供年需求量",
     "🔵 待确认", "装配线500K pcs/yr，注塑500K pcs/yr", "需客户提供年需求量以确认"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
EXCEPTION_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TBD_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

# 客户模板列 + 供应商判定列
HEADERS = [
    "Item",
    "Reference\nSpecification",
    "Reference\nParagraph",
    "Name\n(test/requirement/standard)",
    "Requirement\nDescription",
    "Limit /\nIndication",
    "判定\nJudgment",
    "供应商现状\nSupplier Status",
    "替代方案/备注\nAlternative / Remarks",
]
WIDTHS = [6, 18, 12, 28, 35, 22, 12, 30, 35]

JUDGMENT_FILLS = {
    "✅ 符合": OK_FILL,
    "❌ 例外": EXCEPTION_FILL,
    "🔵 待确认": TBD_FILL,
}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDR6 Exception List"

    # Header
    ws.append(HEADERS)
    ws.row_dimensions[1].height = 40
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    # Data
    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        judgment = str(ws.cell(r, 7).value or "")
        fill = JUDGMENT_FILLS.get(judgment)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # 判定列着色
        if fill:
            ws.cell(r, 7).fill = fill

    # 统计 Sheet
    ws2 = wb.create_sheet("统计 Summary")
    stats = [
        ["判定", "数量"],
        ["✅ 符合", sum(1 for r in DATA if r[6] == "✅ 符合")],
        ["❌ 例外", sum(1 for r in DATA if r[6] == "❌ 例外")],
        ["🔵 待确认", sum(1 for r in DATA if r[6] == "🔵 待确认")],
        ["合计", len(DATA)],
    ]
    for i, row in enumerate(stats):
        for j, val in enumerate(row):
            cell = ws2.cell(i + 1, j + 1, val)
            cell.font = HEADER_FONT if i == 0 else BODY_FONT
            if i == 0:
                cell.fill = HEADER_FILL
            cell.border = THIN
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 10

    # 例外项汇总
    ws2.cell(7, 1, "例外项汇总").font = Font(name='Arial', bold=True, size=12)
    exc_row = 8
    ws2.cell(exc_row, 1, "Item").font = Font(name='Arial', bold=True, size=10)
    ws2.cell(exc_row, 2, "Name").font = Font(name='Arial', bold=True, size=10)
    ws2.cell(exc_row, 3, "替代方案").font = Font(name='Arial', bold=True, size=10)
    ws2.column_dimensions['C'].width = 50
    exc_row += 1
    for r in DATA:
        if r[6] in ("❌ 例外", "🔵 待确认"):
            ws2.cell(exc_row, 1, r[0]).font = BODY_FONT
            ws2.cell(exc_row, 2, r[3]).font = BODY_FONT
            ws2.cell(exc_row, 3, r[8]).font = BODY_FONT
            fill = EXCEPTION_FILL if r[6] == "❌ 例外" else TBD_FILL
            for c in range(1, 4):
                ws2.cell(exc_row, c).fill = fill
                ws2.cell(exc_row, c).border = THIN
            exc_row += 1

    wb.save(CONFIG["输出路径"])
    ok = sum(1 for r in DATA if r[6] == "✅ 符合")
    exc = sum(1 for r in DATA if r[6] == "❌ 例外")
    tbd = sum(1 for r in DATA if r[6] == "🔵 待确认")
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"总计: {len(DATA)} 条 — ✅符合:{ok} | ❌例外:{exc} | 🔵待确认:{tbd}")


if __name__ == "__main__":
    main()
