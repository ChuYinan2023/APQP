"""
TDR8 补充材料/零部件清单生成器 — 使用客户模板格式
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CONFIG = {"输出路径": "TDR8_补充材料清单.xlsx"}

# 客户模板列: Item, Component, q.ty, Note
HEADERS = ["Item", "Component", "q.ty", "Note"]
WIDTHS = [8, 50, 10, 35]

# BOM from README (CTS + PPTX data)
PARTS = [
    [1, "QC Female 90° SAE 5/16\" (Filter side)", 1, "PF.90298, 滤清器侧快插接头"],
    [2, "QC Female 90° SAE 3/8\" w/ sec-lock (Engine side)", 1, "PF.90298, 发动机侧快插接头+二次锁扣"],
    [3, "Plastic pipe (Feed) Ø8mm×6mm", 1, "PF.90197, PA12多层管"],
    [4, "Damper 52183408 (4.5 bar)", 1, "Nobel Auto自供, 含Capot+Spring+Plug+Membrane+Body"],
    [5, "  ├─ Capot (端盖)", 1, "PA66+PA6"],
    [6, "  ├─ Spring compression 4.5 bar", 1, "NF EN 10270-3 弹簧钢丝"],
    [7, "  ├─ Plug (堵头)", 1, "PA66 GF30"],
    [8, "  ├─ Membrane (膜片)", 1, "54U6002+AgN91S"],
    [9, "  └─ Damper body 8×10", 1, "PA66 GF30"],
    [10, "Anti-abrasion sleeve (防磨套管)", 1, "Braided sleeve CPN 5302 (Meihe), SP.91220/02"],
    [11, "Clips / Brackets (卡扣/支架)", "TBD", "待3D数据确认数量和类型"],
    [12, "Protective caps (管端保护帽)", "TBD", "CTS §3.7要求所有管端需保护帽"],
]

HEADER_FONT = Font(name='Cambria', bold=True, size=12)
HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
BODY_FONT = Font(name='Arial', size=10)
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical='top')


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parts List"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 25
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[chr(64 + c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in PARTS:
        ws.append(row)
    for r in range(2, len(PARTS) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # TBD items highlight yellow
        qty_val = str(ws.cell(r, 3).value or "")
        if qty_val == "TBD":
            for c in range(1, len(HEADERS) + 1):
                ws.cell(r, c).fill = YELLOW_FILL

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"零部件: {len(PARTS)} 项")


if __name__ == "__main__":
    main()
