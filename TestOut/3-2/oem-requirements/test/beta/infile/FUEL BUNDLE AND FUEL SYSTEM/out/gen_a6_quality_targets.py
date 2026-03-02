"""
A6质量目标表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_a6_quality_targets.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "A6_质量目标表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 指标名称, 目标值, 来源文件, 验证方法, 备注]
# 类别: 市场质量目标 / 可靠性目标 / 过程能力 / 清洁度目标 / 安全法规 / 环保合规 / 持续符合性
# 同类别的行会自动合并单元格
# ============================================================
DATA = [
    # ===== 市场质量目标 =====
    ["市场质量目标", "CCP目标（Customer Car Profile）", "供应商须承诺达成CCP目标", "SSTS", "CCP跟踪系统", "具体数值由客户定义"],
    ["市场质量目标", "QT目标（Quality Tracking）", "供应商须承诺达成QT目标", "SSTS", "QT跟踪系统", "具体数值由客户定义"],
    ["市场质量目标", "ICP目标（Initial Customer Perception）", "0（单个零部件）", "SSTS", "ICP评审", ""],

    # ===== 可靠性目标 =====
    ["可靠性目标", "TESIS可靠性（车底 3M）", "0.00 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（车底 12M）", "0.035 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（车底 24M）", "0.089 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（车底 36M）", "0.120 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（舱内 3M）", "0.00 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（舱内 12M）", "0.065 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（舱内 24M）", "0.160 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "TESIS可靠性（舱内 36M）", "0.365 c/1000", "SSTS", "TESIS系统跟踪", ""],
    ["可靠性目标", "功能测试可靠性", "P99C90（99%可靠度，90%置信度）", "PF.90197 Annex A", "DV/PV功能测试（15件/项）", "适用于全部功能类测试"],
    ["可靠性目标", "耐久测试可靠性", "R95C90（95%可靠度，90%置信度）", "PF.90197 Annex A", "DV/PV耐久测试（45件/项）", "适用于寿命/振动/脉冲等"],
    ["可靠性目标", "设计寿命", "15年 / 150,000 miles", "PF.90197 §9", "寿命循环+振动+脉冲综合验证", ""],
    ["可靠性目标", "脉冲耐久（P5等级）", "300,000次(S&S) / 600,000次(S&S+FW), 0~2×工作压力", "PF.90197 §9.5", "脉冲压力测试 45件 R95C90", "2.2L柴油4.5bar"],

    # ===== 清洁度目标 =====
    ["清洁度目标", "管路总成颗粒总量", "≤ 1.5 mg/dm²", "PF.90197 §7.1", "清洁度检测 15件 P99C90", ""],
    ["清洁度目标", "Diesel Filter A下游 50-100μm", "< 50颗粒", "PF.90197 Table 6", "颗粒计数", ""],
    ["清洁度目标", "Diesel Filter A下游 100-200μm", "< 15颗粒", "PF.90197 Table 6", "颗粒计数", ""],
    ["清洁度目标", "Diesel Filter A下游 200-500μm", "< 2颗粒", "PF.90197 Table 6", "颗粒计数", ""],
    ["清洁度目标", "Diesel Filter A下游 ≥500μm", "0颗粒", "PF.90197 Table 6", "颗粒计数", "零容忍"],
    ["清洁度目标", "QC接头内部清洁度", "≤ 0.75 mg/dm²", "PF.90298 §6.2.4", "清洁度检测 15件 P99C90", ""],

    # ===== 安全法规 =====
    ["安全法规", "ESD静电放散电阻", "≤ 1,000,000 Ω", "PF.90298 §6.4.1", "ESD电阻测试", "防止燃油起火"],
    ["安全法规", "静电荷消散能力", "符合LP.7A004要求", "PF.90197 §6.2.5", "静电消散测试", ""],
    ["安全法规", "防火安全", "符合FMVSS 301（如适用北美市场）", "PF.90197 §5.8", "火焰蔓延/防火测试", "法规强制"],
    ["安全法规", "有害物质管控", "符合CS.00265 Substances of Concern", "PF.90197, CS.00265", "IMDS提交 + 材料检测", "REACH/RoHS等"],

    # ===== 环保合规 =====
    ["环保合规", "蒸发排放（MiniSHED）", "DV: P99C90 / PV: R95C90", "PF.90197 §8.2", "MiniSHED测试 15件", "LP.7A005"],
    ["环保合规", "排放件标识", "零件需标注排放件标识", "PF-EMISSIONS", "目视检查", ""],

    # ===== 持续符合性 =====
    ["持续符合性", "年度自认证（Self Validation Plan）", "每年重复执行验证计划", "SSTS", "年度SVP执行 + 报告提交", ""],
    ["持续符合性", "最低样件数", "按PF.90068 Annex A，需与Specialist协商", "SSTS", "与Stellantis Specialist确认", ""],
    ["持续符合性", "材料/过程变更认可", "任何变更需QR.00001 GPAT审批", "PF.90197 §3.1", "GPAT流程", ""],

    # ===== 过程能力 =====
    ["过程能力", "关键尺寸Cpk", "Cpk ≥ 1.67（CC特性）", "Stellantis通用要求", "SPC统计 + 三坐标测量", "CC特性初始能力"],
    ["过程能力", "重要尺寸Cpk", "Cpk ≥ 1.33（SC特性）", "Stellantis通用要求", "SPC统计", "SC特性持续能力"],
    ["过程能力", "拉脱力过程能力", "Cpk ≥ 1.67", "PF.90197 §6.3.5", "在线拉脱力SPC", "CC特性"],
    ["过程能力", "泄漏检测过程能力", "100%在线检测，0缺陷流出", "PF.90197 §7.2, PF.90298 §7.1.4", "在线检漏系统", "CC特性"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "指标名称", "目标值", "来源文件", "验证方法", "备注"]
WIDTHS = [18, 25, 30, 20, 20, 20]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质量目标表"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        cat_cell = ws.cell(r, 1)
        cat_cell.font = CAT_FONT
        cat_cell.fill = CAT_FILL

    # 合并同类别单元格
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    wb.save(CONFIG["输出路径"])
    cats = len(set(r[0] for r in DATA)) if DATA else 0
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"质量指标: {len(DATA)} 条, {cats} 个类别")


if __name__ == "__main__":
    main()
