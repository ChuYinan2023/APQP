"""
TDR5 SDT团队表生成器 — 使用客户模板格式
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CONFIG = {"输出路径": "TDR5_SDT团队表.xlsx"}

# 客户模板列: Item, NAME, LAST NAME, ROLE, EMAIL
HEADERS = ["Item", "NAME", "LAST NAME", "ROLE", "EMAIL"]
WIDTHS = [8, 15, 15, 25, 35]

# company-profile/SDT团队表.xlsx 数据
TEAM = [
    [1, "Qiang", "Wang", "PM - Project Manager", "wang.qiang@nobelauto.com"],
    [2, "Ming", "Li", "DE - Design Engineer", "li.ming@nobelauto.com"],
    [3, "Wei", "Zhang", "QE - Quality Engineer", "zhang.wei@nobelauto.com"],
    [4, "Yang", "Liu", "TE - Test Engineer", "liu.yang@nobelauto.com"],
    [5, "Hao", "Chen", "ME - Manufacturing Engineer", "chen.hao@nobelauto.com"],
    [6, "Lei", "Zhao", "PE - Purchasing Engineer", "zhao.lei@nobelauto.com"],
    [7, "Jing", "Sun", "SCM - Logistics Manager", "sun.jing@nobelauto.com"],
    [8, "Jianguo", "Zhou", "Quality Director (VP)", "zhou.jianguo@nobelauto.com"],
]

HEADER_FONT = Font(name='Cambria', bold=True, size=12)
HEADER_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
BODY_FONT = Font(name='Arial', size=10)
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical='center')


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SDT Team"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 25
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[chr(64 + c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in TEAM:
        ws.append(row)
    for r in range(2, len(TEAM) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"团队成员: {len(TEAM)} 人")


if __name__ == "__main__":
    main()
