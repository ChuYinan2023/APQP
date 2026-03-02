"""
TDR7 RASI矩阵生成器 — 复制客户模板，标注适用SI等级
SSTS标注 Supplier Integration Level 2，对应模板中 SI Level II 列 (H-J)
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

TEMPLATE = "_嵌入文件/TDR_RASI责任矩阵.xlsx"
OUTPUT = "TDR7_RASI矩阵.xlsx"


def main():
    # 复制客户模板
    shutil.copy2(TEMPLATE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)

    ws = wb["Medium-Level RASI"]

    # 在模板顶部添加项目适用说明
    # 插入说明到空白区域（右侧列）
    highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    note_font = Font(name='Arial', bold=True, size=11, color='C00000')
    info_font = Font(name='Arial', size=10)

    # 在V列(col 22)写项目说明
    ws.cell(2, 22, "PROJECT: KP1 A&B — FC00SAA78530 Fuel Supply Line").font = note_font
    ws.cell(3, 22, "SUPPLIER: Nobel Auto (诺贝尔汽车零部件有限公司)").font = info_font
    ws.cell(4, 22, "SI LEVEL: Level II (as per SSTS)").font = note_font
    ws.cell(5, 22, "适用列: H (Fiat) / I (Liv. 1 供应商) / J (Liv. strategico 2)").font = info_font
    ws.cell(6, 22, "本矩阵已按客户标准模板预填，供应商确认SI Level II列的R/A/S/I分配。").font = info_font

    ws.column_dimensions['V'].width = 60

    # 高亮适用的SI Level II列 (H, I, J) 表头
    level2_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for col in [8, 9, 10]:  # H, I, J
        for row in [8, 9]:
            cell = ws.cell(row, col)
            cell.fill = level2_fill

    # 添加说明sheet
    ws_note = wb.create_sheet("说明 Notes", 0)
    notes = [
        ["TDR7 RASI责任矩阵 — 使用说明"],
        [""],
        ["项目", "KP1 A&B — FC00SAA78530 Fuel Supply Line Filter to Engine"],
        ["OEM", "Stellantis (原FCA)"],
        ["供应商", "Nobel Auto (诺贝尔汽车零部件有限公司)"],
        ["SI等级", "Level II (SSTS标注)"],
        [""],
        ["本RASI矩阵为Stellantis标准模板，包含214行中等级别任务，分12个大类。"],
        [""],
        ["适用列说明:"],
        ["", "SI Level II (列H-J) 是本项目适用的责任分配列"],
        ["", "H = Fiat/Stellantis (OEM) 的职责"],
        ["", "I = Liv. 1 (一级供应商 = Nobel Auto) 的职责"],
        ["", "J = Liv. strategico 2 (战略二级供应商) 的职责"],
        [""],
        ["R/A/S/I 含义:"],
        ["", "R = Responsible (负责执行)"],
        ["", "A = Approval (审批权)"],
        ["", "S = Support (提供支持)"],
        ["", "I = Informed (需被通知)"],
        [""],
        ["⚠️ 供应商需确认:"],
        ["", "1. 列I中标记为R的任务，确认有能力承担"],
        ["", "2. 如有异议，标注为例外项并在TDR6中体现"],
        ["", "3. 二级供应商(列J)的分配需根据实际委外情况调整"],
    ]
    note_header = Font(name='Arial', bold=True, size=14)
    for i, row in enumerate(notes):
        for j, val in enumerate(row):
            cell = ws_note.cell(i + 1, j + 1, val)
            if i == 0:
                cell.font = note_header
            else:
                cell.font = info_font
    ws_note.column_dimensions['A'].width = 15
    ws_note.column_dimensions['B'].width = 70

    wb.save(OUTPUT)
    print(f"已保存: {OUTPUT}")
    print("基于客户标准RASI模板（214行×12类），标注SI Level II适用")
    print("⚠️ 供应商需确认列I(Liv.1)中所有R标记任务的可执行性")


if __name__ == "__main__":
    main()
