"""
TDR3 CDS开发进度表生成器
无客户模板，按Stellantis标准里程碑格式生成
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG = {"输出路径": "TDR3_CDS开发表.xlsx"}

HEADERS = ["#", "里程碑\nMilestone", "活动\nActivity", "负责方\nOwner",
           "计划开始\nPlan Start", "计划完成\nPlan End", "状态\nStatus", "备注\nRemarks"]
WIDTHS = [5, 22, 35, 15, 14, 14, 10, 30]

# 标准Stellantis APQP开发里程碑 + 项目特定活动
DATA = [
    # Phase 0 — Sourcing
    [1, "TKO (Tool Kick Off)", "合同签订/项目启动", "PM / PC", "", "", "黄色待填", "SSTS Deliverables时间节点"],
    [2, "TDR文件提交", "提交全套TDR1-8报价文件", "PM", "", "", "黄色待填", "含本文件"],
    [3, "SDT团队确认", "确认供应商开发团队名单", "PM", "", "", "黄色待填", "TDR5"],
    [4, "RASI矩阵签署", "确认双方责任分工", "PM / PC", "", "", "黄色待填", "TDR7"],
    [5, "例外事项确认", "确认例外项替代方案", "QE / PM", "", "", "黄色待填", "TDR6"],

    # Phase 1 — Concept
    [6, "3D CAD概念设计", "初步3D模型（Soft Tooling阶段）", "DE", "", "", "黄色待填", "需解决UG/Teamcenter问题"],
    [7, "CAE虚拟分析", "流体/结构/热分析", "DE", "", "", "黄色待填", "ANSYS 2023R1"],
    [8, "DFMEA启动", "产品FMEA初版", "DE / QE", "", "", "黄色待填", "APIS IQ 8.0"],
    [9, "供应商选定（委外）", "确定镀镍/振动测试/MiniSHED委外供应商", "PE", "", "", "黄色待填", "例外项对应"],

    # Phase 2 — Development
    [10, "完整2D图纸", "含GD&T的正式工程图纸", "DE", "", "", "黄色待填", "by TKO（SSTS Deliverables #9）"],
    [11, "完整3D CAD", "最终3D模型（含所有子零件）", "DE", "", "", "黄色待填", "by TKO（SSTS Deliverables #12）"],
    [12, "PFMEA", "过程FMEA", "ME / QE", "", "", "黄色待填", "by TKO（SSTS Deliverables #11）"],
    [13, "模具/工装制作", "注塑模具+弯管夹具+装配工装+检具", "ME / PE", "", "", "黄色待填", ""],
    [14, "控制计划", "预生产+量产控制计划", "QE / ME", "", "", "黄色待填", ""],

    # Phase 3 — DV验证
    [15, "DV样件制作", "DV验证样件生产", "ME", "", "", "黄色待填", "数量见PF.90197 Annex A"],
    [16, "DV测试执行", "全部DV测试项（功能15件P99C90，耐久45件R95C90）", "TE", "", "", "黄色待填", "含委外项目"],
    [17, "DV测试报告", "DV测试结果汇总报告", "TE / QE", "", "", "黄色待填", ""],

    # Phase 4 — PV验证
    [18, "PV样件制作", "PV验证样件（量产工装）", "ME", "", "", "黄色待填", ""],
    [19, "PV测试执行", "全部PV测试项", "TE", "", "", "黄色待填", ""],
    [20, "PV测试报告", "PV测试结果汇总报告", "TE / QE", "", "", "黄色待填", ""],
    [21, "PPAP提交", "PPAP全套文件提交（等级待确认）", "QE / PM", "", "", "黄色待填", ""],

    # Phase 5 — Launch
    [22, "PSO (Process Sign Off)", "过程签核", "QE / ME", "", "", "黄色待填", ""],
    [23, "SOP支持", "总装厂SOP现场支持", "ME / QE", "", "", "黄色待填", "SSTS Deliverables #15"],
    [24, "量产启动", "正式量产开始", "PM", "", "", "黄色待填", ""],
    [25, "年度SVP", "首次年度自认证验证计划", "QE / TE", "", "", "黄色待填", "SSTS要求每年执行"],
]

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PHASE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BODY_FONT = Font(name='Arial', size=10)
PHASE_FONT = Font(name='Arial', bold=True, size=10)
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Phase row indices (insert before these DATA indices)
PHASES = [
    (0, "Phase 0 — Sourcing（报价/定点）"),
    (5, "Phase 1 — Concept（概念设计）"),
    (9, "Phase 2 — Development（开发）"),
    (14, "Phase 3 — DV Validation（设计验证）"),
    (17, "Phase 4 — PV Validation（生产验证）"),
    (21, "Phase 5 — Launch（量产启动）"),
]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CDS Development Schedule"

    # Header
    ws.append(HEADERS)
    ws.row_dimensions[1].height = 35
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    # Build rows with phase headers
    phase_idx = 0
    data_idx = 0
    row_num = 2
    phase_rows = []

    while data_idx < len(DATA):
        # Check if we need a phase header
        if phase_idx < len(PHASES) and PHASES[phase_idx][0] == data_idx:
            phase_label = PHASES[phase_idx][1]
            ws.cell(row_num, 1, "").border = THIN
            ws.merge_cells(start_row=row_num, end_row=row_num, start_column=2, end_column=len(HEADERS))
            cell = ws.cell(row_num, 2, phase_label)
            cell.font = PHASE_FONT
            cell.fill = PHASE_FILL
            cell.alignment = WRAP
            for c in range(1, len(HEADERS) + 1):
                ws.cell(row_num, c).border = THIN
                ws.cell(row_num, c).fill = PHASE_FILL
            phase_rows.append(row_num)
            row_num += 1
            phase_idx += 1

        # Write data row
        row = DATA[data_idx]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row_num, c, val)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # Yellow fill for date and status columns (待填)
        for c in [5, 6, 7]:  # Plan Start, Plan End, Status
            ws.cell(row_num, c).fill = YELLOW_FILL
        data_idx += 1
        row_num += 1

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"里程碑: {len(DATA)} 项, {len(PHASES)} 个阶段")


if __name__ == "__main__":
    main()
