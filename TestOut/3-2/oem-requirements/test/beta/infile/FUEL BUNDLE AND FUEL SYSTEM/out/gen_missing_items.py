"""
缺失项跟踪表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_missing_items.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "缺失项跟踪表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [序号, 缺失文档, 引用来源, 紧急度, 来源方, 影响范围, 处理状态]
# 紧急度: 最高/高/中/低
# 来源方: 向客户要/自行购买/内部收集
# 处理状态: 未发起/已发起/已收到/不适用
# ============================================================
DATA = [
    # ===== 最高（阻塞）=====
    [1, "3D CAD 模型（UG/Teamcenter格式）", "TDR交付物要求 TDR#4, CTS §图纸要求", "最高", "内部收集", "无法完成TDR#4 3D模型交付，无法验证接口尺寸", "未发起"],
    [2, "2D 工程图纸（燃油管路总成）", "CTS §图纸引用, PF.90197 Annex A", "最高", "向客户要", "无法确认关键尺寸公差、装配接口、材料标注", "未发起"],
    [3, "阻尼器（Damper）详细图纸/规格", "PF.90197 §5.7 Damper测试, CTS BOM", "最高", "向客户要", "阻尼器为BOM核心件，无图纸无法报价和测试", "未发起"],

    # ===== 高（TKO前需要）=====
    [4, "07740 — Stellantis标准件规范", "PF.90197 §1 Scope", "高", "向客户要", "燃油管路标准件选型依据缺失", "未发起"],
    [5, "QR.00001 — 材料与过程认可程序", "PF.90197 §3.1 Material Approval", "高", "向客户要", "材料认可流程不明，影响供应商选择", "未发起"],
    [6, "SAE J2044 — Quick Connect接头标准", "PF.90197 §5.1, PF.90298 §1 Scope", "高", "自行购买", "快接接头设计和测试标准缺失", "未发起"],
    [7, "SAE J2045 — Quick Connect性能标准", "PF.90298 §1 Scope, §5 Tests", "高", "自行购买", "快接接头性能验收标准缺失", "未发起"],
    [8, "SAE J2260 — Non-metallic Fuel Line标准", "PF.90197 §1 Scope, §5.3 Permeation", "高", "自行购买", "非金属管材渗透性测试标准缺失", "未发起"],
    [9, "PRO.00004 — Stellantis过程认可程序", "PF.90197 §3.1", "高", "向客户要", "过程认可要求不明", "未发起"],
    [10, "PPAP等级确认（客户明确）", "SSTS Deliverables, TDR交付物要求", "高", "向客户要", "不确定PPAP提交等级，影响样件和文件准备", "未发起"],
    [11, "CS.00251 — Stellantis清洁度规范", "PF.90197 §5.11 Cleanliness", "高", "向客户要", "清洁度验收标准缺失，影响产线规划", "未发起"],
    [12, "SD-11597 — 发动机罩下温度场规范", "PF.90197 §5.4 Thermal", "高", "向客户要", "热环境温度分布数据缺失，影响耐热设计", "未发起"],
    [13, "CD.80064 — Stellantis耐久测试标准", "PF.90197 §5.6 Durability", "高", "向客户要", "耐久测试振动谱/循环数缺失", "未发起"],

    # ===== 中（DV前到位即可）=====
    [14, "SAE J2334 — 循环腐蚀测试方法", "PF.90197 §5.5 Corrosion", "中", "自行购买", "腐蚀测试方法参考", "未发起"],
    [15, "SAE J400 — 碎石冲击测试", "PF.90197 §5.5.2 Chipping", "中", "自行购买", "碎石冲击测试参数", "未发起"],
    [16, "SAE J1645 — 燃油管总成渗透测试", "PF.90197 §5.3 Permeation", "中", "自行购买", "渗透测试详细方法", "未发起"],
    [17, "DIN 51604 — 参考燃油规范", "PF.90197 §4.2 Test Fluids", "中", "自行购买", "测试用燃油配方", "未发起"],
    [18, "EN ISO 527-2 — 塑料拉伸测试", "PF.90197 §5.2 Mechanical", "中", "自行购买", "管材拉伸测试方法", "未发起"],
    [19, "DIN 53504 — 橡胶拉伸测试", "PF.90197 §5.2 Mechanical", "中", "自行购买", "橡胶件拉伸测试方法", "未发起"],
    [20, "EN 14214 — 生物柴油规范", "PF.90197 §4.2 Test Fluids", "中", "自行购买", "生物柴油兼容性测试", "未发起"],
    [21, "LP.7A004 — Stellantis实验室测试程序", "PF.90197 §5.4 Thermal Aging", "中", "向客户要", "热老化测试详细程序", "未发起"],
    [22, "LP.7A005 — Stellantis实验室测试程序", "PF.90197 §5.5 Corrosion", "中", "向客户要", "腐蚀测试详细程序", "未发起"],
    [23, "PS.50005 / PS.50006 — 橡胶材料规范", "PF.90197 §4.1 Materials", "中", "向客户要", "橡胶材料牌号及性能要求", "未发起"],
    [24, "PS-8688 — 管材材料规范", "PF.90197 §4.1 Materials", "中", "向客户要", "管材材料规范", "未发起"],
    [25, "9.02145/02 — Stellantis内部标准", "PF.90197 §5.8 Fire", "中", "向客户要", "防火测试程序", "未发起"],
    [26, "QR-10012 — 质量报告模板", "PF.90197 §6 Quality", "中", "向客户要", "质量报告格式要求", "未发起"],
    [27, "7.G0200 — 包装物流规范", "SSTS Deliverables", "中", "向客户要", "包装和物流要求", "未发起"],
    [28, "CS-9003 — 供应商质量手册", "PF.90197 §6 Quality", "中", "向客户要", "供应商质量管理要求", "未发起"],
    [29, "MS.50017 / MS.50015 — 金属材料规范", "PF.90197 §4.1 Materials", "中", "向客户要", "金属接头材料性能要求", "未发起"],
    [30, "FMVSS 301 — 联邦燃油系统安全标准", "PF.90197 §5.8 Fire/Safety", "中", "自行购买", "法规安全要求（如适用北美市场）", "未发起"],

    # ===== 低（可后续确认）=====
    [31, "SAE J2027 — 燃油管材料标准", "PF.90197 §4.1 Materials", "低", "自行购买", "管材料补充参考", "未发起"],
    [32, "SAE J1677 — 燃油管测试标准", "PF.90197 §5 Tests", "低", "自行购买", "补充测试方法参考", "未发起"],
    [33, "7.N0007 — Stellantis噪声规范", "PF.90197 §5.9 NVH", "低", "向客户要", "NVH要求（如适用）", "未发起"],
    [34, "DS-107 — 设计标准", "PF.90197 §3 Design", "低", "向客户要", "通用设计标准参考", "未发起"],
    [35, "CS.CCP-ROUTING — 管路布置规范", "PF.90197 §3.2 Routing", "低", "向客户要", "管路布置参考（需3D数据配合）", "未发起"],
    [36, "PF.90271 — 相关性能规范", "PF.90197 §1 Scope引用", "低", "向客户要", "关联规范，适用性待确认", "未发起"],
    [37, "SD-M0008/03 — 系统设计规范", "PF.90197 §3 Design", "低", "向客户要", "系统级设计参考", "未发起"],
    [38, "PS-11346 — 材料规范", "PF.90197 §4.1 Materials", "低", "向客户要", "补充材料规范", "未发起"],
    [39, "CDS开发进度模板（客户版）", "SSTS Deliverables, TDR#3", "低", "向客户要", "如客户有指定CDS模板则使用客户版", "未发起"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
URGENCY_FILLS = {
    "最高": PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid"),
    "高": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "中": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

HEADERS = ["序号", "缺失文档", "引用来源", "紧急度", "来源方", "影响范围", "处理状态"]
WIDTHS = [6, 35, 25, 8, 12, 30, 10]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "缺失项跟踪"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        urgency = str(ws.cell(r, 4).value or "")
        fill = URGENCY_FILLS.get(urgency)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    # 数据验证
    dv_urgency = DataValidation(type="list", formula1='"最高,高,中,低"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"未发起,已发起,已收到,不适用"', allow_blank=True)
    dv_source = DataValidation(type="list", formula1='"向客户要,自行购买,内部收集"', allow_blank=True)
    ws.add_data_validation(dv_urgency)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_source)
    last = len(DATA) + 50
    dv_urgency.add(f"D2:D{last}")
    dv_status.add(f"G2:G{last}")
    dv_source.add(f"E2:E{last}")

    # 统计区
    sr = len(DATA) + 4
    ws.cell(sr, 1, "统计").font = Font(name='Arial', bold=True, size=11)
    for i, (label, formula) in enumerate([
        ("最高", f'=COUNTIF(D2:D{len(DATA)+1},"最高")'),
        ("高", f'=COUNTIF(D2:D{len(DATA)+1},"高")'),
        ("中", f'=COUNTIF(D2:D{len(DATA)+1},"中")'),
        ("低", f'=COUNTIF(D2:D{len(DATA)+1},"低")'),
    ]):
        ws.cell(sr + 1 + i, 1, label).font = BODY_FONT
        ws.cell(sr + 1 + i, 2, formula).font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"缺失项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
