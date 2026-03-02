"""
TDR1 成本分解生成器 — 复制客户ED&D模板，填充供应商数据
"""
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill

TEMPLATE = "_嵌入文件/TDR_ED&D成本分解模板.xlsx"
OUTPUT = "TDR1_成本分解.xlsx"

# company-profile/成本费率.xlsx 数据 (CNY→EUR, 假设汇率 1EUR=7.8CNY)
EUR_RATE = 7.8

# 工时费率数据 (来自company-profile)
# CAD率: DE工程师, 按¥150/hr → ~19 EUR/hr (fully fringed估算 ×2 = 38 EUR/hr)
# Engineering率: TE/QE/ME, 按¥120/hr → ~15 EUR/hr (fully fringed ×2 = 30 EUR/hr)
CAD_RATE = 38  # EUR/hr
ENG_RATE = 30  # EUR/hr

# ============================================================
# 劳动力成本数据 (Zone B: rows 23-37)
# [Activity, Start, End, FTE_formula, Location, CAD_hrs, CAD_rate, Eng_hrs, Eng_rate]
# ============================================================
LABOR = [
    ["Pre-feasibility & Quotation", "", "", "", "China",
     0, 0, 80, ENG_RATE, 0, 0],
    ["3D CAD Modelling (UG/NX)", "", "", "", "China",
     200, CAD_RATE, 0, 0, 0, 0],
    ["2D CAD Drawings, GD&T", "", "", "", "China",
     120, CAD_RATE, 0, 0, 0, 0],
    ["CAE Virtual Analysis (ANSYS)", "", "", "", "China",
     0, 0, 80, ENG_RATE, 0, 0],
    ["DFMEA / PFMEA", "", "", "", "China",
     0, 0, 60, ENG_RATE, 0, 0],
    ["Tooling Design & Manufacturing Follow-up", "", "", "", "China",
     0, 0, 40, ENG_RATE, 0, 0],
    ["DV Test Planning & Execution", "", "", "", "China",
     0, 0, 200, ENG_RATE, 0, 0],
    ["PV Test Planning & Execution", "", "", "", "China",
     0, 0, 160, ENG_RATE, 0, 0],
    ["Program Management", "", "", "", "China",
     0, 0, 300, ENG_RATE, 0, 0],
    ["PPAP Documentation", "", "", "", "China",
     0, 0, 80, ENG_RATE, 0, 0],
]

# ============================================================
# 测试成本数据 (Zone C: DV/PV/Self-cert/Homologation)
# [Description, Type, Num_tests, Cost_per_test_EUR]
# ============================================================
DV_TESTS = [
    ["Burst/Leak/Pull-off/Flow/Cleanliness (in-house)", "Lab test", 8, 1500],
    ["Pulse endurance P5 (in-house, 4 stations)", "Endurance", 3, 5000],
    ["Thermal cycling + Chemical resistance (in-house)", "Lab test", 5, 2000],
    ["Salt spray corrosion (in-house)", "Lab test", 3, 800],
    ["Vibration endurance (outsourced - CATARC)", "Outsourced", 3, 8000],
    ["MiniSHED emission (outsourced - SGS/TÜV)", "Outsourced", 2, 6000],
    ["Gravelometer SAE J400 (outsourced)", "Outsourced", 2, 3000],
    ["DSC material test (outsourced)", "Outsourced", 1, 500],
]

SELFCERT_TESTS = [
    ["Annual SVP (self-validation plan, year 1)", "Lab test", 1, 5000],
]

# ============================================================
# 零件成本数据 (Zone D: DV/PV mock-up parts)
# [Part name, Purpose, QTY, EUR/pc]
# ============================================================
DV_PARTS = [
    ["DV sample assemblies (fuel line + damper + QC)", "DV testing", 50, 25],
    ["Spare QC connectors (SAE 5/16 + 3/8)", "DV spares", 20, 8],
]

PV_PARTS = [
    ["PV sample assemblies (production tooling)", "PV testing", 50, 20],
    ["Spare QC connectors", "PV spares", 20, 8],
]

# ============================================================
# 团队数据 (Zone F)
# [Name, Role, Years]
# ============================================================
TEAM = [
    ["Wang Qiang", "PM - Project Manager", 15],
    ["Li Ming", "DE - Design Engineer", 12],
    ["Zhang Wei", "QE - Quality Engineer", 10],
    ["Liu Yang", "TE - Test Engineer", 8],
    ["Chen Hao", "ME - Manufacturing Engineer", 10],
    ["Zhao Lei", "PE - Purchasing Engineer", 7],
]


def main():
    # 复制客户模板
    shutil.copy2(TEMPLATE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)

    # 使用Blank sheet
    ws = wb["ED&D PBD (Blank)"]

    # --- Header: Supplier info ---
    ws["O11"] = "Nobel Auto (诺贝尔汽车零部件有限公司)"
    ws["E5"] = "FC00SAA78530 Fuel Supply Line Filter to Engine"
    ws["O14"] = "TBD (待客户提供年需求量)"

    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font = Font(name='Arial', size=10)

    # --- Zone B: Labor (rows 23-37) ---
    for i, row_data in enumerate(LABOR):
        r = 23 + i
        ws.cell(r, 2, row_data[0]).font = font  # Activity
        ws.cell(r, 3, row_data[1]).font = font   # Start date
        ws.cell(r, 4, row_data[2]).font = font   # End date
        ws.cell(r, 5, row_data[3]).font = font   # FTE
        ws.cell(r, 6, row_data[4]).font = font   # Location
        # CAD
        if row_data[5] > 0:
            ws.cell(r, 7, row_data[5]).font = font   # CAD hrs
            ws.cell(r, 8, row_data[6]).font = font   # CAD rate
        # Engineering
        if row_data[7] > 0:
            ws.cell(r, 10, row_data[7]).font = font   # Eng hrs
            ws.cell(r, 12, row_data[8]).font = font   # Eng rate (col L)

        # Date columns yellow (待填)
        ws.cell(r, 3).fill = yellow
        ws.cell(r, 4).fill = yellow

    # --- Zone C: Testing (DV rows 43-44 area) ---
    # DV Testing
    test_row = 43
    for i, t in enumerate(DV_TESTS):
        r = test_row + i
        ws.cell(r, 2, t[0]).font = font
        ws.cell(r, 5, t[1]).font = font
        ws.cell(r, 6, t[2]).font = font
        ws.cell(r, 7, t[3]).font = font

    # Self-cert
    cert_row = 46
    for i, t in enumerate(SELFCERT_TESTS):
        r = cert_row + i
        ws.cell(r, 2, t[0]).font = font
        ws.cell(r, 5, t[1]).font = font
        ws.cell(r, 6, t[2]).font = font
        ws.cell(r, 7, t[3]).font = font

    # --- Zone D: Parts (right side, cols J-Q) ---
    # DV Parts (rows 43-44)
    for i, p in enumerate(DV_PARTS):
        r = 43 + i
        ws.cell(r, 10, p[0]).font = font  # J: Part name
        ws.cell(r, 13, p[1]).font = font   # M: Purpose
        ws.cell(r, 15, p[2]).font = font   # O: QTY
        ws.cell(r, 16, p[3]).font = font   # P: EUR/pc

    # PV Parts (rows 46-48)
    for i, p in enumerate(PV_PARTS):
        r = 46 + i
        ws.cell(r, 10, p[0]).font = font
        ws.cell(r, 13, p[1]).font = font
        ws.cell(r, 15, p[2]).font = font
        ws.cell(r, 16, p[3]).font = font

    # --- Zone F: Team (rows 71-79) ---
    for i, t in enumerate(TEAM):
        r = 71 + i
        ws.cell(r, 2, t[0]).font = font
        ws.cell(r, 9, t[1]).font = font
        ws.cell(r, 11, t[2]).font = font

    wb.save(OUTPUT)
    print(f"已保存: {OUTPUT}")

    # 计算汇总
    labor_total = sum(r[5] * r[6] + r[7] * r[8] for r in LABOR)
    test_total = sum(t[2] * t[3] for t in DV_TESTS + SELFCERT_TESTS)
    parts_total = sum(p[2] * p[3] for p in DV_PARTS + PV_PARTS)
    grand = labor_total + test_total + parts_total
    print(f"劳动力: €{labor_total:,.0f} | 测试: €{test_total:,.0f} | 零件: €{parts_total:,.0f}")
    print(f"合计估算: €{grand:,.0f}")
    print("⚠️ 日期列为黄色待填，需确认项目时间表后补充")


if __name__ == "__main__":
    main()
