import pandas as pd
import openpyxl
import zipfile
import os
import shutil

BASE_DIR = "/home/chu2026/Documents/github/APQP/TestOut/3-2/oem-requirements/test/beta/infile/FUEL BUNDLE AND FUEL SYSTEM"
OUT_EMBED_DIR = os.path.join(BASE_DIR, "out", "_嵌入文件")
os.makedirs(OUT_EMBED_DIR, exist_ok=True)

FILES = [
    "SSTS KP1 Fuel line.xlsx",
    "STLA-DVPR模板.xlsx",
]

def analyze_file(filename):
    filepath = os.path.join(BASE_DIR, filename)
    stem = os.path.splitext(filename)[0]

    print("=" * 100)
    print(f"FILE: {filename}")
    print("=" * 100)

    # ── 1. pandas: all sheets ──────────────────────────────────────────────────
    print("\n[pandas] Reading all sheets …\n")
    xl = pd.ExcelFile(filepath)
    print(f"  Sheet names: {xl.sheet_names}\n")
    for sheet in xl.sheet_names:
        print(f"{'─'*80}")
        print(f"  SHEET: «{sheet}»")
        print(f"{'─'*80}")
        df = pd.read_excel(filepath, sheet_name=sheet, header=None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_rows', None)
        pd.set_option('display.width', 300)
        pd.set_option('display.max_colwidth', 200)
        print(df.to_string())
        print(f"\n  [Shape: {df.shape[0]} rows × {df.shape[1]} cols]\n")

    # ── 2. openpyxl: formatting & merged cells ─────────────────────────────────
    print(f"\n[openpyxl] Formatting / Merged cells …\n")
    wb = openpyxl.load_workbook(filepath)
    for ws in wb.worksheets:
        print(f"  Sheet «{ws.title}»: dimensions={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}")
        merges = list(ws.merged_cells.ranges)
        if merges:
            print(f"    Merged cells ({len(merges)}):")
            for m in merges:
                print(f"      {m}")
        else:
            print("    Merged cells: none")

    # ── 3. zipfile: check xl/embeddings/ ──────────────────────────────────────
    print(f"\n[zipfile] Checking xl/embeddings/ …\n")
    with zipfile.ZipFile(filepath, 'r') as zf:
        all_names = zf.namelist()
        embed_names = [n for n in all_names if n.startswith('xl/embeddings/')]
        if embed_names:
            print(f"  Found {len(embed_names)} embedded file(s):")
            for idx, ename in enumerate(embed_names, start=1):
                ext = os.path.splitext(ename)[1]
                out_name = f"{stem}_嵌入{idx}{ext}"
                out_path = os.path.join(OUT_EMBED_DIR, out_name)
                with zf.open(ename) as src, open(out_path, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                print(f"    [{idx}] {ename}  →  {out_path}")
        else:
            print("  No embedded files found in xl/embeddings/")

        # Also list all zip entries for completeness
        print(f"\n  All zip entries ({len(all_names)}):")
        for n in sorted(all_names):
            print(f"    {n}")

    print()

for f in FILES:
    analyze_file(f)

print("=" * 100)
print("DONE")
print("=" * 100)
