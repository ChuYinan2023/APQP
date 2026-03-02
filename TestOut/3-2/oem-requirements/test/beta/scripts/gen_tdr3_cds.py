"""
TDR3 CDS零部件开发表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr3_cds.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR3_CDS零部件开发表.xlsx",
    "零件号": "",
    "零件名": "",
    "OEM": "",
    "车型": "",
    "SOP目标": "",
}

# ============================================================
# MILESTONE_DATA — 里程碑
# 列顺序: [序号, 里程碑, 计划日期, 实际日期, 负责人, 备注]
# 计划日期/实际日期: AI留空（黄色待填），或从项目时间推算
# ============================================================
MILESTONE_DATA = [
    # [1, "RFQ收到", "2025-01-15", "", "PM", ""],
    # [2, "TKO (Technical Kick-Off)", "", "", "PM", ""],
]

# ============================================================
# CONSTRAINT_DATA — 时间约束
# 列顺序: [约束项, 约束值, 来源, 备注]
# ============================================================
CONSTRAINT_DATA = [
    # ["SOP目标", "2026-03", "客户通知", ""],
]

# ============================================================
# BLOCKER_DATA — 阻塞项
# 列顺序: [阻塞项, 影响里程碑, 当前状态, 负责人, 备注]
# ============================================================
BLOCKER_DATA = [
    # ["3D CAD模型", "模具设计", "未收到", "DE", "需客户提供"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name='Arial', bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BLOCKER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)


def write_section(ws, row, title, headers, widths, data, special_cols=None):
    ws.cell(row, 1, title).font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=len(headers))
    row += 1
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, w)
    for r_data in data:
        row += 1
        for c, v in enumerate(r_data, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
            if special_cols and c in special_cols and not v:
                cell.fill = INPUT_FILL
    return row + 2


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CDS开发表"

    # 项目信息
    info = [("零件号", CONFIG['零件号']), ("零件名", CONFIG['零件名']),
            ("OEM", CONFIG['OEM']), ("车型", CONFIG['车型']), ("SOP目标", CONFIG['SOP目标'])]
    for i, (k, v) in enumerate(info):
        ws.cell(i + 1, 1, k).font = Font(name='Arial', bold=True, size=10)
        ws.cell(i + 1, 2, v).font = BODY_FONT
        if not v:
            ws.cell(i + 1, 2).fill = INPUT_FILL

    row = len(info) + 2

    # 里程碑
    row = write_section(ws, row, "一、开发里程碑",
                        ["序号", "里程碑", "计划日期", "实际日期", "负责人", "备注"],
                        [6, 30, 14, 14, 10, 20],
                        MILESTONE_DATA, special_cols={3, 4})

    # 时间约束
    row = write_section(ws, row, "二、时间约束",
                        ["约束项", "约束值", "来源", "备注"],
                        [20, 20, 20, 20],
                        CONSTRAINT_DATA)

    # 阻塞项
    for r_data in BLOCKER_DATA:
        pass  # just for counting
    ws.cell(row, 1, "三、阻塞项").font = SECTION_FONT
    ws.cell(row, 1).fill = BLOCKER_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=5)
    row += 1
    b_headers = ["阻塞项", "影响里程碑", "当前状态", "负责人", "备注"]
    b_widths = [25, 20, 12, 10, 20]
    for c, (h, w) in enumerate(zip(b_headers, b_widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    for r_data in BLOCKER_DATA:
        row += 1
        for c, v in enumerate(r_data, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.fill = BLOCKER_FILL

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"里程碑: {len(MILESTONE_DATA)}, 约束: {len(CONSTRAINT_DATA)}, 阻塞: {len(BLOCKER_DATA)}")


if __name__ == "__main__":
    main()
