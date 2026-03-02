"""
TDR8补充材料清单生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr8_materials.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR8_补充材料清单.xlsx",
}

# ============================================================
# DATA — 测试样件与材料
# 列顺序: [序号, 测试项, 测试类型, 验收标准, 样件类型, DV数量, PV数量, 单价(¥), 备注]
# 测试类型: DV / PV / DV+PV
# 单价从company-profile/成本费率.xlsx自动填，无则黄色待填
# ============================================================
DATA = [
    # [1, "爆破压力", "DV+PV", "≥40 bar @23°C", "成品总成", 5, 5, 15.00, ""],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
MONEY_FMT = '¥ #,##0.00'

HEADERS = ["序号", "测试项", "测试类型", "验收标准", "样件类型", "DV数量", "PV数量", "合计数量", "单价(¥)", "费用小计(¥)", "备注"]
WIDTHS = [6, 20, 10, 25, 12, 8, 8, 8, 12, 12, 15]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "补充材料清单"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for i, row in enumerate(DATA):
        r = i + 2
        # 写入基础数据（序号~PV数量 = 列1~7，单价=列9，备注=列11）
        for c, v in enumerate(row[:7], 1):
            cell = ws.cell(r, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
        # 合计数量（公式）
        ws.cell(r, 8, f'=F{r}+G{r}')
        ws.cell(r, 8).fill = FORMULA_FILL
        ws.cell(r, 8).border = THIN
        ws.cell(r, 8).font = BODY_FONT
        ws.cell(r, 8).alignment = CENTER
        # 单价
        ws.cell(r, 9, row[7] if len(row) > 7 else "")
        ws.cell(r, 9).number_format = MONEY_FMT
        ws.cell(r, 9).border = THIN
        ws.cell(r, 9).font = BODY_FONT
        if not (len(row) > 7 and row[7]):
            ws.cell(r, 9).fill = INPUT_FILL
        # 费用小计（公式）
        ws.cell(r, 10, f'=H{r}*I{r}')
        ws.cell(r, 10).number_format = MONEY_FMT
        ws.cell(r, 10).fill = FORMULA_FILL
        ws.cell(r, 10).border = THIN
        ws.cell(r, 10).font = BODY_FONT
        # 备注
        ws.cell(r, 11, row[8] if len(row) > 8 else "")
        ws.cell(r, 11).border = THIN
        ws.cell(r, 11).font = BODY_FONT

    # 汇总行
    sr = len(DATA) + 2
    ws.cell(sr, 1, "合计").font = Font(name='Arial', bold=True, size=10)
    ws.cell(sr, 1).border = THIN
    ws.cell(sr, 8, f'=SUM(H2:H{sr-1})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(sr, 8).border = THIN
    ws.cell(sr, 8).fill = FORMULA_FILL
    ws.cell(sr, 10, f'=SUM(J2:J{sr-1})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(sr, 10).number_format = MONEY_FMT
    ws.cell(sr, 10).border = THIN
    ws.cell(sr, 10).fill = FORMULA_FILL

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"测试项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
