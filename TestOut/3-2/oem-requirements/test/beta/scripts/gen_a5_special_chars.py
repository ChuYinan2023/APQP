"""
A5特殊特性清单生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_a5_special_chars.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "A5_特殊特性清单.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [序号, 特性名称, 类别, 规范要求, 来源文件, 验证方法, 控制方法建议, 备注]
# 类别: "CC"=关键特性(安全/法规) / "SC"=重要特性(功能/耐久)
# ============================================================
DATA = [
    # [1, "爆破压力", "CC", "≥40 bar @23°C", "PF.90197 §5.2", "爆破试验", "100%在线气密+抽检爆破", ""],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CC_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
SC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

HEADERS = ["序号", "特性名称", "类别", "规范要求", "来源文件", "验证方法", "控制方法建议", "备注"]
WIDTHS = [6, 25, 8, 30, 20, 20, 25, 20]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "特殊特性清单"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        cls = str(ws.cell(r, 3).value or "")
        fill = CC_FILL if cls == "CC" else SC_FILL if cls == "SC" else None
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    wb.save(CONFIG["输出路径"])
    cc = sum(1 for r in DATA if len(r) > 2 and r[2] == "CC")
    sc = sum(1 for r in DATA if len(r) > 2 and r[2] == "SC")
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"特殊特性: {len(DATA)} 条 (CC:{cc}, SC:{sc})")


if __name__ == "__main__":
    main()
