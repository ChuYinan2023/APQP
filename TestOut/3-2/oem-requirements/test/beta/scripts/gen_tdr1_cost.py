"""
TDR1成本分解生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr1_cost.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR1_成本分解.xlsx",
    "零件号": "",
    "零件名": "",
    "管理费率": 0.15,
    "废料率": 0.03,
    "利润率": 0.08,
}

# ============================================================
# BOM_DATA — 原材料（从Step1 BOM提取）
# 列顺序: [序号, 零件/材料, 材料牌号, 单位, 用量, 单价(¥), 备注]
# 单价从company-profile/成本费率.xlsx自动填，无则黄色待填
# ============================================================
BOM_DATA = [
    # [1, "主管体", "PA66 GF30", "kg", 0.085, 28.00, ""],
]

# ============================================================
# PROCESS_DATA — 加工工艺
# 列顺序: [序号, 工序, 设备, 时间(min), 费率(¥/min), 备注]
# 费率从company-profile/成本费率.xlsx自动填，无则黄色待填
# ============================================================
PROCESS_DATA = [
    # [1, "注塑", "海天160T", 0.45, 1.50, ""],
]

# ============================================================
# OTHER_DATA — 其他费用
# 列顺序: [序号, 项目, 金额(¥), 分摊方式, 备注]
# ============================================================
OTHER_DATA = [
    # [1, "注塑模具", 80000, "50万件分摊", ""],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FONT = Font(name='Arial', bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FORMULA_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
MONEY_FMT = '¥ #,##0.00'


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成本分解"

    # 项目信息
    ws['A1'] = '零件号'
    ws['B1'] = CONFIG['零件号']
    ws['C1'] = '零件名'
    ws['D1'] = CONFIG['零件名']
    for c in ['A', 'C']:
        ws[f'{c}1'].font = Font(name='Arial', bold=True, size=10)
    row = 3

    # === 原材料 ===
    ws.cell(row, 1, "一、原材料成本").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    bom_headers = ["序号", "零件/材料", "材料牌号", "单位", "用量", "单价(¥)", "小计(¥)"]
    bom_widths = [6, 20, 18, 6, 10, 12, 12]
    for c, (h, w) in enumerate(zip(bom_headers, bom_widths), 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    bom_start = row + 1
    for i, r in enumerate(BOM_DATA):
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.cell(row, 7, f'=E{row}*F{row}')
        ws.cell(row, 7).number_format = MONEY_FMT
        ws.cell(row, 7).fill = FORMULA_FILL
        ws.cell(row, 7).border = THIN
    bom_end = row
    row += 1
    ws.cell(row, 5, "材料小计").font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 7, f'=SUM(G{bom_start}:G{bom_end})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 7).number_format = MONEY_FMT
    ws.cell(row, 7).fill = FORMULA_FILL
    ws.cell(row, 7).border = THIN
    mat_total_row = row

    # === 加工工艺 ===
    row += 2
    ws.cell(row, 1, "二、加工成本").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    proc_headers = ["序号", "工序", "设备", "时间(min)", "费率(¥/min)", "小计(¥)", "备注"]
    for c, h in enumerate(proc_headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    proc_start = row + 1
    for i, r in enumerate(PROCESS_DATA):
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT
            cell.alignment = WRAP
        ws.cell(row, 6, f'=D{row}*E{row}')
        ws.cell(row, 6).number_format = MONEY_FMT
        ws.cell(row, 6).fill = FORMULA_FILL
        ws.cell(row, 6).border = THIN
    proc_end = row
    row += 1
    ws.cell(row, 4, "加工小计").font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 6, f'=SUM(F{proc_start}:F{proc_end})').font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 6).number_format = MONEY_FMT
    ws.cell(row, 6).fill = FORMULA_FILL
    ws.cell(row, 6).border = THIN
    proc_total_row = row

    # === 其他费用 ===
    row += 2
    ws.cell(row, 1, "三、其他费用").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    row += 1
    other_headers = ["序号", "项目", "金额(¥)", "分摊方式", "备注"]
    for c, h in enumerate(other_headers, 1):
        cell = ws.cell(row, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    for r in OTHER_DATA:
        row += 1
        for c, v in enumerate(r, 1):
            cell = ws.cell(row, c, v)
            cell.border = THIN
            cell.font = BODY_FONT

    # === 汇总 ===
    row += 2
    ws.cell(row, 1, "四、成本汇总").font = SECTION_FONT
    ws.cell(row, 1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=7)
    summary = [
        ("材料成本", f'=G{mat_total_row}'),
        ("废料成本", f'=G{mat_total_row}*{CONFIG["废料率"]}'),
        ("加工成本", f'=F{proc_total_row}'),
        ("管理费", f'=F{proc_total_row}*{CONFIG["管理费率"]}'),
        ("利润", f'=(G{mat_total_row}+F{proc_total_row})*{CONFIG["利润率"]}'),
    ]
    for label, formula in summary:
        row += 1
        ws.cell(row, 2, label).font = BODY_FONT
        ws.cell(row, 2).border = THIN
        ws.cell(row, 3, formula).font = BODY_FONT
        ws.cell(row, 3).number_format = MONEY_FMT
        ws.cell(row, 3).fill = FORMULA_FILL
        ws.cell(row, 3).border = THIN
    row += 1
    ws.cell(row, 2, "单件总成本").font = Font(name='Arial', bold=True, size=11)
    ws.cell(row, 2).border = THIN
    sum_rows = [row - i for i in range(len(summary), 0, -1)]
    ws.cell(row, 3, f'=SUM(C{sum_rows[0]}:C{sum_rows[-1]})').font = Font(name='Arial', bold=True, size=11)
    ws.cell(row, 3).number_format = MONEY_FMT
    ws.cell(row, 3).fill = FORMULA_FILL
    ws.cell(row, 3).border = THIN

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"BOM: {len(BOM_DATA)} 项, 工序: {len(PROCESS_DATA)} 项, 其他: {len(OTHER_DATA)} 项")


if __name__ == "__main__":
    main()
