"""
A6质量目标表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_a6_quality_targets.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "A6_质量目标表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 指标名称, 目标值, 来源文件, 验证方法, 备注]
# 类别: 市场质量目标 / 可靠性目标 / 过程能力 / 清洁度目标 / 安全法规 / 环保合规 / 持续符合性
# 同类别的行会自动合并单元格
# ============================================================
DATA = [
    # ["市场质量目标", "12MIS故障率", "≤50 PPM", "SSTS §3.2", "SPC统计", ""],
    # ["市场质量目标", "0km投诉率", "0 PPM", "SSTS §3.2", "出货检验", ""],
    # ["可靠性目标", "脉冲寿命", "P99C90 ≥ 15万次", "PF.90197 §6.3", "脉冲耐久试验", ""],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "指标名称", "目标值", "来源文件", "验证方法", "备注"]
WIDTHS = [18, 25, 30, 20, 20, 20]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "质量目标表"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        cat_cell = ws.cell(r, 1)
        cat_cell.font = CAT_FONT
        cat_cell.fill = CAT_FILL

    # 合并同类别单元格
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    wb.save(CONFIG["输出路径"])
    cats = len(set(r[0] for r in DATA)) if DATA else 0
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"质量指标: {len(DATA)} 条, {cats} 个类别")


if __name__ == "__main__":
    main()
