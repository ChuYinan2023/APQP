"""
缺失项跟踪表生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_missing_items.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "缺失项跟踪表.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [序号, 缺失文档, 引用来源, 紧急度, 来源方, 影响范围, 处理状态]
# 紧急度: 最高/高/中/低
# 来源方: 向客户要/自行购买/内部收集
# 处理状态: 未发起/已发起/已收到/不适用
# ============================================================
DATA = [
    # [1, "PF.90197 燃油管路性能规范", "CTS §2.1, SSTS §4.3", "最高", "向客户要", "核心性能测试无法定义", "未发起"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
URGENCY_FILLS = {
    "最高": PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid"),
    "高": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "中": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    "低": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)

HEADERS = ["序号", "缺失文档", "引用来源", "紧急度", "来源方", "影响范围", "处理状态"]
WIDTHS = [6, 35, 25, 8, 12, 30, 10]


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "缺失项跟踪"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        urgency = str(ws.cell(r, 4).value or "")
        fill = URGENCY_FILLS.get(urgency)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    # 数据验证
    dv_urgency = DataValidation(type="list", formula1='"最高,高,中,低"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"未发起,已发起,已收到,不适用"', allow_blank=True)
    dv_source = DataValidation(type="list", formula1='"向客户要,自行购买,内部收集"', allow_blank=True)
    ws.add_data_validation(dv_urgency)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_source)
    last = len(DATA) + 50
    dv_urgency.add(f"D2:D{last}")
    dv_status.add(f"G2:G{last}")
    dv_source.add(f"E2:E{last}")

    # 统计区
    sr = len(DATA) + 4
    ws.cell(sr, 1, "统计").font = Font(name='Arial', bold=True, size=11)
    for i, (label, formula) in enumerate([
        ("最高", f'=COUNTIF(D2:D{len(DATA)+1},"最高")'),
        ("高", f'=COUNTIF(D2:D{len(DATA)+1},"高")'),
        ("中", f'=COUNTIF(D2:D{len(DATA)+1},"中")'),
        ("低", f'=COUNTIF(D2:D{len(DATA)+1},"低")'),
    ]):
        ws.cell(sr + 1 + i, 1, label).font = BODY_FONT
        ws.cell(sr + 1 + i, 2, formula).font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"缺失项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
