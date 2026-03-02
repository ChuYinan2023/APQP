"""
TDR6例外事项清单生成器
使用方法：复制到out/，填写DATA区，运行 python3 gen_tdr6_exceptions.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "TDR6_例外事项清单.xlsx",
}

# ============================================================
# DATA — 修改此区域
# 列顺序: [类别, 序号, 客户要求, 要求来源, 判定, 我方现状, 替代方案, 备注]
# 判定: "符合"/"例外"/"待确认" — 有company-profile时AI自动填
# 同类别行自动合并
# ============================================================
DATA = [
    # ["质量体系", 1, "IATF 16949认证", "SSTS §2.1", "符合", "已认证", "", ""],
    # ["制造工艺", 2, "镀镍处理", "CTS §4.3", "例外", "无镀镍线", "委外XX厂", "需签委外协议"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PEND_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
CAT_FONT = Font(name='Arial', size=10, bold=True)

HEADERS = ["类别", "序号", "客户要求", "要求来源", "判定", "我方现状", "替代方案", "备注"]
WIDTHS = [15, 6, 30, 15, 10, 25, 25, 20]
JUDGMENT_MAP = {"符合": PASS_FILL, "例外": FAIL_FILL, "待确认": PEND_FILL}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "例外事项清单"

    ws.append(HEADERS)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"

    for row in DATA:
        ws.append(row)
    for r in range(2, len(DATA) + 2):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP
            cell.font = BODY_FONT
        # 类别列
        ws.cell(r, 1).font = CAT_FONT
        ws.cell(r, 1).fill = CAT_FILL
        # 判定列颜色
        j = str(ws.cell(r, 5).value or "")
        if j in JUDGMENT_MAP:
            ws.cell(r, 5).fill = JUDGMENT_MAP[j]
        elif not j:
            ws.cell(r, 5).fill = INPUT_FILL
        # 我方现状（无值则黄色）
        if not ws.cell(r, 6).value:
            ws.cell(r, 6).fill = INPUT_FILL

    # 合并同类别
    if DATA:
        start_row = 2
        for i in range(1, len(DATA)):
            if DATA[i][0] != DATA[i - 1][0]:
                if start_row < i + 1:
                    ws.merge_cells(start_row=start_row, end_row=i + 1, start_column=1, end_column=1)
                start_row = i + 2
        if start_row < len(DATA) + 2:
            ws.merge_cells(start_row=start_row, end_row=len(DATA) + 1, start_column=1, end_column=1)

    # 统计
    sr = len(DATA) + 4
    ws.cell(sr, 1, "统计").font = Font(name='Arial', bold=True, size=11)
    for i, (label, fill) in enumerate([("符合", PASS_FILL), ("例外", FAIL_FILL), ("待确认", PEND_FILL)]):
        ws.cell(sr + 1 + i, 1, label).font = BODY_FONT
        ws.cell(sr + 1 + i, 1).fill = fill
        ws.cell(sr + 1 + i, 2, f'=COUNTIF(E2:E{len(DATA)+1},"{label}")').font = BODY_FONT

    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"要求项: {len(DATA)} 条")


if __name__ == "__main__":
    main()
