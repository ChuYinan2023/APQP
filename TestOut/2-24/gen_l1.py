"""
L1工程特性清单生成器
使用方法：复制到output/gen_l1.py，填写数据区，运行 python3 gen_l1.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "阶段2_L1工程特性清单.xlsx",
}

# ============================================================
# L1_DATA — 修改此区域
# 列顺序: [ID, 特性分类, 工程特性名称, 目标值/要求, 原文描述, 类别, 文件来源, 章节, 页码, 备注]
# 类别: "C(关键)"=影响安全/核心功能 / "A(一般)"=其他
# 冲突行在备注中标注，如 "[冲突] PF.90150=XX, CTS=YY"
# ============================================================
L1_DATA = [
    # ===== 工作条件 =====
    [1, "工作条件", "工作环境温度范围", "+45°C ~ 120°C", "Working ambient temperature +45°C to 120°C", "A(一般)", "CTS", "3.1", "9", ""],
    [2, "工作条件", "工作燃油温度范围", "-45°C ~ 90°C", "Working Fuel temperature -45°C to 90°C", "A(一般)", "CTS", "3.1", "9", ""],
    [3, "工作条件", "工作燃油压力范围", "450 ~ 700 kPa", "Operating Fuel pressure 450 to 700 kPa", "C(关键)", "CTS", "3.1", "9", "Filter Spec §1.2: 0~700kPa，CTS收窄下限至450kPa"],
    [4, "工作条件", "系统最大工作压力", "0 ~ 700 kPa (0 ~ 101.5 psi)", "Working pressure range from 0 to 700 [kPa]", "C(关键)", "Filter Spec", "1.2", "4", ""],
    [5, "工作条件", "使用环境相对湿度", "0 ~ 100%", "Relative humidity in use: 0 to 100%", "A(一般)", "CTS", "3.1", "9", ""],
    [6, "工作条件", "存储环境相对湿度", "≤60%", "Storage: max. 60%", "A(一般)", "CTS", "3.1", "9", ""],
    [7, "工作条件", "极端运输/存储温度", "-30°C ~ +70°C，湿度≤60%，6个月", "Extreme transport condition: -30°C to +70°C, humidity ≤60% for 6 months", "A(一般)", "CTS", "3.1", "9", ""],

    # ===== 过滤性能 =====
    [8, "过滤性能", "标称流量", "190 l/h", "Nominal Flow Rate 190 l/h", "C(关键)", "CTS", "3.1", "9", ""],
    [9, "过滤性能", "过滤效率(初始) @4μm", "≥98% @ 190 l/h", "Initial Filtration Efficiency 98% @4 um @ Nominal Flow Rate", "C(关键)", "CTS", "3.1", "9", "[冲突] Filter Spec Table 5: ≥90%@4μm(初始+平均); CTS: 98%@4μm(初始)。按CTS优先取98%"],
    [10, "过滤性能", "过滤效率(初始+平均) @4μm", "≥90%", "Initial and average efficiency ≥90% @ >4μm(c) per ISO 19438", "C(关键)", "Filter Spec", "7.1/Table5", "20", "通用Spec多粒径要求，与CTS 98%并行"],
    [11, "过滤性能", "过滤效率(初始+平均) @5μm", "≥95%", "Initial and average efficiency ≥95% @ >5μm(c) per ISO 19438", "C(关键)", "Filter Spec", "7.1/Table5", "20", ""],
    [12, "过滤性能", "过滤效率(初始+平均) @10μm", "≥99%", "Initial and average efficiency ≥99% @ >10μm(c) per ISO 19438", "C(关键)", "Filter Spec", "7.1/Table5", "20", ""],
    [13, "过滤性能", "过滤效率(初始+平均) @15μm", "≥99%", "Initial and average efficiency ≥99% @ >15μm(c) per ISO 19438", "A(一般)", "Filter Spec", "7.1/Table5", "20", ""],
    [14, "过滤性能", "过滤效率(初始+平均) @20μm", "≥99.9%", "Initial and average efficiency ≥99.9% @ >20μm(c) per ISO 19438", "A(一般)", "Filter Spec", "7.1/Table5", "20", ""],
    [15, "过滤性能", "过滤效率(初始+平均) @30μm", "≥99.9%", "Initial and average efficiency ≥99.9% @ >30μm(c) per ISO 19438", "A(一般)", "Filter Spec", "7.1/Table5", "20", ""],
    [16, "过滤性能", "单通过滤粒子保持率 @5μm", "≥95%", "Single pass particle retention ≥95% at 5 microns", "C(关键)", "Filter Spec", "7.2", "20", ""],
    [17, "过滤性能", "容尘量(DHC)", "≥30g @ 500kPa寿命末期压力", "DHC 30g @ 500kPa @ end of life pressure (ISO 4020)", "C(关键)", "CTS", "3.1", "9", ""],
    [18, "过滤性能", "过滤效率测试终止压降", "250 kPa (36.3 psi)", "End of test at pressure drop of 250 [kPa]", "A(一般)", "Filter Spec", "7.1", "19", ""],

    # ===== 水分离性能 =====
    [19, "水分离性能", "水分离效率(ISO 16332)", "≥93% @ 190 l/h", "Water Separation efficiency 93% @ Nominal Flow Rate (ISO 16332)", "C(关键)", "CTS", "3.1", "9", "[冲突] Filter Spec §7.12: >95%@200l/h; CTS: 93%@190l/h。客户未明确，暂取CTS 93%"],
    [20, "水分离性能", "水分离效率(通用Spec)", ">95% @ 200 l/h，含1500ppm+2%水，EN590+30%FAME", "Efficiency higher than 95% (ISO 16332, 200 l/h, 1500ppm and 2% water, EN590+30%FAME)", "C(关键)", "Filter Spec", "7.12", "25", "[冲突] 见#19"],
    [21, "水分离性能", "储水容积", "150 ml", "Water storage capacity: 150 ml", "C(关键)", "CTS", "3.1", "9", ""],
    [22, "水分离性能", "水位报警量", "100 ml", "Water alarm level: 100 ml", "C(关键)", "CTS", "3.1", "9", ""],
    [23, "水分离性能", "乳化水分离效率", ">1500 PPM", "Overall water separation efficiency > 1500 PPM", "C(关键)", "Filter Spec", "7.17", "28", "测试条件：IFT oF15，0.25%和2%水"],

    # ===== 压降性能 =====
    [24, "压降性能", "新品压降 @ 190 l/h (CTS)", "≤10 kPa (ISO 4020)", "Pressure drop according to ISO 4020: 10kPa", "C(关键)", "CTS", "3.1", "9", "[冲突] Filter Spec §7.7: ≤30kPa@200l/h。工况不同，两者并行"],
    [25, "压降性能", "新品压降 @ 200 l/h (通用)", "≤30 kPa (4.3 psi)", "Max Δp: 30 [kPa] - 4.3 [psi] at 200 l/h", "C(关键)", "Filter Spec", "7.7", "22", "测试液: FPW.55520/01"],
    [26, "压降性能", "低温工况压降", "≤400 kPa @ 最小流量", "Pressure drop on filter ≤ 400 [kPa] at minimum flow rate", "C(关键)", "Filter Spec", "7.13", "26", "供电12.5V和8V两工况均需满足"],

    # ===== 密封性能 =====
    [27, "密封性能", "外部气密性(空气)", "泄漏值 < 3 cm³/min @ -40°C/23°C/80°C", "Leak value < 3 cm³/min at -40/23/80°C", "C(关键)", "Filter Spec", "7.4.1", "21", "最大工作压力下测试"],
    [28, "密封性能", "外部密封性(燃油)", "无泄漏 @ 10.3 bar，-40°C/23°C/80°C", "No leaks allowed at 10.3 bar (150 psi) at -40/23/80°C", "C(关键)", "Filter Spec", "7.4.2", "21", "EN590 Artic fuel class 4"],
    [29, "密封性能", "滤芯垫片密封(脏侧/净侧)", "泄漏 ≤ 3 cm³/min @ ≥500 kPa", "Maximum leak of 3 cm³/min with pressure ≥ 500 kPa", "C(关键)", "Filter Spec", "7.5", "22", "多温度：-25°C/+23°C/+80°C；含FAME老化后复测"],
    [30, "密封性能", "内部密封(气泡试验)", "出口泄漏 < 100 cm³/min @ 1.3±0.1 kPa", "Air leak from outlet < 100 cm³/min at 1.3±0.1 kPa", "C(关键)", "Filter Spec", "7.6", "22", "ISO 4020 / ISO 2942"],

    # ===== 机械强度 =====
    [31, "机械强度", "滤芯塌陷压力", "> 800 kPa (116 psi)", "Filter element collapse pressure > 800 [kPa]", "C(关键)", "Filter Spec", "7.9", "24", "ISO 4020"],
    [32, "机械强度", "爆破压力", "≥ 20 bar", "Filter burst resistance up to 20 bar", "C(关键)", "Filter Spec", "7.10", "24", "ISO 4020"],
    [33, "机械强度", "金属接头压装扭矩", "≥ 7.3 Nm 不脱落", "Fuel fittings withstand minimum 7.3 Nm torque without breaking free", "C(关键)", "Filter Spec", "7.11.1", "25", ""],
    [34, "机械强度", "塑料接头侧向载荷", "满足 SAE J2044 §6.4", "Plastic fittings meet SAE J2044 6.4 Side load Capability", "A(一般)", "Filter Spec", "7.11.2", "25", ""],
    [35, "机械强度", "螺纹元件拧紧力矩", "在图纸标注范围内", "Tightening torque within drawing limits", "A(一般)", "Filter Spec", "7.15", "27", "需工程图纸确定具体值"],

    # ===== 清洁度 =====
    [36, "清洁度", "滤清器内部清洁度(总重)", "< 0.5 mg/part", "Total weight of contamination < 0.5mg/part", "C(关键)", "Filter Spec", "7.8", "24", ""],
    [37, "清洁度", "清洁度 — 颗粒 >50~100μm", "< 90个", "Particle class >50-100μm: < 90", "A(一般)", "Filter Spec", "7.8/Table6", "24", ""],
    [38, "清洁度", "清洁度 — 颗粒 >100~200μm", "< 25个", "Particle class >100-200μm: < 25", "A(一般)", "Filter Spec", "7.8/Table6", "24", ""],
    [39, "清洁度", "清洁度 — 颗粒 >200~500μm", "0个", "Particle class >200-500μm: 0", "C(关键)", "Filter Spec", "7.8/Table6", "24", ""],

    # ===== 电气 — 加热器 =====
    [40, "电气-加热器", "加热器最低效率", "η ≥ 0.75 (6分钟后)", "Minimum efficiency η = 0.75 after 6 minutes", "C(关键)", "Filter Spec", "6.4.10.1", "18", "蒸馏水测试条件"],
    [41, "电气-加热器", "加热器最大峰值电流", "40A", "Max peak current: 40A", "C(关键)", "Filter Spec", "6.4.10.2", "18", "CTS示例值"],
    [42, "电气-加热器", "加热器峰值持续时间", "≤ 2.5s", "Max peak duration: 2.5s", "A(一般)", "Filter Spec", "6.4.10.2", "18", "CTS示例值"],
    [43, "电气-加热器", "加热器最大工作电流", "30A", "Max working current: 30A", "C(关键)", "Filter Spec", "6.4.10.2", "18", "CTS示例值"],
    [44, "电气-加热器", "加热器最小ΔT", "≥ 3°C (6分钟后, 130l/h, 12.5V, -26°C)", "ΔT minimum: 3°C measured after 6 minutes", "C(关键)", "Filter Spec", "6.4.10.2", "18", "CTS示例值，Arctic Diesel Class 3"],
    [45, "电气-加热器", "加热器恒温器关断 — 满油", "< 10.0 分钟", "Heater turns off before 10.0 minutes when filter is full of fuel", "C(关键)", "Filter Spec", "6.4.9", "16", ""],
    [46, "电气-加热器", "加热器恒温器关断 — 空滤", "< 3.0 分钟", "Heater turns off before 3.0 minutes when filter is empty", "C(关键)", "Filter Spec", "6.4.9", "16", ""],
    [47, "电气-加热器", "加热器失效安全性", "恒温器失效(闭合)时无外部/内部损伤", "Heater fails 'safe' without external/internal damage when thermostat failed in closed position", "C(关键)", "Filter Spec", "6.4.9", "17", ""],
    [48, "电气-加热器", "恒温器开关信号质量", "尖锐无抖动，可重复温度", "Off signal must be sharp (no dithering) at a repeatable temperature", "A(一般)", "Filter Spec", "7.18", "28", "10次重复测试"],

    # ===== 电气 — NTC温度传感器 =====
    [49, "电气-NTC", "NTC工作温度范围", "-40°C ~ 125°C", "NTC sensor operating temperature range: -40 ÷ 125 [°C]", "A(一般)", "Filter Spec", "7.19", "31", ""],
    [50, "电气-NTC", "NTC供电电压", "5V ±2% (ECU供电)", "NTC powered by ECU with power supply voltage = 5V +/- 2%", "A(一般)", "Filter Spec", "7.19", "31", "示例值，正式值在Product Card"],
    [51, "电气-NTC", "NTC温度测量精度 @80°C", "T4-T2 < 15°C (300s后)", "Difference T4-T2 < 15°C after 300s when fuel at 80±5°C", "C(关键)", "Filter Spec", "7.20", "32", "流量200 l/h"],
    [52, "电气-NTC", "NTC特性曲线 (25°C标称)", "2.057 kΩ (1.941~2.173)", "Nom 2.057kΩ, Min 1.941, Max 2.173 @ 25°C", "A(一般)", "Filter Spec", "7.19/Table7", "29", "示例值，正式值在Product Card"],
    [53, "电气-NTC", "SENT协议通信", "PIN 5: SENT Protocol", "SENT Protocol on Pin 5", "C(关键)", "CTS", "3.3", "10", ""],

    # ===== 电气 — WiF传感器 =====
    [54, "电气-WiF", "WiF传感器类型", "主动式(Active)", "WIF sensor type: active", "A(一般)", "Filter Spec", "7.21", "32", "示例值"],
    [55, "电气-WiF", "WiF正常状态电压", "V out > 5.5V", "Normal condition V out > 5.5 V", "C(关键)", "Filter Spec", "7.21", "33", ""],
    [56, "电气-WiF", "WiF报警状态电压", "V out < 2.2V", "Alarm condition V out < 2.2 V", "C(关键)", "Filter Spec", "7.21", "33", ""],
    [57, "电气-WiF", "WiF最大吸收电流", "< 50 mA", "Absorbed current < 50 mA in every condition", "A(一般)", "Filter Spec", "7.21", "34", ""],
    [58, "电气-WiF", "WiF自检时间 (-30~110°C)", "3s ±20% (2.4~3.6s)", "Self-test time 3s ±20% in -30÷110°C", "A(一般)", "Filter Spec", "7.21", "34", ""],
    [59, "电气-WiF", "WiF自检时间 (-40~125°C)", "1.7 ~ 4.5s", "Self-test time 1.7s ÷ 4.5s in -40÷125°C", "A(一般)", "Filter Spec", "7.21", "34", ""],
    [60, "电气-WiF", "WiF供电电压范围", "10V ~ 16V (A1分类)", "Supply Voltage 10V ÷ 16V (A1 classification)", "A(一般)", "Filter Spec", "7.21", "34", ""],
    [61, "电气-WiF", "WiF工作温度范围", "-40°C ~ 125°C", "Operating temperature -40 ÷ 125°C", "A(一般)", "Filter Spec", "7.21", "34", ""],
    [62, "电气-WiF", "WIF探针抗电解", "168h @ 14V后仍正常工作", "After 168h electrolysis test at 14V, probe passes WIF Signal Level test", "A(一般)", "Filter Spec", "6.4.6", "15", ""],

    # ===== 电气 — 通用 =====
    [63, "电气-通用", "电气连接器型号", "APTIV p/n 211FYT0523, KEY CODE A", "APTIV p/n 211FYT0523, KEY CODE A", "C(关键)", "CTS", "3.3", "10", "6-pin连接器"],
    [64, "电气-通用", "防护等级", "IP6K9K", "Water-penetration classification IP6K9K", "C(关键)", "Filter Spec", "5.1/Table4", "13", "加热器/WiF/NTC三个设备均为IP6K9K"],
    [65, "电气-通用", "ESD分类", "SE2", "ESD classification SE2", "A(一般)", "Filter Spec", "5.1/Table4", "13", ""],
    [66, "电气-通用", "振动等级", "V2", "Vibrational classification V2", "A(一般)", "Filter Spec", "5.1/Table4", "13", ""],
    [67, "电气-通用", "静电电荷防护", "满足 SAE J1645", "Meet electrostatic charge mitigation criteria per SAE J1645", "A(一般)", "Filter Spec", "7.16", "27", ""],

    # ===== 液压接口 =====
    [68, "液压接口", "进口类型及尺寸", "10mm 快接, SAE J2044 / PF.90298 / B12 2813", "Inlet port 10mm diameter, ref B12 2813, per SAE J2044", "C(关键)", "CTS", "3.2", "9-10", ""],
    [69, "液压接口", "出口类型及尺寸", "10mm 快接, SAE J2044 / PF.90298 / B12 2813", "Outlet port 10mm diameter, ref B12 2813, per SAE J2044", "C(关键)", "CTS", "3.2", "9-10", ""],
    [70, "液压接口", "接头尺寸符合性", "满足 SAE J2044 尺寸要求", "Fuel fittings comply with SAE J2044 dimensional requirements", "A(一般)", "Filter Spec", "6.4.7", "15", ""],

    # ===== 环境耐久 =====
    [71, "环境耐久", "热冲击(TS)", "125/-40°C, 500 cycles", "Thermal Shock 125/-40°C, 500 cycles per CS.00056", "A(一般)", "Filter Spec", "5.1.1", "7", "详细profile依赖CS.00056"],
    [72, "环境耐久", "带电热循环(PTCE)", "125/-40°C, 500 cycles", "Powered Thermal Cycle 125/-40°C, 500 cycles per CS.00056", "A(一般)", "Filter Spec", "5.1.2", "8", ""],
    [73, "环境耐久", "高温工作耐久(HTOE)", "125°C, 1300 hours", "High Temp Operating Endurance 125°C, 1300 hours", "A(一般)", "Filter Spec", "5.1.3", "8", ""],
    [74, "环境耐久", "高温高湿耐久(HTHE)", "85°C/85%RH, 1000 hours", "HTHE 85°C & 85%RH, 1000 hours", "A(一般)", "Filter Spec", "5.1.4", "8", ""],
    [75, "环境耐久", "存储温度暴露(SSTE)", "-40~95°C, 54 hours", "SSTE -40 to 95°C, 54 hours", "A(一般)", "Filter Spec", "5.1.5", "8", ""],
    [76, "环境耐久", "低温工作耐久(LTOE)", "-40°C, 48 hours(24 cycles)", "LTOE -40°C, 48 hours (24 cycles)", "A(一般)", "Filter Spec", "5.1.6", "8", ""],
    [77, "环境耐久", "热湿循环(THC)", "10~65°C, 240 hours, 93%RH", "THC 10-65°C, 240 hours, 93%RH intermittent", "A(一般)", "Filter Spec", "5.1.7", "8", ""],

    # ===== 耐腐蚀/化学 =====
    [78, "耐腐蚀", "腐蚀试验", "180 cycles (CS.00251 + SAE J2334)", "Corrosion 180 cycles per CS.00251, SAE J2334 Immersion", "A(一般)", "Filter Spec", "5.2", "10", "30 cycles后中检，180 cycles后功能全检"],
    [79, "耐腐蚀", "发动机舱流体耐受", "5种流体浸泡后功能正常", "Resistance to engine washing/oil/brake/coolant/battery fluids", "A(一般)", "Filter Spec", "5.3", "10-11", "含22h 110°C烘箱"],

    # ===== 可靠性/耐久 =====
    [80, "可靠性", "耐热循环(总成级)", "20 cycles: Tmax=120°C(机舱)/90°C(底盘), Tmin=-40°C, 2h/2h", "Resistance to Thermal Cycles: 20 cycles 2h@Tmax/2h@Tmin, 2°C/min ramp", "C(关键)", "Filter Spec", "9.3", "35", "含50次排水阀开关测试"],
    [81, "可靠性", "耐脉动压力", "600,000 cycles, 0~700kPa方波, 80°C", "Pulsating Pressure 600,000 cycles 0-700kPa square wave at 80°C per ISO 4020:2001", "C(关键)", "Filter Spec", "9.4", "35-36", ""],
    [82, "可靠性", "耐振动", "Class V2 per CS.00056 (空滤+满油两工况)", "Resistance to Vibrations V2 per CS.00056 (empty + filled with ISO 4113)", "C(关键)", "Filter Spec", "9.5", "36-37", "底盘=V2, 发动机=V1"],
    [83, "可靠性", "燃油老化(动态)", "800h, 100°C环境, 80°C液体, EN590+30%FAME, 每200h换液", "Dynamic Ageing 800h at 100°C env/80°C liquid, EN590+30%FAME", "C(关键)", "Filter Spec", "9.6", "37", "北美：MS-10756+20%生物柴油"],
    [84, "可靠性", "低温功能(含石蜡)", "T = CFPP-12°C下滤芯完整性", "Filter integrity at CFPP-12°C with paraffin clogging", "C(关键)", "Filter Spec", "7.14", "27", ""],
    [85, "可靠性", "加热器寿命", "5000 cycles@14V + 500@18V + 10@24V", "Heater Life: 5000 cycles@14V, 500@18V, 10@24V without failure", "C(关键)", "Filter Spec", "9.7", "37-39", "模拟10年/15万英里"],

    # ===== 外观/物理 =====
    [86, "外观", "外观要求", "无裂纹/缺损/变色/应力痕/凹痕/划伤", "Free from cracks/fractures, missing portions, discolorations, stress marks, dents, scratches", "A(一般)", "Filter Spec", "6.1", "11", "所有测试后检查"],
    [87, "物理", "尺寸符合性", "满足工程图纸及QR-10012", "Meet dimensional requirements in released engineering drawing per QR-10012", "C(关键)", "Filter Spec", "6.2.2", "11", "需工程图纸，当前缺失"],
    [88, "物理", "材料符合性", "满足S材料规格", "All materials conform to applicable S materials spec", "A(一般)", "Filter Spec", "6.2.1", "11", ""],

    # ===== NVH =====
    [89, "NVH", "NVH主观评分", "≥ 8分 (jury evaluation)", "Minimum of 8 subjective rating for jury evaluation per Table 3", "A(一般)", "Filter Spec", "6.3.2", "12", "LP.7R011"],
    [90, "NVH", "消音室噪音", "≤25dB且低于车辆系统噪音3dB", "25 dB max and 3 dB less than vehicle system noise for anechoic chamber", "A(一般)", "Filter Spec", "6.3.2", "12", ""],

    # ===== 安全/法规 =====
    [91, "安全", "法规符合性", "满足所有适用法规", "Compliance to all applicable regulatory requirements in markets where vehicles are sold", "C(关键)", "Filter Spec", "8", "34", ""],
    [92, "安全", "保护帽", "所有接口和电连接器须加保护帽", "Protection caps on all port extremities and electric connector", "A(一般)", "CTS", "3.4", "11", "取消需S品质部同意"],

    # ===== 可维护性 =====
    [93, "可维护性", "免维护里程", "≥ 240,000 km", "Maintenance free ≥ 240,000 km", "C(关键)", "CTS", "5.2.1.5.2", "12", ""],
    [94, "可维护性", "可更换性", "整模块可更换，生产件=售后件", "Replaceable as module. Production parts identical to service parts", "A(一般)", "CTS", "5.2.1.1", "12", ""],
    [95, "可维护性", "拆装耐久性", "≥10次拆装无损伤", "Attachments allow ≥10 remove/refit without damage", "A(一般)", "CTS", "5.2.1.7.2", "13", ""],

    # ===== 验证 — DV/PV =====
    [96, "验证", "功能试验验收标准", "P99C90 (大部分项), DV/PV各15件", "P99C90 acceptance, 15 DV / 15 PV samples for functional tests", "A(一般)", "Filter Spec", "Annex A", "40-42", ""],
    [97, "验证", "耐久试验验收标准", "R95C90, DV/PV各45件", "R95C90 acceptance, 45 DV / 45 PV samples for durability tests", "A(一般)", "Filter Spec", "Annex A", "42", ""],
    [98, "验证", "寿命末期性能保持", "需用-3σ泵保证末期性能", "Performances guaranteed with -3 sigma pump at end of life", "C(关键)", "CTS", "3.1", "9", ""],

    # ===== 电连接器Pinout =====
    [99, "电气接口", "PIN 1: +5V电源", "+5V Power Supply", "Pin 1: +5V Power Supply", "C(关键)", "CTS", "3.3", "10", ""],
    [100, "电气接口", "PIN 2: WiF数字输出", "WiF Digital Output", "Pin 2: WiF Digital Output", "C(关键)", "CTS", "3.3", "10", ""],
    [101, "电气接口", "PIN 3: PTC-", "PTC负极", "Pin 3: PTC -", "C(关键)", "CTS", "3.3", "10", ""],
    [102, "电气接口", "PIN 4: PTC+", "PTC正极", "Pin 4: PTC +", "C(关键)", "CTS", "3.3", "10", ""],
    [103, "电气接口", "PIN 5: SENT协议", "SENT Protocol (NTC温度)", "Pin 5: SENT Protocol", "C(关键)", "CTS", "3.3", "10", ""],
    [104, "电气接口", "PIN 6: 传感器接地", "GROUND (Sensors)", "Pin 6: GROUND (Sensors)", "C(关键)", "CTS", "3.3", "10", ""],
]

# ============================================================
# GAP_DATA — 缺口清单（无则留空）
# 列顺序: [缺失信息, 影响范围, 所需文档, 优先级]
# 优先级: "高" / "中" / "低"
# ============================================================
GAP_DATA = [
    ["CS.00056/CS.00263 电子电气环境规范", "环境试验(TS/PTCE/HTOE等)详细profile和设备分类定义", "CS.00056 或 CS.00263", "高"],
    ["CS.00244 电气及EMC性能要求", "EMC/EMI测试项目和限值", "CS.00244", "高"],
    ["工程图纸 (CAD/Engineering Drawing)", "尺寸要求、关键公差、配合关系", "工程图纸", "高"],
    ["CS.00251 腐蚀要求", "腐蚀试验详细profile", "CS.00251", "中"],
    ["Product Card (PC)", "NTC特性曲线正式值、WiF参数正式值（当前仅为example）", "Product Card", "中"],
    ["PF.90298 快接燃油系统规范", "液压接口详细密封/配合要求", "PF.90298", "中"],
    ["CTS附件 02b_CTS att_K0_SFF", "3D模型构型、安装方式", "CTS附件", "中"],
    ["NTC和WiF传感器具体型号/规格", "传感器详细性能参数", "供应商数据", "中"],
    ["滤芯详细规格（滤材类型/层数/面积）", "滤芯设计优化所需", "供应商内部数据", "低"],
]

# ============================================================
# CONFLICT_DATA — 冲突清单（无则留空）
# 列顺序: [参数, 文档A值, 文档B值, 建议处理]
# ============================================================
CONFLICT_DATA = [
    ["过滤效率 @4μm", "Filter Spec Table 5: ≥90% (初始+平均, ISO 19438)", "CTS §3.1: 98% @4μm (初始, @ 190 l/h)", "以CTS 98%为准；同时保留Spec多粒径要求"],
    ["水分离效率", "Filter Spec §7.12: >95% @ 200l/h (1500ppm+2%水, EN590+30%FAME)", "CTS §3.1: 93% @ 190 l/h", "客户未明确，暂取CTS 93%。可能因流量/FAME条件不同"],
    ["新品压降", "Filter Spec §7.7: ≤30kPa @ 200 l/h", "CTS §3.1: ≤10kPa (ISO 4020)", "可能非冲突——不同工况下的两个指标，建议并行保留"],
    ["E/E环境规范编号", "Filter Spec: CS.00056", "CTS §2.4.2: CS.00263 (01446_24_00439)", "可能是S Harmonized后新编号替代旧版，待客户确认"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
CLASS_C_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
CONFLICT_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_HIGH = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
GAP_MID = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_LOW = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

L1_HEADERS = ["ID", "特性分类", "工程特性名称", "目标值/要求", "原文描述", "类别", "文件来源", "章节", "页码", "备注"]
L1_WIDTHS = [6, 15, 40, 30, 50, 10, 20, 10, 6, 30]


def apply_header(ws, headers, widths):
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"


def build_l1(wb):
    ws = wb.active
    ws.title = "工程特性清单"
    apply_header(ws, L1_HEADERS, L1_WIDTHS)
    for row in L1_DATA:
        ws.append(row)
    for r in range(2, len(L1_DATA) + 2):
        cls = ws.cell(r, 6).value  # 类别列
        note = str(ws.cell(r, 10).value or "")  # 备注列
        for c in range(1, len(L1_HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP_TOP
            if "[冲突]" in note:
                cell.fill = CONFLICT_FILL
            elif str(cls).startswith("C"):
                cell.fill = CLASS_C_FILL


def build_gaps(wb):
    ws = wb.create_sheet("缺口与冲突")

    # 缺口部分
    ws.cell(1, 1, "— 缺口 —").font = Font(bold=True, size=12)
    gap_headers = ["缺失信息", "影响范围", "所需文档", "优先级"]
    gap_widths = [40, 30, 30, 10]
    for c, (h, w) in enumerate(zip(gap_headers, gap_widths), 1):
        cell = ws.cell(2, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = w

    priority_fills = {"高": GAP_HIGH, "中": GAP_MID, "低": GAP_LOW}
    for i, row in enumerate(GAP_DATA, 3):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
        p = row[3] if len(row) > 3 else ""
        if p in priority_fills:
            for c in range(1, len(gap_headers) + 1):
                ws.cell(i, c).fill = priority_fills[p]

    # 冲突部分
    conflict_start = len(GAP_DATA) + 5
    ws.cell(conflict_start, 1, "— 冲突 —").font = Font(bold=True, size=12)
    conflict_headers = ["参数", "文档A值", "文档B值", "建议处理"]
    conflict_widths = [30, 30, 30, 30]
    for c, (h, w) in enumerate(zip(conflict_headers, conflict_widths), 1):
        cell = ws.cell(conflict_start + 1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = max(
            ws.column_dimensions[get_column_letter(c)].width or 0, w
        )
    for i, row in enumerate(CONFLICT_DATA, conflict_start + 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
            cell.fill = CONFLICT_FILL


def main():
    wb = openpyxl.Workbook()
    build_l1(wb)
    if GAP_DATA or CONFLICT_DATA:
        build_gaps(wb)
    wb.save(CONFIG["输出路径"])
    c_count = sum(1 for r in L1_DATA if len(r) > 5 and str(r[5]).startswith("C"))
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"L1特性: {len(L1_DATA)} 条 (C类:{c_count}, A类:{len(L1_DATA)-c_count})")
    if GAP_DATA:
        print(f"缺口: {len(GAP_DATA)} 项")
    if CONFLICT_DATA:
        print(f"冲突: {len(CONFLICT_DATA)} 项")


if __name__ == "__main__":
    main()
