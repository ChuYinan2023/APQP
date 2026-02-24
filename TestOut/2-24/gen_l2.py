"""
L2零部件特性清单生成器
使用方法：复制到output/gen_l2.py，填写数据区，运行 python3 gen_l2.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "阶段3_L2零部件特性清单.xlsx",
}

# ============================================================
# L2_DATA — 修改此区域
# 列顺序: [ID, 零件/组件名称, 零件特性, 目标值/要求, 对应L1 ID, 文件来源, 章节, 备注]
# ID格式: PC-01, PC-02...（顺序编号）
# 按零部件分组排列
# ============================================================
L2_DATA = [
    # ===== 壳体 (Housing) =====
    ["PC-01", "壳体", "材料符合性", "满足S材料规格", "88", "Filter Spec", "6.2.1", "具体材料牌号需图纸确认"],
    ["PC-02", "壳体", "尺寸符合性", "满足工程图纸及QR-10012", "87", "Filter Spec", "6.2.2", "需工程图纸，当前缺失"],
    ["PC-03", "壳体", "爆破压力", "≥20 bar", "32", "Filter Spec", "7.10", "ISO 4020"],
    ["PC-04", "壳体", "外观", "无裂纹/缺损/变色/应力痕/凹痕/划伤", "86", "Filter Spec", "6.1", ""],
    ["PC-05", "壳体", "耐腐蚀性", "180 cycles (CS.00251 + SAE J2334)无锈蚀", "78", "Filter Spec", "5.2", "30 cycles后中检，180 cycles后功能全检"],
    ["PC-06", "壳体", "耐化学性", "5种发动机舱流体浸泡后完整", "79", "Filter Spec", "5.3", ""],
    ["PC-07", "壳体", "零件标识", "含品牌/零件号/回收标记/材料缩写/供应商号", "86", "Filter Spec", "6.3.3", ""],
    ["PC-08", "壳体", "耐脉动压力", "600,000 cycles @ 0~700kPa方波, 80°C", "81", "Filter Spec", "9.4", "ISO 4020:2001"],
    ["PC-09", "壳体", "耐热循环", "20 cycles: Tmax=120°C/-40°C", "80", "Filter Spec", "9.3", ""],
    ["PC-10", "壳体", "防护等级", "IP6K9K", "64", "Filter Spec", "5.1/Table4", ""],

    # ===== 滤芯 (Filter Element) =====
    ["PC-11", "滤芯", "过滤效率(初始) @4μm", "≥98% @ 190 l/h", "9", "CTS", "3.1", ""],
    ["PC-12", "滤芯", "过滤效率(初始+平均) @4μm", "≥90%", "10", "Filter Spec", "7.1/Table5", "ISO 19438"],
    ["PC-13", "滤芯", "过滤效率(初始+平均) @5μm", "≥95%", "11", "Filter Spec", "7.1/Table5", ""],
    ["PC-14", "滤芯", "过滤效率(初始+平均) @10μm", "≥99%", "12", "Filter Spec", "7.1/Table5", ""],
    ["PC-15", "滤芯", "过滤效率(初始+平均) @15μm", "≥99%", "13", "Filter Spec", "7.1/Table5", ""],
    ["PC-16", "滤芯", "过滤效率(初始+平均) @20μm", "≥99.9%", "14", "Filter Spec", "7.1/Table5", ""],
    ["PC-17", "滤芯", "过滤效率(初始+平均) @30μm", "≥99.9%", "15", "Filter Spec", "7.1/Table5", ""],
    ["PC-18", "滤芯", "单通过滤粒子保持率 @5μm", "≥95%", "16", "Filter Spec", "7.2", "ISO 19438"],
    ["PC-19", "滤芯", "容尘量(DHC)", "≥30g @ 500kPa寿命末期压降", "17", "CTS", "3.1", "ISO 4020"],
    ["PC-20", "滤芯", "新品压降 @ 190 l/h", "≤10 kPa", "24", "CTS", "3.1", "ISO 4020"],
    ["PC-21", "滤芯", "新品压降 @ 200 l/h", "≤30 kPa", "25", "Filter Spec", "7.7", ""],
    ["PC-22", "滤芯", "塌陷压力", ">800 kPa (116 psi)", "31", "Filter Spec", "7.9", "ISO 4020"],
    ["PC-23", "滤芯", "内部密封(气泡试验)", "出口泄漏 < 100 cm³/min @ 1.3±0.1 kPa", "30", "Filter Spec", "7.6", "ISO 2942"],
    ["PC-24", "滤芯", "清洁度(总重)", "< 0.5 mg/part", "36", "Filter Spec", "7.8", ""],
    ["PC-25", "滤芯", "清洁度 — 颗粒 >200~500μm", "0个", "39", "Filter Spec", "7.8/Table6", ""],
    ["PC-26", "滤芯", "耐石蜡堵塞性", "CFPP-12°C下完整性保持", "84", "Filter Spec", "7.14", ""],
    ["PC-27", "滤芯", "燃油老化耐受", "800h @ 100°C环境, EN590+30%FAME", "83", "Filter Spec", "9.6", "每200h换液"],
    ["PC-28", "滤芯", "滤材类型/层数/面积", "待确认", "9,17", "—", "—", "需供应商内部数据"],

    # ===== 集水碗/水分离器 (Water Separator Bowl) =====
    ["PC-29", "集水碗", "水分离效率(ISO 16332)", "≥93% @ 190 l/h", "19", "CTS", "3.1", ""],
    ["PC-30", "集水碗", "水分离效率(通用Spec)", ">95% @ 200 l/h, EN590+30%FAME", "20", "Filter Spec", "7.12", ""],
    ["PC-31", "集水碗", "乳化水分离效率", ">1500 PPM", "23", "Filter Spec", "7.17", "IFT oF15"],
    ["PC-32", "集水碗", "储水容积", "150 ml", "21", "CTS", "3.1", ""],
    ["PC-33", "集水碗", "水位报警量", "100 ml", "22", "CTS", "3.1", "与WiF传感器配合"],
    ["PC-34", "集水碗", "排水阀耐久", "50次开关 (热循环试验中)", "80", "Filter Spec", "9.3", ""],
    ["PC-35", "集水碗", "材料", "满足S材料规格", "88", "Filter Spec", "6.2.1", "需图纸确认具体材料"],

    # ===== PTC加热器 (PTC Heater) =====
    ["PC-36", "PTC加热器", "最低效率", "η ≥ 0.75 (6分钟后)", "40", "Filter Spec", "6.4.10.1", "蒸馏水,130l/h,12.5V"],
    ["PC-37", "PTC加热器", "最大峰值电流", "40A", "41", "Filter Spec", "6.4.10.2", "CTS示例值"],
    ["PC-38", "PTC加热器", "峰值持续时间", "≤2.5s", "42", "Filter Spec", "6.4.10.2", ""],
    ["PC-39", "PTC加热器", "最大工作电流", "30A", "43", "Filter Spec", "6.4.10.2", ""],
    ["PC-40", "PTC加热器", "最小ΔT", "≥3°C (6分钟后,130l/h,12.5V,-26°C)", "44", "Filter Spec", "6.4.10.2", "Arctic Diesel Class 3"],
    ["PC-41", "PTC加热器", "恒温器关断 — 满油", "< 10.0 分钟", "45", "Filter Spec", "6.4.9", "14V供电"],
    ["PC-42", "PTC加热器", "恒温器关断 — 空滤", "< 3.0 分钟", "46", "Filter Spec", "6.4.9", "14V供电"],
    ["PC-43", "PTC加热器", "失效安全性", "恒温器闭合失效时无外部/内部损伤", "47", "Filter Spec", "6.4.9", ""],
    ["PC-44", "PTC加热器", "恒温器开关信号质量", "尖锐无抖动，可重复温度", "48", "Filter Spec", "7.18", "10次重复"],
    ["PC-45", "PTC加热器", "寿命", "5000 cycles@14V + 500@18V + 10@24V", "85", "Filter Spec", "9.7", "模拟10年/15万英里"],
    ["PC-46", "PTC加热器", "低温工况性能", "≤400 kPa @ 最小流量 (12.5V和8V)", "26", "Filter Spec", "7.13", ""],
    ["PC-47", "PTC加热器", "耐腐蚀后功能", "180 cycles腐蚀后正常工作", "78", "Filter Spec", "5.2", ""],
    ["PC-48", "PTC加热器", "PIN 3/4 接线", "PIN 3: PTC-, PIN 4: PTC+", "101,102", "CTS", "3.3", ""],

    # ===== NTC温度传感器 (NTC Sensor) =====
    ["PC-49", "NTC温度传感器", "工作温度范围", "-40°C ~ 125°C", "49", "Filter Spec", "7.19", ""],
    ["PC-50", "NTC温度传感器", "供电电压", "5V ±2% (ECU供电)", "50", "Filter Spec", "7.19", "示例值，正式值在Product Card"],
    ["PC-51", "NTC温度传感器", "特性曲线 (25°C标称)", "2.057 kΩ (1.941~2.173)", "52", "Filter Spec", "7.19/Table7", "示例值"],
    ["PC-52", "NTC温度传感器", "温度测量精度 @80°C", "T4-T2 < 15°C (300s后, 200 l/h)", "51", "Filter Spec", "7.20", ""],
    ["PC-53", "NTC温度传感器", "SENT协议通信", "PIN 5: SENT Protocol", "53", "CTS", "3.3", ""],
    ["PC-54", "NTC温度传感器", "耐腐蚀后功能", "180 cycles腐蚀后正常工作", "78", "Filter Spec", "5.2", ""],

    # ===== WiF水位传感器 (WiF Sensor) =====
    ["PC-55", "WiF水位传感器", "传感器类型", "主动式(Active)", "54", "Filter Spec", "7.21", "示例值"],
    ["PC-56", "WiF水位传感器", "正常状态电压", "V out > 5.5V", "55", "Filter Spec", "7.21", ""],
    ["PC-57", "WiF水位传感器", "报警状态电压", "V out < 2.2V", "56", "Filter Spec", "7.21", ""],
    ["PC-58", "WiF水位传感器", "最大吸收电流", "< 50 mA", "57", "Filter Spec", "7.21", ""],
    ["PC-59", "WiF水位传感器", "自检时间 (-30~110°C)", "3s ±20% (2.4~3.6s)", "58", "Filter Spec", "7.21", ""],
    ["PC-60", "WiF水位传感器", "自检时间 (-40~125°C)", "1.7 ~ 4.5s", "59", "Filter Spec", "7.21", ""],
    ["PC-61", "WiF水位传感器", "供电电压范围", "10V ~ 16V (A1分类)", "60", "Filter Spec", "7.21", ""],
    ["PC-62", "WiF水位传感器", "工作温度范围", "-40°C ~ 125°C", "61", "Filter Spec", "7.21", ""],
    ["PC-63", "WiF水位传感器", "探针抗电解", "168h @ 14V后仍正常工作", "62", "Filter Spec", "6.4.6", ""],
    ["PC-64", "WiF水位传感器", "水位报警触发量", "100 ml时触发信号", "22", "CTS", "3.1", "与集水碗配合"],
    ["PC-65", "WiF水位传感器", "PIN 2 接线", "PIN 2: WiF Digital Output", "100", "CTS", "3.3", ""],
    ["PC-66", "WiF水位传感器", "耐腐蚀后功能", "180 cycles腐蚀后正常工作", "78", "Filter Spec", "5.2", ""],

    # ===== 电连接器 (Electrical Connector) =====
    ["PC-67", "电连接器", "型号", "APTIV p/n 211FYT0523, KEY CODE A", "63", "CTS", "3.3", "6-pin"],
    ["PC-68", "电连接器", "PIN 1: +5V电源", "+5V Power Supply", "99", "CTS", "3.3", ""],
    ["PC-69", "电连接器", "PIN 6: 传感器接地", "GROUND (Sensors)", "104", "CTS", "3.3", ""],
    ["PC-70", "电连接器", "防护等级", "IP6K9K", "64", "Filter Spec", "5.1/Table4", ""],
    ["PC-71", "电连接器", "ESD分类", "SE2", "65", "Filter Spec", "5.1/Table4", ""],
    ["PC-72", "电连接器", "接线要求", "满足CS.00050/PF.90303/PF.90012/SAE USCAR-21", "63", "Filter Spec", "6.4.1", ""],
    ["PC-73", "电连接器", "连接器标准", "满足PF.90012/SAE USCAR-25", "63", "Filter Spec", "6.4.2", ""],

    # ===== 液压快接 (Hydraulic Quick Connect) =====
    ["PC-74", "液压快接-进口", "类型及尺寸", "10mm 快接, SAE J2044 / B12 2813", "68", "CTS", "3.2", ""],
    ["PC-75", "液压快接-出口", "类型及尺寸", "10mm 快接, SAE J2044 / B12 2813", "69", "CTS", "3.2", ""],
    ["PC-76", "液压快接", "尺寸符合性", "满足SAE J2044尺寸要求", "70", "Filter Spec", "6.4.7", ""],
    ["PC-77", "液压快接", "金属接头压装扭矩", "≥7.3 Nm不脱落", "33", "Filter Spec", "7.11.1", ""],
    ["PC-78", "液压快接", "塑料接头侧向载荷", "满足SAE J2044 §6.4", "34", "Filter Spec", "7.11.2", ""],
    ["PC-79", "液压快接", "静电电荷防护", "满足SAE J1645", "67", "Filter Spec", "7.16", ""],

    # ===== 密封件 (Seals/Gaskets) =====
    ["PC-80", "密封件", "外部气密性(空气)", "泄漏 < 3 cm³/min @ -40/23/80°C", "27", "Filter Spec", "7.4.1", "最大工作压力下"],
    ["PC-81", "密封件", "外部密封性(燃油)", "无泄漏 @ 10.3 bar @ -40/23/80°C", "28", "Filter Spec", "7.4.2", "EN590 Arctic fuel class 4"],
    ["PC-82", "密封件-滤芯垫片", "脏侧/净侧密封", "泄漏 ≤ 3 cm³/min @ ≥500 kPa", "29", "Filter Spec", "7.5", "多温度+FAME老化后复测"],
    ["PC-83", "密封件", "O型圈规格", "满足S 14618标准", "27,28", "CTS", "2.4.2", ""],
    ["PC-84", "密封件", "燃油老化后密封性", "800h EN590+30%FAME老化后密封保持", "83", "Filter Spec", "9.6", ""],

    # ===== 保护帽 (Protection Caps) =====
    ["PC-85", "保护帽", "覆盖范围", "所有接口端和电连接器", "92", "CTS", "3.4", "取消需S品质部同意"],
    ["PC-86", "保护帽", "形状规格", "满足protection cap.xlsx", "92", "CTS", "3.4", "文件缺失"],
]

# ============================================================
# GAP_DATA — 缺口清单（无则留空）
# 列顺序: [缺失信息, 影响L2 ID, 关联L1 ID, 所需文档, 优先级]
# ============================================================
GAP_DATA = [
    # 高优先级
    ["壳体/滤芯/集水碗尺寸公差", "PC-02", "87", "工程图纸", "高"],
    ["壳体具体材料牌号", "PC-01", "88", "工程图纸", "高"],
    ["EMC/EMI测试限值", "PC-67~73", "64,65", "CS.00244", "高"],
    ["环境试验详细profile（振动/热冲击等）", "PC-10,PC-09", "71~77,82", "CS.00056/CS.00263", "高"],
    # 中优先级
    ["NTC特性曲线正式值（当前为示例值）", "PC-51", "52", "Product Card", "中"],
    ["WiF传感器正式参数（当前为示例值）", "PC-55~63", "54~61", "Product Card", "中"],
    ["液压快接详细密封/配合要求", "PC-74~76", "68,69,70", "PF.90298", "中"],
    ["集水碗材料牌号", "PC-35", "88", "工程图纸", "中"],
    ["腐蚀试验详细cycle参数", "PC-05,PC-47,PC-54,PC-66", "78", "CS.00251", "中"],
    ["3D模型构型及安装方式", "PC-02", "87", "CTS附件02b_CTS att_K0_SFF", "中"],
    # 低优先级
    ["滤材类型/层数/过滤面积", "PC-28", "9,17", "供应商内部数据", "低"],
    ["保护帽形状和规格", "PC-86", "92", "protection cap.xlsx", "低"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
GAP_HIGH = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
GAP_MID = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_LOW = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

L2_HEADERS = ["ID", "零件/组件名称", "零件特性", "目标值/要求", "对应L1 ID", "文件来源", "章节", "备注"]
L2_WIDTHS = [8, 20, 40, 30, 15, 20, 10, 25]


def apply_header(ws, headers, widths):
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"


def build_l2(wb):
    ws = wb.active
    ws.title = "零件特性清单"
    apply_header(ws, L2_HEADERS, L2_WIDTHS)
    for row in L2_DATA:
        ws.append(row)
    for r in range(2, len(L2_DATA) + 2):
        for c in range(1, len(L2_HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP_TOP


def build_gaps(wb):
    ws = wb.create_sheet("缺口与冲突")
    headers = ["缺失信息", "影响L2 ID", "关联L1 ID", "所需文档", "优先级"]
    widths = [40, 15, 15, 30, 10]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    priority_fills = {"高": GAP_HIGH, "中": GAP_MID, "低": GAP_LOW}
    for i, row in enumerate(GAP_DATA, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
        p = row[4] if len(row) > 4 else ""
        if p in priority_fills:
            for c in range(1, len(headers) + 1):
                ws.cell(i, c).fill = priority_fills[p]


def main():
    wb = openpyxl.Workbook()
    build_l2(wb)
    if GAP_DATA:
        build_gaps(wb)
    wb.save(CONFIG["输出路径"])
    parts = {}
    for row in L2_DATA:
        parts[row[1]] = parts.get(row[1], 0) + 1
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"L2特性: {len(L2_DATA)} 条, {len(parts)} 个零部件")
    if GAP_DATA:
        print(f"缺口: {len(GAP_DATA)} 项")


if __name__ == "__main__":
    main()
