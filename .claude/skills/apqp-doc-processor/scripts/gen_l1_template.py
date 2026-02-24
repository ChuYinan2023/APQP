"""
L1工程特性清单生成器 —— 标准模板 v2.0
使用说明：
  1. 填写 CONFIG 区（项目元数据）
  2. 填写 L1_DATA 区（特性数据行）
  3. 填写 CC_DATA 区（持续符合性数据，无则留空列表）
  4. 填写 RV_DATA 区（反向校验数据）
  5. 运行脚本：python gen_l1.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "零部件名称": "",        # 例: "柴油滤清器"
    "OEM":        "",        # 例: "Stellantis"
    "来源文档":   "",        # 例: "PF.90150 + CTS"
    "提取日期":   "",        # 例: "2026-02-21"
    "输出路径":   "阶段3_L1工程特性清单.xlsx",  # 相对于脚本所在目录
}

# ============================================================
# L1_DATA — 修改此区域
# 列顺序: [ID, 特性分类, 工程特性名称, 目标值/要求, 文件名称, 层级, 页码, 章节, 图表, 原文描述, 类别, 冲突项ID, 冲突/差异说明, 分析建议, 客户裁决]
# ============================================================
L1_DATA = [
    # 示例行（删除后填入真实数据）:
    # ["F-01", "功能", "过滤效率 ≥4μm(c)", "≥98%", "CTS 3.1", "CTS", "12", "7.1", "Table 3", "Filtration efficiency...", "C", "", "", "", ""],
]

# ============================================================
# CC_DATA — Annex CC 持续符合性（无则留空）
# 列顺序: [测试项目, 规范章节, 样本量, 接收准则, 频次]
# ============================================================
CC_DATA = [
    # ["过滤效率", "7.1", "6", "无失效", "每季"],
]

# ============================================================
# RV_DATA — 反向校验
# 列顺序: [章节, 标题, 覆盖状态, L1 ID, 备注]
# 覆盖状态: "✓已提取" / "未涉及" / "[缺口]"
# ============================================================
RV_DATA = [
    # ["1", "General", "✓已提取", "F-01", ""],
]

# ============================================================
# 以下为格式代码，不需要修改
# ============================================================

# ----- 样式定义（固定，不修改）-----
HEADER_FONT  = Font(bold=True, size=11)
HEADER_FILL  = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
TOTAL_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GAP_FILL     = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
# 类别着色: C类=粉红, A类=无色
CLASS_C_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_header(ws, headers, col_widths, row_height=30):
    ws.append(headers)
    ws.row_dimensions[1].height = row_height
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"


def apply_data_border(ws, start_row, end_row, ncols, class_col=None):
    for r in range(start_row, end_row + 1):
        cls = ws.cell(r, class_col).value if class_col else None
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP_TOP
            if cls and str(cls).upper() == "C":
                cell.fill = CLASS_C_FILL


def build_cover(wb):
    ws = wb.active
    ws.title = "封面"
    fields = [
        ("零部件名称", CONFIG["零部件名称"]),
        ("OEM",        CONFIG["OEM"]),
        ("来源文档",   CONFIG["来源文档"]),
        ("提取日期",   CONFIG["提取日期"]),
        ("阶段",       "3 - L1工程特性清单"),
    ]
    for r, (k, v) in enumerate(fields, 1):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2, v).border = THIN
        ws.cell(r, 2).alignment = WRAP_TOP
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 70


def build_l1(wb):
    ws = wb.create_sheet("工程特性清单")
    headers    = ["ID", "特性分类", "工程特性名称", "目标值/要求", "文件名称", "层级", "页码", "章节", "图表", "原文描述", "类别", "冲突项ID", "冲突/差异说明", "分析建议", "客户裁决"]
    col_widths = [8,    15,         50,              30,            40,         12,     8,      12,     12,     60,         8,      15,          30,              30,         20]
    apply_header(ws, headers, col_widths)
    for row in L1_DATA:
        ws.append(row)
    # class_col=11 is the "类别" column (K)
    apply_data_border(ws, 2, len(L1_DATA) + 1, len(headers), class_col=11)


def build_cc(wb):
    ws = wb.create_sheet("CC持续符合性")
    headers    = ["测试项目", "规范章节", "样本量", "接收准则", "频次"]
    col_widths = [35,         14,          10,        20,          15]
    apply_header(ws, headers, col_widths)
    for row in CC_DATA:
        ws.append(row)
    apply_data_border(ws, 2, len(CC_DATA) + 1, len(headers))


def build_rv(wb):
    ws = wb.create_sheet("反向校验")
    headers    = ["章节",  "标题",   "覆盖状态",  "L1 ID",  "备注"]
    col_widths = [18,       30,        12,           22,        35]
    apply_header(ws, headers, col_widths)
    for row in RV_DATA:
        ws.append(row)
    apply_data_border(ws, 2, len(RV_DATA) + 1, len(headers))
    # 高亮缺口行
    for r in range(2, len(RV_DATA) + 2):
        status_cell = ws.cell(r, 3)
        if status_cell.value and "[缺口]" in str(status_cell.value):
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = GAP_FILL


def build_stats(wb):
    ws = wb.create_sheet("统计汇总")
    headers    = ["类别", "数量"]
    col_widths = [12,      10]
    apply_header(ws, headers, col_widths)
    counts = {"C": 0, "A": 0}
    for row in L1_DATA:
        cls = str(row[10]).upper() if len(row) > 10 and row[10] else ""
        if cls in counts:
            counts[cls] += 1
    for r, (code, cnt) in enumerate(counts.items(), 2):
        ws.cell(r, 1, code).border = THIN
        ws.cell(r, 2, cnt).border = THIN
    total_r = len(counts) + 2
    for c, v in enumerate(["合计", sum(counts.values())], 1):
        cell = ws.cell(total_r, c, v)
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN


def main():
    wb = openpyxl.Workbook()
    build_cover(wb)
    build_l1(wb)
    if CC_DATA:
        build_cc(wb)
    build_rv(wb)
    build_stats(wb)
    wb.save(CONFIG["输出路径"])
    counts = {"C": 0, "A": 0}
    for row in L1_DATA:
        cls = str(row[10]).upper() if len(row) > 10 and row[10] else ""
        if cls in counts:
            counts[cls] += 1
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"L1特性合计: {len(L1_DATA)} 条")
    print(f"  C类(关键): {counts['C']}  A类(一般): {counts['A']}")


if __name__ == "__main__":
    main()
