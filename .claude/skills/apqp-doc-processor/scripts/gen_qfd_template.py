"""
QFD质量屋矩阵生成器 —— 标准模板 v2.0
使用说明：
  1. 填写 CONFIG 区
  2. 填写 L1_ITEMS / L2_ITEMS 区（从已生成的L1/L2清单导入标题）
  3. 填写 MATRIX 区（L1×L2关联强度）
  4. 填写 ROOF 区（L2×L2屋顶矩阵）
  5. 运行脚本：python gen_qfd.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "零部件名称": "",
    "OEM":        "",
    "来源文档":   "",
    "提取日期":   "",
    "输出路径":   "阶段5_QFD质量屋矩阵.xlsx",
}

# ============================================================
# L1_ITEMS — L1工程特性列表
# 格式: [(ID, 特性名称), ...]
# ============================================================
L1_ITEMS = [
    # ("F-01", "过滤效率 ≥4μm(c)"),
]

# ============================================================
# L2_ITEMS — L2零部件特性列表
# 格式: [(L2_ID, 零部件名称, 特性名称, 目标值), ...]
# ============================================================
L2_ITEMS = [
    # ("L2-1.01", "壳体", "材料牌号", "PA66-GF30"),
]

# ============================================================
# MATRIX — L1×L2关联矩阵
# 格式: {(L1_ID, L2_ID): 关联强度}
# 关联强度: 9=强(◎) / 3=中(○) / 1=弱(△) / 不填=无关联
# 只填有关联的格子，无关联不需要写
# ============================================================
MATRIX = {
    # ("F-01", "L2-2.01"): 9,   # 过滤效率 ← 滤芯材料（强关联）
    # ("F-01", "L2-1.01"): 3,   # 过滤效率 ← 壳体材料（中关联）
}

# ============================================================
# ROOF — L2×L2屋顶矩阵（相关性）
# 格式: {(L2_ID_A, L2_ID_B): 相关性}
# 相关性: "+" = 正相关（协同）/ "-" = 负相关（冲突）
# 只填有相关的组合，且A < B（按L2_ITEMS顺序）
# ============================================================
ROOF = {
    # ("L2-2.01", "L2-2.02"): "+",   # 滤芯材料 与 滤芯面积 正相关
}

# ============================================================
# 以下为格式代码，不需要修改
# ============================================================

# 关联符号（无背景色）
SCORE_MAP  = {9: "◎", 3: "○", 1: "△"}
ROOF_FILL = {
    "+": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "-": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}
# 标题区样式
TITLE_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT   = Font(bold=True, size=12, color="FFFFFF")
# L2列表头样式
L2_ROW1_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # 零件名称行
L2_ROW1_FONT = Font(bold=True, size=10)
L2_ROW2_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # PC-ID + 特性描述行
L1_ID_FILL   = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # L1 ID列底色
# 辅助Sheet表头
HEADER_FONT  = Font(bold=True, size=10)
HEADER_FILL  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin'),
)
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')


def build_cover(wb):
    ws = wb.active
    ws.title = "封面"
    fields = [
        ("零部件名称", CONFIG["零部件名称"]),
        ("OEM",        CONFIG["OEM"]),
        ("来源文档",   CONFIG["来源文档"]),
        ("提取日期",   CONFIG["提取日期"]),
        ("阶段",       "5 - QFD质量屋矩阵"),
        ("L1特性数量", str(len(L1_ITEMS))),
        ("L2特性数量", str(len(L2_ITEMS))),
        ("矩阵规模",   f"{len(L1_ITEMS)} × {len(L2_ITEMS)}"),
    ]
    for r, (k, v) in enumerate(fields, 1):
        ws.cell(r, 1, k).font   = Font(bold=True)
        ws.cell(r, 1).border    = THIN
        ws.cell(r, 2, v).border = THIN
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50


def build_matrix(wb):
    ws = wb.create_sheet("QFD矩阵")
    n_l1 = len(L1_ITEMS)
    n_l2 = len(L2_ITEMS)

    # 固定列: L1 ID(A), L1特性名称(B), 空白分隔(C)，L2列从 col=4(D) 开始
    L2_START_COL = 4

    # --- 标题区: A1:C3 合并 ---
    ws.merge_cells("A1:C3")
    title_cell = ws.cell(1, 1, "QFD第二阶段质量屋\n工程特性 → 零件特性")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    title_cell.border = THIN
    # 给合并区域加边框
    for r in range(1, 4):
        for c in range(1, 4):
            ws.cell(r, c).border = THIN

    # --- L2列3行表头 ---
    # 按零件名称分组，确定每个零件的起始列和列数
    part_groups = {}  # {零件名: [col_indices]}
    for j, (l2id, part, name, target) in enumerate(L2_ITEMS):
        col = L2_START_COL + j
        part_groups.setdefault(part, []).append(col)
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Row 1: 零件名称（合并同一零件的列）
    for part, cols in part_groups.items():
        start_col = min(cols)
        end_col = max(cols)
        if start_col != end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col, part)
        cell.font = L2_ROW1_FONT
        cell.fill = L2_ROW1_FILL
        cell.alignment = CENTER
        cell.border = THIN
        # 边框 for all merged cells
        for c in cols:
            ws.cell(1, c).border = THIN
            ws.cell(1, c).fill = L2_ROW1_FILL

    # Row 2: PC-ID + 特性描述
    for j, (l2id, part, name, target) in enumerate(L2_ITEMS):
        col = L2_START_COL + j
        cell = ws.cell(2, col, f"{l2id}\n{name}")
        cell.fill = L2_ROW2_FILL
        cell.alignment = CENTER
        cell.border = THIN

    # Row 3: 目标值
    for j, (l2id, part, name, target) in enumerate(L2_ITEMS):
        col = L2_START_COL + j
        cell = ws.cell(3, col, target)
        cell.alignment = CENTER
        cell.border = THIN

    # L1固定列表头 (rows 1-3 already used for title merge in A:C)
    # Row 1-3 A:C is title area, no separate L1 header needed there

    # --- 数据行 (from row 4) ---
    DATA_START_ROW = 4
    for i, (l1id, name) in enumerate(L1_ITEMS):
        row = DATA_START_ROW + i
        # L1 ID column
        id_cell = ws.cell(row, 1, l1id)
        id_cell.alignment = CENTER
        id_cell.fill = L1_ID_FILL
        id_cell.border = THIN
        # L1 特性名称
        name_cell = ws.cell(row, 2, name)
        name_cell.alignment = WRAP_TOP
        name_cell.border = THIN
        # 空白分隔列
        ws.cell(row, 3).border = THIN

        # 关联格
        for j, (l2id, _, _, _) in enumerate(L2_ITEMS):
            col = L2_START_COL + j
            score = MATRIX.get((l1id, l2id), 0)
            cell = ws.cell(row, col)
            cell.alignment = CENTER
            cell.border    = THIN
            if score:
                cell.value = SCORE_MAP[score]

    # 固定列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 2

    # 冻结窗格: D4
    ws.freeze_panes = "D4"


def build_roof(wb):
    ws = wb.create_sheet("屋顶矩阵")
    n = len(L2_ITEMS)

    # 标题行
    ws.cell(1, 1, "L2 ID").font    = HEADER_FONT
    ws.cell(1, 1).fill             = HEADER_FILL
    ws.cell(1, 1).alignment        = CENTER
    ws.cell(1, 1).border           = THIN
    for j, (l2id, part, name, target) in enumerate(L2_ITEMS):
        col = j + 2
        ws.cell(1, col, f"{l2id}\n{name}").font = HEADER_FONT
        ws.cell(1, col).fill       = HEADER_FILL
        ws.cell(1, col).alignment  = CENTER
        ws.cell(1, col).border     = THIN
        ws.column_dimensions[get_column_letter(col)].width = 12

    # 数据行
    for i, (l2id_a, _, name_a, _) in enumerate(L2_ITEMS):
        row = i + 2
        ws.cell(row, 1, f"{l2id_a}\n{name_a}").alignment = WRAP_TOP
        ws.cell(row, 1).border = THIN
        ws.column_dimensions['A'].width = 20
        for j, (l2id_b, _, _, _) in enumerate(L2_ITEMS):
            col = j + 2
            cell = ws.cell(row, col)
            cell.border = THIN
            cell.alignment = CENTER
            if i == j:  # 对角线
                cell.fill = HEADER_FILL
            elif i < j:
                rel = ROOF.get((l2id_a, l2id_b), "")
                if rel:
                    cell.value = rel
                    cell.fill  = ROOF_FILL.get(rel, PatternFill())
                    cell.font  = Font(bold=True)


def build_legend(wb):
    ws = wb.create_sheet("判定依据")
    ws.cell(1, 1, "符号").font  = HEADER_FONT
    ws.cell(1, 1).fill          = HEADER_FILL
    ws.cell(1, 1).border        = THIN
    ws.cell(1, 2, "分值").font  = HEADER_FONT
    ws.cell(1, 2).fill          = HEADER_FILL
    ws.cell(1, 2).border        = THIN
    ws.cell(1, 3, "判定标准").font = HEADER_FONT
    ws.cell(1, 3).fill             = HEADER_FILL
    ws.cell(1, 3).border           = THIN
    legend = [
        ("◎", "9", "强关联：L2是L1的直接技术实现，L2变化直接决定L1达标"),
        ("○", "3", "中关联：L2对L1有间接贡献，是达标的辅助条件之一"),
        ("△", "1", "弱关联：L2对L1有微弱影响，仅极端条件下关联"),
        ("(空)", "0", "无关联"),
        ("+", "-", "屋顶正相关：两个L2特性协同，同向变化"),
        ("-", "-", "屋顶负相关：两个L2特性冲突，需设计平衡"),
    ]
    for r, (sym, score, desc) in enumerate(legend, 2):
        ws.cell(r, 1, sym).border    = THIN
        ws.cell(r, 1).alignment      = CENTER
        ws.cell(r, 2, score).border  = THIN
        ws.cell(r, 2).alignment      = CENTER
        ws.cell(r, 3, desc).border   = THIN
        ws.cell(r, 3).alignment      = WRAP_TOP
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 60


def build_review(wb):
    ws = wb.create_sheet("全局审查")
    ws.cell(1, 1, "审查项").font    = HEADER_FONT
    ws.cell(1, 1).fill              = HEADER_FILL
    ws.cell(1, 1).border            = THIN
    ws.cell(1, 2, "结果").font      = HEADER_FONT
    ws.cell(1, 2).fill              = HEADER_FILL
    ws.cell(1, 2).border            = THIN
    ws.cell(1, 3, "说明").font      = HEADER_FONT
    ws.cell(1, 3).fill              = HEADER_FILL
    ws.cell(1, 3).border            = THIN

    # 自动计算空行/空列
    l1_ids = [item[0] for item in L1_ITEMS]
    l2_ids = [item[0] for item in L2_ITEMS]

    empty_l1 = [l1id for l1id in l1_ids if not any(MATRIX.get((l1id, l2id)) for l2id in l2_ids)]
    empty_l2 = [l2id for l2id in l2_ids if not any(MATRIX.get((l1id, l2id)) for l1id in l1_ids)]

    l2_scores = {
        l2id: sum(MATRIX.get((l1id, l2id), 0) for l1id in l1_ids)
        for l2id in l2_ids
    }
    top_l2 = sorted(l2_scores.items(), key=lambda x: -x[1])[:5]

    conflicts = [(a, b) for (a, b), v in ROOF.items() if v == "-"]

    checks = [
        ("空行L1（无L2关联）", str(empty_l1) if empty_l1 else "无", "若存在，可能遗漏L2特性"),
        ("空列L2（无L1关联）", str(empty_l2) if empty_l2 else "无", "若存在，可能冗余或遗漏L1"),
        ("关联密度最高L2（设计关键件）", str(top_l2), "得分越高越关键"),
        ("L2负相关冲突对", str(conflicts) if conflicts else "无", "需设计平衡处理"),
        ("矩阵覆盖率", f"{len(MATRIX)}/{len(L1_ITEMS)*len(L2_ITEMS)} ({len(MATRIX)*100//(len(L1_ITEMS)*len(L2_ITEMS)) if L1_ITEMS and L2_ITEMS else 0}%)", "关联格数/总格数"),
    ]
    for r, (item, result, note) in enumerate(checks, 2):
        ws.cell(r, 1, item).border    = THIN
        ws.cell(r, 1).alignment       = WRAP_TOP
        ws.cell(r, 2, result).border  = THIN
        ws.cell(r, 2).alignment       = WRAP_TOP
        ws.cell(r, 3, note).border    = THIN
        ws.cell(r, 3).alignment       = WRAP_TOP

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35


def main():
    wb = openpyxl.Workbook()
    build_cover(wb)
    build_matrix(wb)
    build_roof(wb)
    build_legend(wb)
    build_review(wb)
    wb.save(CONFIG["输出路径"])
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"矩阵规模: {len(L1_ITEMS)} L1 × {len(L2_ITEMS)} L2")
    print(f"有效关联格: {len(MATRIX)} 个")
    print(f"屋顶相关对: {len(ROOF)} 对")


if __name__ == "__main__":
    main()
