"""
L2零部件特性清单生成器
使用方法：复制到output/gen_l2.py，填写数据区，运行 python3 gen_l2.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG — 修改此区域
# ============================================================
CONFIG = {
    "输出路径": "阶段3_L2零部件特性清单.xlsx",
}

# ============================================================
# L2_DATA — 修改此区域
# 列顺序: [ID, 零件/组件名称, 零件特性, 目标值/要求, 对应L1 ID, 文件来源, 章节, 备注]
# ID格式: PC-01, PC-02...（顺序编号）
# 按零部件分组排列
# ============================================================
L2_DATA = [
    # ["PC-01", "壳体", "材料", "PA66-GF30", "15,16", "PF.90150", "6.2", ""],
]

# ============================================================
# GAP_DATA — 缺口清单（无则留空）
# 列顺序: [缺失信息, 影响L2 ID, 关联L1 ID, 所需文档, 优先级]
# ============================================================
GAP_DATA = [
    # ["壳体材料牌号", "PC-01", "15", "CTS图纸", "高"],
]

# ============================================================
# 以下为格式代码，不修改
# ============================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
GAP_HIGH = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
GAP_MID = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GAP_LOW = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
WRAP_TOP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

L2_HEADERS = ["ID", "零件/组件名称", "零件特性", "目标值/要求", "对应L1 ID", "文件来源", "章节", "备注"]
L2_WIDTHS = [8, 20, 40, 30, 15, 20, 10, 25]


def apply_header(ws, headers, widths):
    ws.append(headers)
    ws.row_dimensions[1].height = 30
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN
    ws.freeze_panes = "A2"


def build_l2(wb):
    ws = wb.active
    ws.title = "零件特性清单"
    apply_header(ws, L2_HEADERS, L2_WIDTHS)
    for row in L2_DATA:
        ws.append(row)
    for r in range(2, len(L2_DATA) + 2):
        for c in range(1, len(L2_HEADERS) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.alignment = WRAP_TOP


def build_gaps(wb):
    ws = wb.create_sheet("缺口与冲突")
    headers = ["缺失信息", "影响L2 ID", "关联L1 ID", "所需文档", "优先级"]
    widths = [40, 15, 15, 30, 10]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    priority_fills = {"高": GAP_HIGH, "中": GAP_MID, "低": GAP_LOW}
    for i, row in enumerate(GAP_DATA, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(i, c, v)
            cell.border = THIN
            cell.alignment = WRAP_TOP
        p = row[4] if len(row) > 4 else ""
        if p in priority_fills:
            for c in range(1, len(headers) + 1):
                ws.cell(i, c).fill = priority_fills[p]


def main():
    wb = openpyxl.Workbook()
    build_l2(wb)
    if GAP_DATA:
        build_gaps(wb)
    wb.save(CONFIG["输出路径"])
    parts = {}
    for row in L2_DATA:
        parts[row[1]] = parts.get(row[1], 0) + 1
    print(f"已保存: {CONFIG['输出路径']}")
    print(f"L2特性: {len(L2_DATA)} 条, {len(parts)} 个零部件")
    if GAP_DATA:
        print(f"缺口: {len(GAP_DATA)} 项")


if __name__ == "__main__":
    main()
